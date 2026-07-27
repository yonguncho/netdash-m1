import logging
import hmac
import os
import sys
import re
import io
import argparse
import tempfile
import ipaddress
import sqlite3
import time
import threading
from functools import wraps
from flask import Flask, jsonify, request, render_template, Response, make_response, redirect
from pathlib import Path

from config import get_config, reset_config
from core import db, collector, correlator, credentials, report_builder, netinfo, connectivity
from core import pcprofile
from core import facility as facility_mod
from core import firewall as firewall_mod
from core.demo import run_demo
from core import flapping as flapping_mod
from core.utils import log_event
from core.collector import _sanitize_error_msg
from core.excel_loader import load_workbook as load_excel_workbook
from core.excel_loader import parse_switch_inventory
from core.excel_loader import parse_server_inventory, parse_firewall_inventory

logging.basicConfig(
    level=logging.INFO,
    format='{"time":"%(asctime)s","level":"%(levelname)s","name":"%(name)s","msg":"%(message)s"}'
)
logger = logging.getLogger(__name__)

_file_log_attached = False


def _safe_stdout():
    """콘솔 출력이 인코딩 오류로 프로세스를 죽이지 않게 한다.

    한국어 Windows 콘솔은 cp949라 '—'(em dash) 같은 문자가 UnicodeEncodeError를
    낸다. 배너·안내문을 찍다가 죽으면 사용자는 원인 안내는커녕 창이 닫히는 것만
    본다(안내를 넣은 목적이 정확히 그 반대다). 인코딩 불가 문자는 '?'로 대체한다.
    """
    for stream in (sys.stdout, sys.stderr):
        try:
            if stream is not None and hasattr(stream, "reconfigure"):
                stream.reconfigure(errors="replace")
        except (OSError, ValueError):
            pass


class _RedactTokenFilter(logging.Filter):
    """접근 로그의 `token=<값>` 을 가린다.

    werkzeug는 요청 라인을 그대로 찍으므로 누가 `?token=...` 으로 접속하면
    netdash.log(공유폴더)에 API 토큰이 평문으로 남는다. 정상 경로는 POST라
    발생하지 않지만, 주소창에 직접 입력하는 경우를 위한 안전망이다.
    """

    _PAT = re.compile(r"(token=)[^\s&\"']+", re.I)

    def filter(self, record):
        try:
            if record.args:
                record.args = tuple(
                    self._PAT.sub(r"\1<redacted>", a) if isinstance(a, str) else a
                    for a in record.args)
            if isinstance(record.msg, str):
                record.msg = self._PAT.sub(r"\1<redacted>", record.msg)
        except Exception:
            pass
        return True


def _install_log_redaction():
    """werkzeug 접근 로그와 루트 로거에 토큰 마스킹 필터를 건다(중복 무해)."""
    f = _RedactTokenFilter()
    for name in ("werkzeug", ""):
        lg = logging.getLogger(name)
        if not any(isinstance(x, _RedactTokenFilter) for x in lg.filters):
            lg.addFilter(f)


def _attach_file_logger(path):
    """루트 로거에 회전 파일 핸들러 부착(중복 방지).

    윈도우 exe는 stdout이 유실돼 '서버 오류'의 실제 내용을 아무도 못 봤다.
    DB 옆 netdash.log에 남겨 재발 오류를 진단 가능하게 한다. 2MB×3 회전(최대 6MB).
    """
    global _file_log_attached
    if _file_log_attached:
        return
    from logging.handlers import RotatingFileHandler
    from pathlib import Path as _Path
    p = _Path(path)
    try:
        p.parent.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    fh = RotatingFileHandler(str(p), maxBytes=2_000_000, backupCount=3,
                             encoding="utf-8")
    fh.setLevel(logging.INFO)
    fh.setFormatter(logging.Formatter(
        '{"time":"%(asctime)s","level":"%(levelname)s","name":"%(name)s","msg":"%(message)s"}'))
    logging.getLogger().addHandler(fh)
    _file_log_attached = True


# ─── 일괄 진단(등록 스위치 전체) — 백그라운드 스레드 + 상태 폴링 ───────────────
_diag_all_lock = threading.Lock()
_diag_all = {"running": False, "total": 0, "done": 0, "corrected": 0,
             "results": [], "error": None}


def _run_diagnose_all(db_path, source_ip):
    """등록된 전 스위치를 동시(스레드풀) 진단 → 벤더 미지정/오지정 자동 교정.

    각 스위치의 저장된 자격증명을 사용(없으면 skip). 결과는 _diag_all에 누적.
    """
    import concurrent.futures as _cf
    try:
        switches = db.get_switches(db_path)
    except Exception as e:
        with _diag_all_lock:
            _diag_all.update(running=False, error=collector._sanitize_error_msg(str(e)))
        return

    # 이 PC 프로필 계정(MAC 키) — 스위치 blob이 다른 PC 것일 때 폴백
    _profile_cred = pcprofile.get_credential(db_path)

    def _one(sw):
        sid = sw.get("id")
        name = sw.get("name")
        blob = db.get_switch_credential(db_path, sid)
        username = password = ""
        if blob:
            dec = credentials.decrypt_credential(blob)
            if dec and "|" in dec:
                username, password = dec.split("|", 1)
        if not (username and password) and _profile_cred:
            username, password = _profile_cred
        if not (username and password):
            return {"id": sid, "name": name, "error": "저장된 계정 없음", "guess": None}
        try:
            res = collector.diagnose_switch(sw, username, password, source_ip=source_ip)
        except Exception as e:
            return {"id": sid, "name": name,
                    "error": collector._sanitize_error_msg(str(e)), "guess": None}
        guess = res.get("guess")
        corrected = None
        if guess and guess != (sw.get("vendor") or "").lower():
            try:
                db.update_switch(db_path, sid, vendor=guess)
                corrected = guess
            except Exception:
                pass
        if guess:
            try:
                diag_text = "\n".join(res.get(k) or "" for k in (
                    "version_head", "banner_head", "sysinfo_head", "inventory_head"))
                osv = res.get("os_version") or collector._parse_os_version(guess, diag_text)
                model = res.get("model") or collector._parse_model(guess, diag_text)
                serial = res.get("serial") or collector._parse_serial(guess, diag_text)
                if osv or model or serial:
                    db.update_switch(db_path, sid, os_version=osv, model=model, serial=serial)
            except Exception:
                pass
        return {"id": sid, "name": name, "guess": guess or "unknown",
                "corrected": corrected, "error": res.get("error") or ""}

    try:
        with _cf.ThreadPoolExecutor(max_workers=8) as ex:
            futs = [ex.submit(_one, sw) for sw in switches]
            for f in _cf.as_completed(futs):
                r = f.result()
                with _diag_all_lock:
                    _diag_all["done"] += 1
                    if r.get("corrected"):
                        _diag_all["corrected"] += 1
                    _diag_all["results"].append(r)
    except Exception as e:
        with _diag_all_lock:
            _diag_all["error"] = collector._sanitize_error_msg(str(e))
    finally:
        with _diag_all_lock:
            _diag_all["running"] = False
    log_event("info", "diagnose_all_done", total=_diag_all["total"],
              corrected=_diag_all["corrected"])


# ─── 방화벽 전체 수집 — 백그라운드 스레드 + 상태 폴링 ───────────────
_fw_all_lock = threading.Lock()
_fw_all = {"running": False, "total": 0, "done": 0, "ok": 0, "message": ""}
# 스위치 일괄 수집 배치 — 진행바·중지용(수집 자체는 collector 워커 큐가 처리)
_sw_bulk = {"ids": [], "total": 0, "started": False}
_sw_bulk_lock = threading.Lock()
# 진단 전용 진행 상태 — 수집과 별개로 둔다(진단이 수집 상태를 덮어쓰지 않게)
_fw_diag = {"running": False, "total": 0, "done": 0, "ok": 0, "message": "",
            "results": []}
_fw_diag_lock = threading.Lock()


def _run_diagnose_all_firewalls(db_path, source_ip, ids=None):
    """방화벽 도달성·인증만 확인한다. 인터페이스/ARP를 저장하지 않고 status도 안 바꾼다.

    예전엔 '전체 진단' 버튼이 collect-all을 호출해 실제 수집을 했다 —
    안내문(도달성·인증 확인)과 달리 데이터를 덮어쓰고 상태를 바꿨다.
    """
    import json as _json
    results = []
    try:
        try:
            firewalls = db.list_firewalls(db_path)
        except Exception as e:
            with _fw_diag_lock:
                _fw_diag.update(message=collector._sanitize_error_msg(str(e)))
            return
        if ids:                      # 서버실 화면처럼 일부만 진단할 때
            _want = set(ids)
            firewalls = [f for f in firewalls if f.get("id") in _want]
        with _fw_diag_lock:
            _fw_diag.update(total=len(firewalls), done=0, ok=0, message="진단 중")
        for fw in firewalls:
            fid = fw.get("id")
            vendor = (fw.get("vendor") or "").lower()
            host = fw.get("host")
            mgmt = fw.get("port") or (443 if vendor == "fortigate" else 22)
            token = username = password = ""
            try:
                blob = db.get_firewall_credential(db_path, fid)
                if blob:
                    dec = credentials.decrypt_text(blob)
                    if dec:
                        saved = _json.loads(dec)
                        token = saved.get("token", "")
                        username = saved.get("username", "")
                        password = saved.get("password", "")
            except Exception:
                pass
            item = {"id": fid, "name": fw.get("name"), "host": host, "vendor": vendor,
                    "mgmt_port": mgmt}
            try:
                item["tcp_mgmt"] = connectivity.test_tcp(host, mgmt, 3, source_ip)
                item["tcp_ssh"] = connectivity.test_tcp(host, 22, 3, source_ip)
                item["has_token"] = bool(token)
                item["has_login"] = bool(username and password)
                res = connectivity.test_firewall(vendor, host, mgmt, token=token,
                                                 username=username, password=password,
                                                 verify_ssl=False, source_ip=source_ip)
                item["auth_ok"] = bool(res.get("ok")) and res.get("stage") == "auth"
                item["detail"] = res.get("detail") or ""
                if item["auth_ok"] or item["tcp_mgmt"]:
                    with _fw_diag_lock:
                        _fw_diag["ok"] += 1
            except Exception as e:
                item["detail"] = collector._sanitize_error_msg(str(e))
            results.append(item)
            with _fw_diag_lock:
                _fw_diag.update(done=len(results), results=list(results),
                                message="진단 중 (%d/%d)" % (len(results), len(firewalls)))
            token = username = password = None
    finally:
        with _fw_diag_lock:
            _fw_diag.update(running=False, results=list(results),
                            message="진단 완료(%d대)" % len(results))
        log_event("info", "firewalls_diagnose_all_done", total=len(results))


def _run_collect_all_firewalls(db_path, source_ip, ids=None, sess_cred=None,
                               common_token=""):
    """방화벽을 순차 수집. ids를 주면 그 방화벽만. 진행은 _fw_all.

    계정 우선순위: 화면에서 입력한 공통 계정/토큰 > 각 방화벽 저장 계정 > 방화벽 세션 계정.
    sess_cred: (username, password) — 화면 입력이 없을 때 쓰는 '방화벽' 세션 계정.
    common_token: FortiGate API 토큰(화면 입력). 백그라운드 스레드에는 request가
    없으므로 호출 시점에 꺼내어 넘긴다.
    """
    import json as _json
    # 이 함수는 백그라운드 스레드에서 돈다. 예상 못한 예외로 스레드가 죽으면
    # running 플래그가 True로 남아 이후 모든 수집이 영구 409가 되므로,
    # 바깥 try/finally로 해제를 보장한다(현장 db_error에서 실제로 재현됨).
    ok = 0
    firewalls = []
    try:
        try:
            firewalls = db.list_firewalls(db_path)
            if ids:
                _idset = set(int(x) for x in ids)
                firewalls = [f for f in firewalls if f.get("id") in _idset]
        except Exception as e:
            with _fw_all_lock:
                _fw_all.update(message=collector._sanitize_error_msg(str(e)))
            return
        for i, fw in enumerate(firewalls):
            with _fw_all_lock:
                if _fw_all.get("stop"):   # 사용자 중지 → 남은 방화벽 건너뜀
                    _fw_all.update(message="중지됨(성공 %d/%d)" % (ok, len(firewalls)))
                    return
            fid = fw.get("id")
            token = username = password = ""
            # ① 화면에서 입력한 공통 계정/토큰이 최우선(방화벽 계정은 장비마다 다르므로
            #    사용자가 이 화면에서 직접 준 것이 가장 정확하다)
            if common_token:
                token = common_token
            if sess_cred and sess_cred[0] and sess_cred[1]:
                username, password = sess_cred
            # ② 화면 입력이 없으면 각 방화벽에 저장된 계정
            if not (token or (username and password)):
                try:
                    blob = db.get_firewall_credential(db_path, fid)
                    if blob:
                        dec = credentials.decrypt_text(blob)
                        if dec:
                            saved = _json.loads(dec)
                            token = saved.get("token", "")
                            username = saved.get("username", "")
                            password = saved.get("password", "")
                except Exception:
                    pass
            try:
                # 상태 갱신도 try 안에 둔다(DB 잠금 시 예외가 스레드를 죽이던 지점).
                db.set_firewall_status(db_path, fid, "collecting")
                result = firewall_mod.collect_firewall(
                    fw["vendor"], fw["host"], fw.get("port"),
                    token=token, username=username, password=password,
                    verify_ssl=False, source_ip=source_ip)
                db.save_firewall_interfaces(db_path, fid, result["interfaces"])
                db.save_firewall_arp(db_path, fid, result["arp"])
                if result.get("ha"):
                    try:
                        db.set_firewall_ha_info(db_path, fid, _json.dumps(result["ha"]))
                    except Exception:
                        pass
                db.set_firewall_status(db_path, fid, "done")
                ok += 1
            except Exception as e:
                try:
                    db.set_firewall_status(db_path, fid, "failed")
                except Exception:
                    pass       # 상태 저장 실패가 남은 방화벽 수집을 막지 않도록
                log_event("warning", "firewall_collect_all_item_error",
                          firewall_id=fid, error=collector._sanitize_error_msg(str(e)))
            finally:
                token = username = password = None
                with _fw_all_lock:
                    _fw_all.update(done=i + 1, ok=ok,
                                   message="방화벽 수집 중 (%d/%d)" % (i + 1, len(firewalls)))
        with _fw_all_lock:
            _fw_all.update(message="완료(성공 %d/%d)" % (ok, len(firewalls)))
        log_event("info", "firewall_collect_all_done", total=len(firewalls), ok=ok)
    finally:
        with _fw_all_lock:
            _fw_all.update(running=False, ok=ok)


def validate_credential(value, max_length=256):
    """CRITICAL FIX (CWE-20): Validate credential string length and printable ASCII only.

    Prevents DoS (oversized input), injection attacks (control chars).
    Allows printable ASCII characters except space (to prevent accidental whitespace in passwords).
    """
    if value is None:
        return None
    if not isinstance(value, str):
        raise ValueError("credentials must be string")
    if len(value) == 0:
        raise ValueError("credentials cannot be empty")
    if len(value) > max_length:
        raise ValueError(f"credentials max length {max_length}")
    # CWE-20: Accept printable ASCII only (ord 33-126); exclude space (ord 32) to prevent whitespace-only credentials
    if not all(33 <= ord(c) <= 126 for c in value):
        raise ValueError("credentials must contain only printable ASCII characters (no spaces, no control chars)")
    return value


def validate_ipv4(ip_str, allowed_ip_ranges=None):
    """HARDENING (CWE-918 SSRF): Validate IPv4 address and reject reserved/dangerous ranges.

    Args:
        ip_str: IP address string
        allowed_ip_ranges: Optional list of allowed CIDR ranges (e.g., ["10.0.0.0/8", "172.16.0.0/12"])

    Raises:
        ValueError: If IP is invalid or in a reserved/dangerous range
    """
    if not isinstance(ip_str, str) or not ip_str.strip():
        raise ValueError("IP address is required and must be a string")

    try:
        ip_obj = ipaddress.IPv4Address(ip_str.strip())
    except ipaddress.AddressValueError:
        raise ValueError(f"Invalid IPv4 address: {ip_str}")

    # Reject reserved/dangerous ranges (loopback, multicast, link-local, etc.)
    if ip_obj.is_loopback:
        raise ValueError(f"Loopback address not allowed: {ip_str}")
    if ip_obj.is_multicast:
        raise ValueError(f"Multicast address not allowed: {ip_str}")
    if ip_obj.is_link_local:
        raise ValueError(f"Link-local address not allowed: {ip_str}")
    if ip_obj.is_reserved:
        raise ValueError(f"Reserved address not allowed: {ip_str}")

    # Check allowed_ip_ranges if provided (whitelist mode)
    if allowed_ip_ranges:
        allowed = False
        for cidr_str in allowed_ip_ranges:
            try:
                network = ipaddress.IPv4Network(cidr_str, strict=False)
                if ip_obj in network:
                    allowed = True
                    break
            except ipaddress.AddressValueError:
                log_event("warning", "invalid_cidr_range", cidr=cidr_str)
        if not allowed:
            raise ValueError(f"IP address not in allowed ranges: {ip_str}")

    return str(ip_obj)


# Rate limiting: IP/token-based request tracking (simple dict-based, no external dependency)
_rate_limit_tracker = {}
_rate_limit_lock = __import__("threading").Lock()

_SERVER_ETH_IP = None   # 서버 이더넷 IP 캐시(루프백 접속 시 대체 표기용)

# 방화벽 동시 수집 가드(스위치의 collector._collecting_switches에 대응).
# 방화벽 collect는 요청 스레드에서 동기 실행이라 같은 fid 동시 POST 시 상태·
# 인터페이스/ARP가 서로 다른 수집본으로 섞일 수 있어 in-progress를 막는다.
_collecting_firewalls = set()
_collecting_fw_lock = threading.Lock()

# 서버 os_type 허용값 — Windows가 아닌 OS는 모두 범용 UNIX 폴백 명령으로 수집한다.
# (linux/windows만 허용하던 제약 때문에 AIX·Solaris·HP-UX·ESXi 서버 등록이 auto로
#  덮여 수집 분기가 어긋나던 문제 해소)
_SERVER_OS_TYPES = ("auto", "linux", "windows", "aix", "solaris", "hpux",
                    "esxi", "bsd", "macos", "unix")


def _server_eth_ip():
    global _SERVER_ETH_IP
    if _SERVER_ETH_IP is None:
        try:
            _SERVER_ETH_IP = netinfo._primary_ip() or ""
        except Exception:
            _SERVER_ETH_IP = ""
    return _SERVER_ETH_IP


def _trusted_proxies():
    """XFF/X-Real-IP를 신뢰할 프록시 주소 목록(config `app.trusted_proxies`).

    비어 있으면(기본) 어떤 요청의 포워딩 헤더도 신뢰하지 않는다.
    """
    try:
        from config import get_config as _gc
        app_cfg = _gc().app
        vals = app_cfg.get("trusted_proxies") if isinstance(app_cfg, dict) else None
        return {str(v).strip() for v in (vals or []) if str(v).strip()}
    except Exception:
        return set()


def _client_ip():
    """실사용자 IP 판별 — 프록시/포워딩 뒤에서도 이더넷 IP가 기록되도록.

    우선순위: (신뢰 프록시 경유 시) X-Forwarded-For → X-Real-IP → remote_addr.
    remote_addr가 루프백(127.0.0.1/localhost/::1 = 같은 PC 접속)이면 접근 로그에
    127.0.0.1 대신 서버 PC의 이더넷 IP로 표기(사용자 구분 목적).

    보안: XFF/X-Real-IP는 클라이언트가 위조할 수 있으므로 **설정에 명시한 프록시**
    (`app.trusted_proxies`)에서 온 요청일 때만 채택한다. 기본값은 빈 목록 = 절대
    신뢰하지 않음.

    예전엔 remote_addr가 사설(RFC1918)이면 신뢰했는데, 이 제품은 폐쇄망 전용이라
    **모든 정상 클라이언트가 사설 IP**다. 즉 그 조건은 공격자 집단과 정확히 일치했고,
    아무 사용자나 `X-Forwarded-For`를 붙여 감사 로그의 행위자를 위조하고
    (감사 추적의 유일한 근거) 레이트리밋 키까지 바꿔 제한을 우회할 수 있었다.
    """
    ra = request.remote_addr or "unknown"
    _trusted = ra in _trusted_proxies()
    if _trusted:
        xff = (request.headers.get("X-Forwarded-For") or "").split(",")[0].strip()
        if xff:
            return xff[:64]
        xri = (request.headers.get("X-Real-IP") or "").strip()
        if xri:
            return xri[:64]
    if ra in ("127.0.0.1", "localhost", "::1"):
        eth = _server_eth_ip()
        if eth:
            return eth
    return ra


def rate_limit(endpoint, max_requests=5, window_seconds=60):
    """HARDENING (CWE-400): Simple rate limiter decorator.

    Limits requests per IP/token combination.
    """
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            # Get identifier: IP + token (or IP alone)
            ip = _client_ip()
            token = request.headers.get("X-API-Token", "")
            identifier = f"{ip}:{token}" if token else ip

            with _rate_limit_lock:
                now = time.time()
                key = f"{endpoint}:{identifier}"

                # Clean old entries (older than window)
                if key in _rate_limit_tracker:
                    timestamps = [t for t in _rate_limit_tracker[key] if now - t < window_seconds]
                    if len(timestamps) >= max_requests:
                        log_event("warning", "rate_limit_exceeded", endpoint=endpoint, identifier=identifier)
                        return jsonify({"error": "Rate limit exceeded"}), 429
                    _rate_limit_tracker[key] = timestamps + [now]
                else:
                    _rate_limit_tracker[key] = [now]

                # 메모리 누수 방지(CWE-770): 주기적으로 만료된 빈 키 정리.
                # ip/token은 클라이언트 제어라 고유 키가 무한 누적될 수 있어 gc 필요.
                _rl = _rate_limit_tracker
                if len(_rl) > 512:
                    for k in [k for k, v in _rl.items()
                              if not any(now - t < window_seconds for t in v)]:
                        del _rl[k]

            return f(*args, **kwargs)
        return wrapper
    return decorator


def _start_primary_services(config, db_path):
    """주 서버 전용 초기화 — 시작 시 DB 정비 쓰기 + 쓰기 백그라운드 스레드.

    정상 기동(주 서버)과 읽기 전용 → 주 서버 자동 승격 양쪽에서 재사용.
    """
    # 런타임 로그를 파일로도 보존(서버 오류 진단용) — DB 옆 netdash.log
    _install_log_redaction()   # 접근 로그의 token= 값 마스킹(파일 핸들러보다 먼저)
    try:
        from pathlib import Path as _Path
        _attach_file_logger(_Path(str(db_path)).parent / "netdash.log")
    except Exception as e:
        log_event("warning", "file_logger_failed", error=str(e))
    db.init_schema(db_path)
    db.validate_schema(db_path)
    # 이전 실행 중단으로 박제된 '수집중' 상태 복구(재시작 후 실제 수집은 없음)
    try:
        db.reset_stale_collecting(db_path)
    except Exception as e:
        log_event("warning", "stale_reset_failed", error=str(e))
    # 벤더 별칭(cisco/extreme...) → 표준 값(cisco_ios/extreme_exos...) 일괄 정규화
    try:
        db.normalize_vendor_values(db_path)
    except Exception as e:
        log_event("warning", "vendor_normalize_failed", error=str(e))
    # 버전이 빈 스위치를 기존 config 백업의 version 줄로 백필(재수집 불필요)
    try:
        db.backfill_versions_from_config(db_path)
    except Exception as e:
        log_event("warning", "version_backfill_failed", error=str(e))

    collector.init_collector()
    # M14: 하루 N회 자동 수집 스케줄러 시작(설정으로 on/off)
    try:
        from core import scheduler
        scheduler.start_scheduler(db_path)
    except Exception as e:
        log_event("warning", "scheduler_start_failed", error=str(e))
    # 스위치 도달성 감시(TCP-22, 부하 없는 1분 내 끊김 감지 — 설정으로 on/off)
    if not config.app.get("demo_mode"):
        try:
            from core import reachability
            reachability.start_monitor(db_path)
        except Exception as e:
            log_event("warning", "reachability_start_failed", error=str(e))
        # 관제/설비가 쓰는 'MAC 최근위치' 맵을 기동 직후 미리 구축(첫 진입 지연 제거)
        try:
            db.warm_mac_last_cache(db_path)
        except Exception as e:
            log_event("warning", "mac_cache_warm_start_failed", error=str(e))
        # 알람 이메일 발송 스레드(설정으로 on/off, 60초 묶음 발송)
        try:
            from core import notifier
            notifier.start_notifier(db_path)
        except Exception as e:
            log_event("warning", "notifier_start_failed", error=str(e))


def _promote_to_primary(app, config, db_path):
    """읽기 전용 인스턴스를 주 서버로 승격(인스턴스 락 획득 후 호출).

    새 DB 연결부터 쓰기가 허용되고, 주 서버 백그라운드가 기동된다.
    """
    db.READONLY = False
    _start_primary_services(config, db_path)
    app.config["IS_READONLY"] = False
    app.config["READONLY_PRIMARY"] = None
    log_event("info", "promoted_to_primary")
    print("=" * 56)
    print("  주 서버가 종료되어 이 프로그램이 [주 서버]로 전환되었습니다.")
    print("  수집·수정 기능이 활성화되었습니다.")
    print("=" * 56, flush=True)


PROMOTE_POLL_SEC = 5  # 주 서버 종료 감지 주기


def _watch_and_promote(app, config, db_path, url):
    """읽기 전용 모드 감시 스레드: 주 서버 종료 시 자동 승격.

    인스턴스 락 획득을 주기적으로 시도 — 주 서버가 종료(크래시 포함)하면
    OS가 락을 해제하므로 획득에 성공하고, 그 순간 승격한다.
    여러 읽기 전용 인스턴스가 있어도 락은 하나만 얻으므로 승격도 하나만 된다.
    """
    from core import instance_lock
    from core.config_loader import get_data_dir
    while app.config.get("IS_READONLY"):
        time.sleep(PROMOTE_POLL_SEC)
        try:
            # allow_unlocked=False: 락 파일을 확인하지 못했으면 승격하지 않는다.
            # (공유폴더 순단 때 '획득 성공'으로 오인해 주 서버가 둘이 되면
            #  같은 SQLite에 두 PC가 동시 쓰기를 한다)
            # 잠금 위치는 기동 경로와 동일하게 DB 폴더 기준이어야 한다
            # (다르면 승격 판정이 다른 파일을 보게 된다)
            try:
                _ld = config.get_db_path().parent
            except Exception:
                _ld = get_data_dir()
            acquired, _ = instance_lock.acquire(_ld, url, allow_unlocked=False)
        except Exception as e:
            log_event("warning", "promote_watch_error",
                      error=collector._sanitize_error_msg(str(e)))
            continue
        if not acquired:
            continue
        try:
            _promote_to_primary(app, config, db_path)
            return
        except Exception as e:
            # 승격 도중 예외 — 여기서 죽으면 락은 이 프로세스가 쥔 채 읽기 전용으로
            # 남아, 어떤 PC도 주 서버가 될 수 없다(배포 전체 정지). 락을 놓고 재시도한다.
            log_event("error", "promote_failed_releasing_lock",
                      error=collector._sanitize_error_msg(str(e)))
            try:
                instance_lock.release()
            except Exception:
                pass
            app.config["IS_READONLY"] = True
            db.READONLY = True


def create_app(demo_mode=None, readonly_info=None, promote_watch=False):
    """Factory function to create and configure Flask app.

    readonly_info: 다른 PC의 주 서버가 DB(인스턴스 락)를 소유 중일 때 그 정보
    ({hostname, url, ...}). 지정되면 읽기 전용 모드로 기동 — 조회는 전부 허용,
    쓰기(수집/수정/삭제) 요청은 423 + 안내 메시지. DB 쓰기 스레드(수집 워커·
    스케줄러·도달성 감시·알림)와 시작 시 DB 정비 쓰기도 모두 건너뛴다.
    """
    readonly = readonly_info is not None
    app = Flask(__name__,
                template_folder=str(Path(__file__).parent / "web" / "templates"),
                static_folder=str(Path(__file__).parent / "web" / "static"))

    # M4: Set 16MB max upload size (CWE-399 fix: prevent DoS via oversized uploads)
    app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16MB

    # Reset config singleton to allow fresh load in tests
    reset_config()
    config = get_config(demo_mode=demo_mode)

    db_path = config.get_db_path()

    app.config["IS_READONLY"] = readonly
    app.config["READONLY_PRIMARY"] = (
        ((readonly_info or {}).get("hostname") or "다른 PC") if readonly else None)

    if readonly:
        # 프로세스 전역 안전벨트: 모든 DB 연결이 query_only로 열림
        db.READONLY = True
        log_event("info", "app_readonly_mode",
                  primary_host=app.config["READONLY_PRIMARY"])

        # NOTE: 등록은 validate_api_token 정의 뒤에서 한다(아래 `app.before_request(...)`).
        # before_request는 등록 순서대로 실행되므로 이걸 먼저 등록하면 토큰 없는 요청이
        # 401 대신 423 + 주 서버 호스트명을 받아 미인증 원격에 호스트명이 새어나간다.
        def _readonly_gate():
            # 쓰기 메서드 전체 차단 — 조회(GET)와 정적 리소스는 그대로 허용.
            # (승격되면 IS_READONLY=False가 되어 게이트가 열린다)
            if not app.config.get("IS_READONLY"):
                return None
            # 로그인(토큰 제출)은 DB를 쓰지 않는다 — 읽기 전용 모드에서도 허용해야
            # 원격 사용자가 조회조차 못 하는 상태가 되지 않는다.
            if request.path == "/session":
                return None
            if request.method in ("POST", "PUT", "DELETE", "PATCH"):
                return jsonify({
                    "error": "다른 사용자(주 서버: %s)가 DB를 사용 중입니다. "
                             "지금은 조회만 가능합니다. 수집·수정은 주 서버에서 "
                             "하세요. (주 서버 종료 시 이 프로그램이 자동으로 "
                             "주 서버가 됩니다)"
                             % app.config.get("READONLY_PRIMARY"),
                    "readonly": True,
                }), 423

        _ro_gate = _readonly_gate
        if promote_watch:
            # 주 서버 종료 감지 → 자동 승격 (프로덕션 exe 전용 — 테스트는 미기동)
            _pw_host = config.app.get("host", "127.0.0.1")
            _pw_port = config.app.get("port", 8082)
            _pw_open = "127.0.0.1" if _pw_host in ("0.0.0.0", "::") else _pw_host
            threading.Thread(
                target=_watch_and_promote,
                args=(app, config, db_path, f"http://{_pw_open}:{_pw_port}"),
                daemon=True, name="promote-watch").start()
    else:
        _ro_gate = None
        db.READONLY = False
        _start_primary_services(config, db_path)

    # Load demo data in demo mode
    if config.app.get("demo_mode") and not readonly:
        run_demo(config)

    # API Token validation for production mode (CWE-306 fix: enforce authentication on all API routes)
    @app.before_request
    def validate_api_token():
        # Skip token validation only for page shells ("/", "/wall"), the token
        # submission form ("/session"), and health checks. 페이지 셸은 자체적으로
        # 토큰을 확인한다(_render_page) — /session은 그 확인 자체를 수행한다.
        if request.path in ("/", "/health", "/wall", "/session"):
            return
        # 데모 모드는 인증을 건너뛴다 — 단, **토큰이 설정돼 있으면 건너뛰지 않는다.**
        # 예전엔 무조건 통과라, 운영 폴더에서 환경변수 하나(DEMO_MODE=true)로 띄우면
        # 같은 config를 읽어 **실제 운영 DB가 인증 없이 전부 열렸다**
        # (자산 목록·감사 이력 조회, 웹 SSH 터미널 포함).
        if config.app.get("demo_mode") and not config.api_token:
            return
        # Loopback-only bind + request originating from localhost is exempt:
        # a closed-network single-host tool is reachable only by the same machine's
        # user, so the local UI works without a token. The token defends remote
        # access only when bound to an externally reachable address (0.0.0.0 등).
        bind_host = config.app.get("host", "127.0.0.1")
        if bind_host in ("127.0.0.1", "localhost", "::1") and request.remote_addr in ("127.0.0.1", "::1"):
            return
        # 웹 SSH 터미널(WebSocket)은 핸들러 내부에서 자체 토큰 검증(쿼리 파라미터)
        if request.path.startswith("/ws/"):
            return
        # Enforce API authentication in production mode (all /api/* routes)
        if request.path.startswith("/api/"):
            token = request.headers.get("X-API-Token")
            expected_token = config.api_token
            # CWE-306 fix: Reject if token is missing or invalid; never accept empty token
            if not token:
                log_event("warning", "api_missing_token", path=request.path)
                return jsonify({"error": "unauthorized"}), 401
            if not expected_token:
                # Production mode requires api_token to be set in config
                log_event("error", "api_token_not_configured", path=request.path)
                return jsonify({"error": "server configuration error"}), 500
            # bytes 비교: 비ASCII 토큰 헤더 쌍에서 compare_digest가 TypeError를
            # 던져 500(미처리 예외)이 되던 것 방지 → 정상적으로 401 처리.
            if not hmac.compare_digest(token.encode("utf-8", "replace"),
                                       expected_token.encode("utf-8", "replace")):
                log_event("warning", "api_invalid_token", path=request.path)
                return jsonify({"error": "unauthorized"}), 401

    # 읽기 전용 게이트는 인증 뒤에 — 미인증 요청은 423(주 서버 호스트명 포함)이 아니라
    # 401을 받아야 한다. (정의는 위 `if readonly:` 블록)
    if _ro_gate is not None:
        app.before_request(_ro_gate)

    # ── 접근(감사) 로그: 어느 PC가 언제 무엇을 했는지 자동 기록 ──
    # 변경 행위(POST/PUT/DELETE)와 다운로드성 GET만 기록(조회 폴링은 제외 — 소음 방지).
    def _audit_label(m, p):
        if m == "POST":
            if p == "/api/switches/bulk-collect":
                return "일괄 수집 실행"
            if p == "/api/switches/diagnose-all":
                return "전체 진단 실행"
            if p == "/api/switches/bulk-zone":
                return "존 일괄 지정"
            if p == "/api/credentials/delete":
                return "저장 계정 삭제"
            if p == "/api/topology/diagram":
                return "토폴로지 구성도 저장"
            if p == "/api/servers":
                return "서버 등록"
            if p == "/api/servers/collect-all":
                return "서버 일괄 수집"
            if p == "/api/servers/import":
                return "서버 일괄등록"
            if p.startswith("/api/servers/") and p.endswith("/collect"):
                return "서버 수집 실행"
            if p.startswith("/api/switches/") and p.endswith("/collect"):
                return "스위치 수집 실행"
            if p.startswith("/api/firewalls/") and p.endswith("/collect"):
                return "방화벽 수집 실행"
            if p == "/api/switches/bulk-delete":
                return "스위치 일괄 삭제"
            if p == "/api/switches/manual":
                return "스위치 등록"
            if p == "/api/firewalls":
                return "방화벽 등록"
            if p == "/api/upload":
                return "엑셀 업로드"
            if p == "/api/switches/import-inventory":  # 실제 라우트(app.py:586)
                return "장비 일괄등록"
            if p == "/api/facility/collect":
                return "설비 대역 수집"
            if p == "/api/facility/rematch":
                return "설비 재매칭"
            if p.startswith("/api/settings/"):
                return "설정 변경"
            if p == "/api/alerts/ack":
                return "알람 확인"
            # 파괴적 작업·계정 보관은 반드시 남긴다(대역 수집 결과를 통째로 지움)
            if p == "/api/facility/delete-subnet":
                return "설비 대역 삭제"
            if p == "/api/switches/bulk-set-type":
                return "스위치 구분 일괄 변경"
            if p == "/api/session/credential":
                return "세션 수집 계정 보관"
            if p == "/api/session/credential/lock":
                return "세션 수집 계정 잠금"
            if p == "/api/servers/diagnose-all":
                return "서버 전체 진단"
            if p == "/api/firewalls/diagnose-all":
                return "방화벽 전체 진단"
            if p == "/api/report/pptx":
                return "보고서(PPTX) 생성"
        elif m == "PUT":
            if p.startswith("/api/switches/"):
                return "스위치 수정"
            if p.startswith("/api/firewalls/"):
                return "방화벽 수정"
            if p.startswith("/api/servers/"):
                return "서버 수정"
        elif m == "DELETE":
            if p.startswith("/api/switches/"):
                return "스위치 삭제"
            if p.startswith("/api/firewalls/"):
                return "방화벽 삭제"
            if p.startswith("/api/servers/"):
                return "서버 삭제"
        elif m == "GET":
            if p == "/api/report":
                return "보고서 다운로드"
            if p == "/api/facility/export":
                return "설비 목록 추출"
            # 화면별 ⬇ 다운로드 — 전체 자산 목록이 나가는 경로라 반드시 기록한다
            # (기존엔 report/facility만 남기고 이것만 빠져 정책이 어긋났다)
            if p.startswith("/api/export/"):
                return "목록 다운로드(%s)" % p.rsplit("/", 1)[-1]
            if p == "/api/serverroom/export":
                return "랙배치 다운로드"
            if p == "/api/configs/export-all":
                return "설정 백업 일괄 다운로드"
            if p.startswith("/api/configs/") and "diff" not in p:
                return "설정 백업 열람"
        return None

    @app.after_request
    def audit_request(response):
        try:
            if response.status_code < 400:
                label = _audit_label(request.method, request.path)
                if label:
                    # 실사용자 IP(X-Forwarded-For 우선) — 127.0.0.1만 찍히던 문제 해결
                    db.save_audit(db_path, _client_ip(), label,
                                  target=request.path, method=request.method,
                                  path=request.path)
        except Exception:
            pass
        return response

    # Security headers for all responses
    @app.after_request
    def set_security_headers(response):
        # script-src는 'self'만(인라인 스크립트/onclick 차단 → 이벤트 위임 사용).
        # style-src는 인라인 style 속성 허용(레이아웃 정상화). 스타일은 스크립트 실행이
        # 아니므로 XSS 위험이 낮다.
        response.headers["Content-Security-Policy"] = (
            "default-src 'self'; script-src 'self'; style-src 'self' 'unsafe-inline'"
        )
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["Referrer-Policy"] = "no-referrer"
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        return response

    # ── 페이지 셸 토큰 (원격 접속용) ──────────────────────────────
    # `/api/*`는 X-API-Token을 요구하는데 페이지 셸(/, /wall)은 면제라, 0.0.0.0
    # 바인드로 원격 접속하면 화면은 뜨고 그 안의 모든 fetch가 401이 됐다
    # (window._API_TOKEN을 넣어주는 코드가 어디에도 없었다).
    # 이제 셸도 토큰을 요구하고(POST 제출 → 쿠키), 통과한 페이지에만
    # window._API_TOKEN을 심어 준다.
    #
    # 토큰은 반드시 POST로 받는다. `?token=` 쿼리로 받으면 werkzeug 접근 로그
    # ("GET /?token=... HTTP/1.1")에 **토큰이 평문으로 남는다** — netdash.log는
    # 공유폴더(DB 옆)에 쌓이므로 공유 접근 권한자 전원에게 토큰이 새어 토큰의
    # 존재 의의가 사라진다. 쿼리 방식은 하위호환으로 남기되 즉시 쿠키로 옮기고
    # 리다이렉트하며, 로그 필터가 쿼리의 토큰 값을 가린다(_redact_token_in_logs).
    _TOKEN_COOKIE = "netdash_token"

    def _is_local_request():
        return request.remote_addr in ("127.0.0.1", "::1")

    def _api_needs_token():
        """validate_api_token과 같은 판정 — 이 요청의 /api 호출에 토큰이 필요한가."""
        if config.app.get("demo_mode") and not config.api_token:
            return False
        bind_host = config.app.get("host", "127.0.0.1")
        return not (bind_host in ("127.0.0.1", "localhost", "::1") and _is_local_request())

    def _render_page(template, **kw):
        if not _api_needs_token():
            # 로컬 전용 배포 — 종전 그대로 토큰 없이 동작
            return make_response(render_template(template, api_token="", **kw))
        expected = config.api_token
        if not expected:
            log_event("error", "api_token_not_configured", path=request.path)
            return jsonify({"error": "server configuration error"}), 500
        # 토큰이 필요한 배포(0.0.0.0 바인드)에서는 **로컬 요청도 예외로 두지 않는다.**
        # 예전엔 "같은 PC 사용자는 config.yaml을 읽을 수 있다"는 이유로 면제했는데,
        # 그러면 `/api/state`는 401로 막는 바로 그 요청자에게 페이지가 토큰을
        # 넘겨준다(권한 없는 로컬 프로세스·리버스 프록시 뒤 전원). API 게이트
        # (validate_api_token)와 판정을 일치시킨다.
        q_token = request.args.get("token") or ""
        given = q_token or request.cookies.get(_TOKEN_COOKIE) or ""
        if not given or not hmac.compare_digest(
                given.encode("utf-8", "replace"), expected.encode("utf-8", "replace")):
            log_event("warning", "page_missing_token", path=request.path)
            return render_template("token_required.html",
                                   next_path=_safe_next(request.path)), 401
        if q_token:
            # 쿼리로 들어왔으면 쿠키로 옮기고 URL에서 즉시 지운다(브라우저
            # 히스토리·리퍼러·접근 로그에 토큰이 남지 않도록).
            resp = redirect(_safe_next(request.path))
            resp.set_cookie(_TOKEN_COOKIE, q_token, httponly=True,
                            samesite="Strict", path="/")
            return resp
        resp = make_response(render_template(template, api_token=given, **kw))
        resp.set_cookie(_TOKEN_COOKIE, given, httponly=True,
                        samesite="Strict", path="/")
        return resp

    def _safe_next(path):
        """리다이렉트 대상은 우리가 아는 페이지로만 — 오픈 리다이렉트 차단."""
        return path if path in ("/", "/wall") else "/"

    def _ws_page_same_origin():
        """폼 POST의 출처 확인 — 다른 사이트가 이 엔드포인트를 대신 호출하지 못하게.
        Origin이 없는 요청(curl 등)은 토큰 검증에 맡긴다."""
        origin = request.headers.get("Origin")
        if not origin:
            return True
        host = request.host or ""
        return origin in ("http://" + host, "https://" + host)

    @app.route("/session", methods=["POST"])
    def submit_token():
        """토큰 제출(POST) — 쿼리스트링을 쓰지 않아 접근 로그에 남지 않는다."""
        expected = config.api_token
        nxt = _safe_next(request.form.get("next") or "/")
        if not _api_needs_token():
            return redirect(nxt)
        if not _ws_page_same_origin():
            log_event("warning", "page_token_bad_origin")
            return jsonify({"error": "forbidden"}), 403
        given = request.form.get("token") or ""
        if not expected or not given or not hmac.compare_digest(
                given.encode("utf-8", "replace"), expected.encode("utf-8", "replace")):
            log_event("warning", "page_token_rejected", path=nxt)
            return render_template("token_required.html", next_path=nxt,
                                   error="토큰이 올바르지 않습니다."), 401
        resp = redirect(nxt)
        resp.set_cookie(_TOKEN_COOKIE, given, httponly=True,
                        samesite="Strict", path="/")
        log_event("info", "page_token_accepted", path=nxt)
        return resp

    @app.route("/", methods=["GET"])
    def index():
        demo_mode = config.app.get("demo_mode", False)
        return _render_page("index.html", demo_mode=demo_mode)

    @app.route("/api/state", methods=["GET"])
    def get_state():
        log_event("info", "api_state")
        try:
            switches = db.get_switches(db_path)
            snapshots = db.get_snapshots(db_path)

            # hostname → TPS 물리 위치 라벨 + 랙 그룹핑 키 주입(포맷 일치 시)
            from core import tps_location, serverroom, topology as _topo
            # 구분 자동 분류(L2/L3/L4) — 최신 running-config + 벤더로 판정(수동 지정 불필요)
            _cfgs = db.get_latest_configs(db_path)
            for sw in switches:
                _kind = _topo.classify_switch_kind(_cfgs.get(sw["id"]), sw.get("vendor"))
                if _kind:
                    sw["kind_auto"] = _kind
            for sw in switches:
                # 존 자동 분류(hostname 명명규칙) — 명시 zone 없을 때 표시용
                if not sw.get("zone"):
                    _za = _topo.infer_zone(sw.get("name"), sw.get("hostname"))
                    if _za:
                        sw["zone_auto"] = _za
                info = tps_location.parse(sw.get("hostname"))
                if info:
                    sw["tps_location"] = info["label"]
                    sw["tps_group"] = "%d공장 · %s(%s) · %d층" % (
                        info["phase"], info["building_name"], info["building_code"], info["floor"])
                    sw["tps_num"] = "TPS" + info["tps"]
                # location "A09U27" → 서버실 랙/유닛 (서버실 현황 탭용)
                room = serverroom.parse_rack(sw.get("location"))
                if room:
                    sw["room_rack"] = room["rack"]
                    sw["room_unit"] = room["unit"]
                    sw["room_label"] = room["label"]
                    sw["room_height"] = room.get("height", 1)

            # 도달성 감시 결과 주입(True=도달, False=불가, 없으면 미확인)
            try:
                from core import reachability
                reach = reachability.get_state()
                for sw in switches:
                    if sw["id"] in reach:
                        sw["reachable"] = reach[sw["id"]]
            except Exception:
                pass

            # TPS 구역 전원다운 의심: 한 구역(tps_group)의 스위치가 2대 이상이고
            # 전부 도달불가면 정전/전원다운으로 본다(개별 장애와 구분해 즉시 식별).
            zone_outages = []
            try:
                _zg = {}
                for sw in switches:
                    g = sw.get("tps_group")
                    if not g:
                        continue
                    z = _zg.setdefault(g, {"total": 0, "down": 0, "sample": sw.get("tps_num")})
                    z["total"] += 1
                    if sw.get("reachable") is False:
                        z["down"] += 1
                _outage_groups = set()
                for g, z in _zg.items():
                    if z["total"] >= 2 and z["down"] == z["total"]:
                        zone_outages.append({"group": g, "total": z["total"], "down": z["down"]})
                        _outage_groups.add(g)
                # 해당 구역 스위치에 플래그(랙뷰 빨간 점등용)
                for sw in switches:
                    if sw.get("tps_group") in _outage_groups:
                        sw["zone_outage"] = True
            except Exception:
                zone_outages = []
            # 이웃 수집 방식(cdp/lldp/disabled) 주입 — 어느 장비가 추론인지 표시
            try:
                for sw in switches:
                    v = db.get_setting(db_path, "nbrsrc_%d" % sw["id"], "")
                    if v:
                        parts = v.split("|", 1)
                        sw["neighbor_source"] = parts[0]
                        if len(parts) > 1:
                            sw["neighbor_note"] = parts[1]
            except Exception:
                pass

            return jsonify({
                "switches": switches,
                "snapshots": snapshots,
                "demo": config.app.get("demo_mode", False),
                "zone_outages": zone_outages,   # TPS 구역 전원다운 의심 목록
                # 승격되면 실시간으로 False가 되어 UI 배너가 해제된다
                "readonly": bool(app.config.get("IS_READONLY")),
                "primary_host": app.config.get("READONLY_PRIMARY"),
            })
        except Exception as e:
            # CWE-532 fix: Sanitize error messages to prevent credential/path exposure in logs
            sanitized_error = _sanitize_error_msg(str(e))
            log_event("error", "api_state_error", error_type=type(e).__name__, error=sanitized_error)
            # DB 오류면 구체적 원인/힌트를 함께 반환(화면 배너에 상세 표시)
            info = db.get_last_db_error()
            if info:
                return jsonify({"error": "db_error", "db_error": info}), 503
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/health", methods=["GET"])
    def health():
        """경량 상태 점검 — DB 접근 가능 여부와 실패 시 원인/힌트."""
        try:
            db.get_switches(db_path)   # 가벼운 DB 프로브
            return jsonify({"ok": True})
        except Exception:
            info = db.get_last_db_error()
            return jsonify({"ok": False, "db_error": info or {"reason": "DB 오류",
                            "hint": "netdash_error.log를 확인하세요."}}), 503

    @app.route("/api/switches", methods=["GET"])
    def get_switches():
        log_event("info", "api_switches")
        try:
            switches = db.get_switches(db_path)
            return jsonify({"switches": switches})
        except Exception as e:
            # CWE-532 fix: Sanitize error messages to prevent credential/path exposure in logs
            sanitized_error = _sanitize_error_msg(str(e))
            log_event("error", "api_switches_error", error_type=type(e).__name__, error=sanitized_error)
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/switches/manual", methods=["POST"])
    @rate_limit("add_switch_manual", max_requests=30, window_seconds=60)
    def add_switch_manual():
        """수동으로 스위치 1대 등록 (SSRF 검증 포함)."""
        try:
            data = request.get_json(silent=True) or {}
            name = data.get("name", "").strip()
            ip = data.get("ip", "").strip()
            hostname = data.get("hostname", "").strip()
            vendor = data.get("vendor", "unknown").strip()
            location = data.get("location", "").strip()

            if not ip:
                return jsonify({"error": "ip is required"}), 400

            # SSRF 안전선만 유지(loopback/link-local/multicast/reserved 차단).
            # 정적 IP 화이트리스트는 강제하지 않고 '실제 도달(ping/TCP)'로 등록을 허용한다.
            try:
                validated_ip = validate_ipv4(ip, allowed_ip_ranges=config.collector.get("allowed_ip_ranges"))
            except ValueError as e:
                log_event("warning", "add_switch_invalid_ip", ip=ip, reason=str(e))
                return jsonify({"error": str(e)}), 400

            # 도달성 게이트: ping 또는 관리 포트(TCP) 응답 시 등록. force=true/데모모드면 생략.
            if not data.get("force") and not config.app.get("demo_mode"):
                src = pcprofile.get_source_ip(db_path)
                if not collector.is_reachable(validated_ip, source_ip=src):
                    log_event("warning", "add_switch_unreachable", ip=validated_ip)
                    return jsonify({"error": "%s 도달 불가 — ping/TCP(22·443·23·80) 응답이 없습니다. "
                                             "연결(케이블/방화벽/경로)을 확인하세요. 그래도 등록하려면 '강제 등록'을 사용하세요."
                                             % validated_ip, "unreachable": True}), 400

            if not name:
                name = hostname or validated_ip

            rows = [{"name": name, "ip": validated_ip, "hostname": hostname,
                     "vendor": collector.canonical_vendor(vendor),
                     "location": location, "note": data.get("note", "")}]
            ids = db.import_switches_bulk(db_path, rows)
            # 구분(장비 유형) — 선택 입력
            dtype = (data.get("device_type") or "").strip()
            if dtype:
                try:
                    db.update_switch(db_path, ids[0], device_type=dtype)
                except Exception:
                    pass
            return jsonify({"ok": True, "switch_id": ids[0]}), 201
        except Exception as e:
            sanitized = collector._sanitize_error_msg(str(e))
            log_event("error", "add_switch_manual_error", error=sanitized)
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/switches/import", methods=["POST"])
    @rate_limit("import_switches_excel", max_requests=5, window_seconds=60)
    def import_switches_excel():
        """엑셀 파일(xlsx)로 스위치 목록 일괄 등록 (압축폭탄 검증 포함).
        컬럼 순서: name, ip, hostname, vendor, location (헤더 행 필수)
        """
        try:
            import openpyxl
            import zipfile
        except ImportError:
            return jsonify({"error": "required libraries not installed"}), 500

        if "file" not in request.files:
            return jsonify({"error": "file field required"}), 400

        file = request.files["file"]
        if not file.filename.endswith((".xlsx", ".xls")):
            return jsonify({"error": "xlsx file required"}), 400

        try:
            file_content = file.read()

            # HARDENING (CWE-409 Zip Bomb DoS): Validate ZIP compression ratio before processing
            max_compressed_size_mb = 16
            max_uncompressed_size_mb = 50
            max_compression_ratio = 100
            max_single_entry_size_mb = 10

            # Check overall file size
            if len(file_content) / (1024 * 1024) > max_compressed_size_mb:
                log_event("warning", "import_excel_file_too_large", size_mb=len(file_content) / (1024 * 1024))
                return jsonify({"error": f"Compressed file size exceeds {max_compressed_size_mb}MB"}), 413

            # Validate ZIP structure before openpyxl processes it
            try:
                with zipfile.ZipFile(io.BytesIO(file_content), 'r') as zf:
                    total_uncompressed = 0
                    for info in zf.infolist():
                        # Check individual entry size
                        if info.file_size / (1024 * 1024) > max_single_entry_size_mb:
                            log_event("warning", "import_excel_entry_too_large", entry=info.filename, size_mb=info.file_size / (1024 * 1024))
                            return jsonify({"error": f"Single ZIP entry exceeds {max_single_entry_size_mb}MB"}), 413

                        total_uncompressed += info.file_size

                        # Check compression ratio bomb
                        if info.compress_size > 0:
                            ratio = info.file_size / info.compress_size
                            if ratio > max_compression_ratio:
                                log_event("warning", "import_excel_compression_bomb", ratio=ratio, entry=info.filename)
                                return jsonify({"error": f"Compression ratio too high (potential zip bomb)"}), 413

                    if total_uncompressed / (1024 * 1024) > max_uncompressed_size_mb:
                        log_event("warning", "import_excel_uncompressed_too_large", size_mb=total_uncompressed / (1024 * 1024))
                        return jsonify({"error": f"Total uncompressed size exceeds {max_uncompressed_size_mb}MB"}), 413
            except zipfile.BadZipFile:
                log_event("warning", "import_excel_invalid_zip")
                return jsonify({"error": "Invalid ZIP/Excel file"}), 400

            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header = [str(c).lower().strip() if c else "" for c in next(rows_iter)]

            parsed_rows = []
            allowed_ip_ranges = config.collector.get("allowed_ip_ranges")
            for row in rows_iter:
                row_dict = dict(zip(header, row))
                ip = str(row_dict.get("ip", "") or "").strip()
                if not ip:
                    continue

                # HARDENING (CWE-918 SSRF): Validate each IP in bulk import
                try:
                    validated_ip = validate_ipv4(ip, allowed_ip_ranges=allowed_ip_ranges)
                except ValueError as e:
                    log_event("warning", "import_excel_invalid_ip", ip=ip, reason=str(e))
                    continue  # Skip invalid IP instead of failing the entire import

                parsed_rows.append({
                    "name": str(row_dict.get("name", "") or "").strip() or validated_ip,
                    "ip": validated_ip,
                    "hostname": str(row_dict.get("hostname", "") or "").strip(),
                    "vendor": str(row_dict.get("vendor", "unknown") or "unknown").strip(),
                    "location": str(row_dict.get("location", "") or "").strip(),
                })

            if not parsed_rows:
                return jsonify({"error": "no valid rows found"}), 400

            ids = db.import_switches_bulk(db_path, parsed_rows)
            return jsonify({"ok": True, "imported": len(ids), "switch_ids": ids}), 201
        except Exception as e:
            sanitized = collector._sanitize_error_msg(str(e))
            log_event("error", "import_excel_error", error=sanitized)
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/switches/import-inventory", methods=["POST"])
    @rate_limit("import_inventory", max_requests=5, window_seconds=60)
    def import_switch_inventory():
        """IP/SUBNET/HOSTNAME 인벤토리 엑셀 → 스위치 일괄 등록(벤더 unknown, 이후 수정)."""
        log_event("info", "import_inventory_requested")
        if "file" not in request.files:
            return jsonify({"error": "file field required"}), 400
        # 서버/방화벽 임포트와 동일한 검증·한국어 안내(.xls·CSV·손상 파일 안내, 오탐 완화)
        content, err = _read_xlsx_safe(request.files["file"])
        if err:
            return err
        try:
            rows = parse_switch_inventory(io.BytesIO(content))
            allowed = config.collector.get("allowed_ip_ranges")
            valid, skipped = [], 0
            for r in rows:
                try:
                    r["ip"] = validate_ipv4(r["ip"], allowed)
                    valid.append(r)
                except ValueError:
                    skipped += 1
            imported = db.import_switches_bulk(db_path, valid) if valid else []
            log_event("info", "inventory_imported", imported=len(imported), skipped=skipped, total=len(rows))
            return jsonify({"ok": True, "imported": len(imported), "skipped": skipped, "total": len(rows)})
        except Exception as e:
            log_event("error", "import_inventory_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    def _read_xlsx_safe(file):
        """업로드 xlsx 수신 + 크기·zip bomb 검증. 반환: (content_bytes, error_response|None)."""
        if not file or not file.filename:
            return None, (jsonify({"error": "파일이 필요합니다."}), 400)
        fn = file.filename.lower()
        if fn.endswith(".xls") and not fn.endswith(".xlsx"):
            return None, (jsonify({"error": "구형 .xls 형식은 지원하지 않습니다. Excel에서 "
                                   "'다른 이름으로 저장 → Excel 통합 문서(*.xlsx)'로 저장해 올려주세요."}), 400)
        if not fn.endswith(".xlsx"):
            return None, (jsonify({"error": ".xlsx 파일이 필요합니다."}), 400)
        content = file.read()
        if len(content) / (1024 * 1024) > 16:
            return None, (jsonify({"error": "파일이 너무 큽니다(16MB 초과)."}), 413)
        import zipfile as _zip
        try:
            with _zip.ZipFile(io.BytesIO(content), "r") as zf:
                total = 0
                for info in zf.infolist():
                    if info.file_size / (1024 * 1024) > 20:
                        return None, (jsonify({"error": "ZIP 내부 항목이 너무 큽니다."}), 413)
                    total += info.file_size
                    # zip bomb 방어: 개별 항목이 크면서(>2MB) 압축비가 과도할 때만 차단
                    # (일반 xlsx의 반복적 XML은 압축비가 높아 오탐 방지 — 총량 상한이 주 방어선)
                    if (info.file_size > 2 * 1024 * 1024 and info.compress_size > 0
                            and info.file_size / info.compress_size > 200):
                        return None, (jsonify({"error": "압축비가 비정상적으로 높습니다(zip bomb 의심)."}), 413)
                if total / (1024 * 1024) > 80:
                    return None, (jsonify({"error": "압축 해제 크기가 너무 큽니다."}), 413)
        except _zip.BadZipFile:
            return None, (jsonify({"error": "올바른 .xlsx 파일이 아닙니다. Excel에서 열어 "
                                   "'다른 이름으로 저장 → Excel 통합 문서(*.xlsx)'로 다시 저장해 올려주세요. "
                                   "(CSV·구형 xls·손상 파일일 수 있습니다)"}), 400)
        return content, None

    @app.route("/api/servers/import", methods=["POST"])
    @rate_limit("import_servers", max_requests=5, window_seconds=60)
    def import_servers():
        """이름/IP(+위치/OS) 엑셀 → 서버 일괄 등록. 상세·계정은 이후 '수집'에서."""
        log_event("info", "import_servers_requested")
        if "file" not in request.files:
            return jsonify({"error": "file field required"}), 400
        content, err = _read_xlsx_safe(request.files["file"])
        if err:
            return err
        try:
            rows = parse_server_inventory(io.BytesIO(content))
            allowed = config.collector.get("allowed_ip_ranges")
            imported = skipped = 0
            for r in rows:
                try:
                    ip = validate_ipv4(r["ip"], allowed)
                except ValueError:
                    skipped += 1
                    continue
                try:
                    # 엑셀에 없는 칸은 넘기지 않는다(None=건드리지 않음) — 같은 엑셀을
                    # 다시 등록해도 이미 배치·수집해 둔 위치·OS가 지워지지 않게.
                    sid = db.save_server(db_path, r["name"], ip,
                                         os_type=(r.get("os_type") or None),
                                         location=(r.get("location") or None))
                    # 엑셀에 있던 hostname·OS 원문도 함께 보존(표시용)
                    _extra = {}
                    if r.get("hostname"):
                        _extra["hostname"] = r["hostname"]
                    if r.get("os_info"):
                        _extra["os_info"] = r["os_info"]
                    if _extra:
                        db.update_server(db_path, sid, **_extra)
                    imported += 1
                except Exception:
                    skipped += 1
            log_event("info", "servers_imported", imported=imported, skipped=skipped, total=len(rows))
            return jsonify({"ok": True, "imported": imported, "skipped": skipped, "total": len(rows)})
        except Exception as e:
            log_event("error", "import_servers_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/firewalls/import", methods=["POST"])
    @rate_limit("import_firewalls", max_requests=5, window_seconds=60)
    def import_firewalls():
        """이름/벤더/호스트IP(+포트/위치) 엑셀 → 방화벽 일괄 등록."""
        log_event("info", "import_firewalls_requested")
        if "file" not in request.files:
            return jsonify({"error": "file field required"}), 400
        content, err = _read_xlsx_safe(request.files["file"])
        if err:
            return err
        try:
            rows = parse_firewall_inventory(io.BytesIO(content))
            allowed = config.collector.get("allowed_ip_ranges")
            imported = skipped = 0
            for r in rows:
                try:
                    host = validate_ipv4(r["host"], allowed)
                except ValueError:
                    skipped += 1
                    continue
                try:
                    db.save_firewall(db_path, r.get("name") or host, r.get("vendor") or "fortigate",
                                     host, port=r.get("port"), location=(r.get("location") or None))
                    imported += 1
                except Exception:
                    skipped += 1
            log_event("info", "firewalls_imported", imported=imported, skipped=skipped, total=len(rows))
            return jsonify({"ok": True, "imported": imported, "skipped": skipped, "total": len(rows)})
        except Exception as e:
            log_event("error", "import_firewalls_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/upload", methods=["POST"])
    @rate_limit("upload_excel", max_requests=5, window_seconds=60)
    def upload_excel():
        """M4: 멀티블록 엑셀 로더 엔드포인트 (압축폭탄 검증 포함).

        스위치/호스트 혼합 엑셀 파일을 자동으로 분리해서 DB에 임포트.
        - 16MB 업로드 제한 강제 (MAX_CONTENT_LENGTH)
        - .xlsx 확장자만 허용
        - 멀티블록 분리 + IP 필터링 + 멱등성(upsert)
        - ZIP 압축폭탄 검증
        """
        log_event("info", "upload_excel_requested")

        if "file" not in request.files:
            log_event("warning", "upload_no_file_field")
            return jsonify({"error": "file field required"}), 400

        file = request.files["file"]
        if not file or not file.filename:
            log_event("warning", "upload_empty_file")
            return jsonify({"error": "file required"}), 400

        # M4: CWE-434 fix - Allow only .xlsx extension (CWE-94 prevention)
        if not file.filename.endswith(".xlsx"):
            log_event("warning", "upload_invalid_extension", filename=file.filename)
            return jsonify({"error": ".xlsx file required"}), 400

        tmp_path = None
        try:
            # HARDENING (CWE-409 Zip Bomb DoS): Validate before tempfile creation
            file_content = file.read()
            import zipfile

            max_compressed_size_mb = 16
            max_uncompressed_size_mb = 50
            max_compression_ratio = 100
            max_single_entry_size_mb = 10

            if len(file_content) / (1024 * 1024) > max_compressed_size_mb:
                log_event("warning", "upload_file_too_large", size_mb=len(file_content) / (1024 * 1024))
                return jsonify({"error": f"Compressed file size exceeds {max_compressed_size_mb}MB"}), 413

            try:
                with zipfile.ZipFile(io.BytesIO(file_content), 'r') as zf:
                    total_uncompressed = 0
                    for info in zf.infolist():
                        if info.file_size / (1024 * 1024) > max_single_entry_size_mb:
                            log_event("warning", "upload_entry_too_large", entry=info.filename)
                            return jsonify({"error": f"Single ZIP entry exceeds {max_single_entry_size_mb}MB"}), 413

                        total_uncompressed += info.file_size

                        if info.compress_size > 0:
                            ratio = info.file_size / info.compress_size
                            if ratio > max_compression_ratio:
                                log_event("warning", "upload_compression_bomb_detected", ratio=ratio)
                                return jsonify({"error": "Compression ratio too high (potential zip bomb)"}), 413

                    if total_uncompressed / (1024 * 1024) > max_uncompressed_size_mb:
                        log_event("warning", "upload_uncompressed_too_large")
                        return jsonify({"error": f"Total uncompressed size exceeds {max_uncompressed_size_mb}MB"}), 413
            except zipfile.BadZipFile:
                log_event("warning", "upload_invalid_zip")
                return jsonify({"error": "Invalid ZIP/Excel file"}), 400

            # M4: Store file in temporary location, delete after processing (CWE-377 fix)
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(file_content)
                tmp_path = tmp.name

            # M4: Parse multiblock excel
            result = load_excel_workbook(tmp_path, read_only=True, data_only=True)

            switches = result.get("switches", [])
            hosts = result.get("hosts", [])
            allowed_ranges = config.collector.get("allowed_ip_ranges")

            # HARDENING (CWE-918 SSRF): Validate IPs in excel_loader output before DB import
            valid_switches = []
            for sw in switches:
                ip = sw.get("ip", "")
                try:
                    sw["ip"] = validate_ipv4(ip, allowed_ranges)
                    valid_switches.append(sw)
                except ValueError as e:
                    log_event("warning", "upload_switch_invalid_ip", ip=ip, reason=str(e))
                    result["diagnostics"].setdefault("warnings", []).append(
                        f"Switch IP rejected (SSRF): {ip} — {e}"
                    )

            valid_hosts = []
            for h in hosts:
                ip = h.get("ip", "")
                try:
                    h["ip"] = validate_ipv4(ip, allowed_ranges)
                    valid_hosts.append(h)
                except ValueError as e:
                    log_event("warning", "upload_host_invalid_ip", ip=ip, reason=str(e))
                    result["diagnostics"].setdefault("warnings", []).append(
                        f"Host IP rejected (SSRF): {ip} — {e}"
                    )

            # 호스트네임/이름으로 방화벽(FW)·스위치(SW) 구분.
            # 사용자 명명 규칙: 방화벽은 hostname에 'FW', 스위치는 'SW'.
            # FIX(Opus): 'fw' 부분일치는 "GBFW-SW01" 같은 스위치를 오분류 →
            # 토큰 경계 기반(-_공백 구분 또는 fw+숫자)으로만 방화벽 판정.
            # 단, 'sw'가 함께 명시된 이름(...FW-SW...)은 스위치 우선.
            import re as _re
            _FW_PAT = _re.compile(r"(?:^|[-_ .])fw(?:[-_ .\d]|$)", _re.IGNORECASE)
            _SW_PAT = _re.compile(r"(?:^|[-_ .])(?:fa)?sw(?:[-_ .\d]|$)", _re.IGNORECASE)
            sw_rows, fw_rows = [], []
            for sw in valid_switches:
                text = ((sw.get("hostname") or "") + " " + (sw.get("name") or "")).lower()
                is_fw = bool(_FW_PAT.search(text)) and not _SW_PAT.search(text)
                (fw_rows if is_fw else sw_rows).append(sw)

            # Import switches (upsert by name/IP)
            imported_switch_ids = []
            if sw_rows:
                imported_switch_ids = db.import_switches_bulk(db_path, sw_rows)
                log_event("info", "upload_switches_imported", count=len(imported_switch_ids))
            # Import firewalls (host 기준 upsert, vendor=unknown → 이후 수정)
            imported_firewall_ids = []
            if fw_rows:
                imported_firewall_ids = db.import_firewalls_bulk(db_path, fw_rows)
                log_event("info", "upload_firewalls_imported", count=len(imported_firewall_ids))

            # Import hosts (upsert by IP)
            # M7: Excel hosts are the operator's LEDGER (expected inventory), not
            # measured data. Route them through save_ledger_hosts so they populate
            # ledger/mac columns WITHOUT clobbering measured location columns
            # (switch_id/port/located) that a prior collection may have set.
            imported_host_ids = []
            if valid_hosts:
                imported_host_ids = db.save_ledger_hosts(db_path, valid_hosts)
                log_event("info", "upload_hosts_imported", count=len(imported_host_ids))

            # WARNING 2 fix: 유효 row 0건인 경우 400 반환
            if not valid_switches and not valid_hosts:
                log_event("warning", "upload_no_valid_rows")
                return jsonify({
                    "error": "no valid rows found after IP validation",
                    "diagnostics": result["diagnostics"],
                }), 400

            # WARNING 3 fix: diagnostics imported count를 실제 DB import count로 덮어씀
            diagnostics = result["diagnostics"]
            diagnostics["imported_switches"] = len(imported_switch_ids)
            diagnostics["imported_firewalls"] = len(imported_firewall_ids)
            diagnostics["imported_hosts"] = len(imported_host_ids)

            return jsonify({
                "ok": True,
                "diagnostics": diagnostics,
                "imported_switch_ids": imported_switch_ids,
                "imported_firewall_ids": imported_firewall_ids,
                "imported_host_ids": imported_host_ids,
            }), 201

        except Exception as e:
            # CWE-532 fix: Sanitize error messages to prevent path/credential exposure
            sanitized = _sanitize_error_msg(str(e))
            log_event("error", "upload_excel_error", error_type=type(e).__name__, error=sanitized)
            return jsonify({"error": "Internal server error"}), 500

        finally:
            # M4: Clean up temporary file immediately after processing (CWE-377 fix)
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.unlink(tmp_path)
                    log_event("debug", "temp_file_deleted", path=tmp_path)
                except Exception as e:
                    # HARDENING: Retry with delay for Windows file lock issue
                    import time
                    time.sleep(0.1)
                    try:
                        os.unlink(tmp_path)
                    except Exception as retry_e:
                        log_event("warning", "temp_file_delete_failed", path=tmp_path, error=str(retry_e))

    @app.route("/api/search", methods=["GET"])
    def search_ip():
        """IP/이름 종합 검색: 등록 스위치·방화벽 + 수집 ARP + 장부 호스트."""
        ip = request.args.get("ip", "").strip()
        if not ip:
            return jsonify({"error": "ip parameter required"}), 400
        try:
            results = db.search_everywhere(db_path, ip)
            return jsonify({"results": results, "count": len(results)})
        except Exception as e:
            sanitized = collector._sanitize_error_msg(str(e))
            log_event("error", "search_ip_error", error=sanitized)
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/facility", methods=["GET"])
    def facility_list():
        """설비 현황 + 수집 진행 상태 조회.

        현재 연결 위치를 모르는 설비(switch_name 없음 — 온라인/오프라인 무관)는,
        과거 스냅샷의 MAC 위치를 배치 조회해 '과거 연결' 정보(hist_*)로 주입한다.
        (MAC 수집 시점엔 어느 스위치 테이블에 있었으므로 과거 이력으로 특정 가능)
        (스캔 진행 중엔 3초 폴링 부하를 피해 생략 — 스캔이 곧 정확히 갱신)
        """
        try:
            hosts = db.get_facility_hosts(db_path)
            status = facility_mod.get_status()
            if not status.get("running"):
                off = [h for h in hosts if not h.get("switch_name")]
                if off:
                    mac_last = db.get_mac_last_seen(db_path, [h.get("mac") for h in off])
                    for h in off:
                        _hx = re.sub(r"[^0-9a-f]", "", (h.get("mac") or "").lower())
                        hh = mac_last.get(_hx) if len(_hx) == 12 else None
                        if hh and hh.get("switch_name"):
                            h["hist_switch"] = hh.get("switch_name")
                            h["hist_port"] = hh.get("port")
                            h["hist_ts"] = hh.get("ts")
            return jsonify({"hosts": hosts, "status": status})
        except Exception as e:
            log_event("error", "facility_list_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/alerts", methods=["GET"])
    def list_alerts():
        """장비 변경/알람 이벤트 목록 + 미확인 개수. ?kind=&days=&unack=1 필터."""
        try:
            only = request.args.get("unack") == "1"
            kind = (request.args.get("kind") or "").strip() or None
            days_raw = (request.args.get("days") or "").strip()
            days = int(days_raw) if days_raw.isdigit() else None
            return jsonify({"events": db.list_device_events(db_path, limit=300, only_unack=only,
                                                            kind=kind, days=days),
                            "unacked": db.count_unacked_events(db_path)})
        except Exception as e:
            log_event("error", "alerts_list_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/audit", methods=["GET"])
    def list_audit_log():
        """툴 접근(감사) 로그 조회 — 어느 PC가 언제 무엇을 했는지."""
        try:
            return jsonify({"logs": db.list_audit(db_path, limit=300)})
        except Exception as e:
            log_event("error", "audit_list_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/switches/<int:switch_id>/port-history", methods=["GET"])
    def port_history(switch_id):
        """포트 이력: 스냅샷 이력에서 (포트,MAC) 최초/최근 관측. ?port=로 특정 포트만."""
        try:
            port = (request.args.get("port") or "").strip() or None
            return jsonify({"history": db.get_port_history(db_path, switch_id, port=port)})
        except Exception as e:
            log_event("error", "port_history_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    # ── 서버(리눅스/윈도우) 현황 ───────────────────────────────────
    # SQLite INTEGER 상한. Flask <int:id>는 자릿수 제한이 없어 20자리 id가 그대로
    # 들어오면 바인딩에서 OverflowError → 500 + DB 오류 배너 오염. 없는 id이므로 404가 맞다.
    _MAX_ROWID = 2 ** 63 - 1

    def _bad_id(n):
        return not isinstance(n, int) or n <= 0 or n > _MAX_ROWID

    def _sv_text(v, limit=100):
        """요청 본문 값 → 안전한 문자열. dict/list/bool/None은 빈 문자열로 취급.

        (str(None)=='None', 123.strip() AttributeError 등으로 500이 나던 것 방지)
        """
        if v is None or isinstance(v, (dict, list, bool)):
            return ""
        return str(v).strip()[:limit]

    @app.route("/api/servers", methods=["GET"])
    def list_servers_endpoint():
        """서버 목록. 물리 서버 + location이 랙 형식(A09U27)이면 room_* 주입
        (서버실 현황 탭 포함용 — VM은 물리 위치가 없으므로 제외)."""
        try:
            from core import serverroom
            # 구분=Server로 분류된 스위치 행을 서버로 편입(멱등) — 서버 현황에 노출
            try:
                db.adopt_server_switches(db_path)
            except Exception:
                pass
            servers = db.list_servers(db_path)
            for sv in servers:
                if not sv.get("is_vm"):
                    room = serverroom.parse_rack(sv.get("location"))
                    if room:
                        sv["room_rack"] = room["rack"]
                        sv["room_unit"] = room["unit"]
                        sv["room_label"] = room["label"]
                        sv["room_height"] = room.get("height", 1)
            return jsonify({"servers": servers})
        except Exception as e:
            log_event("error", "list_servers_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/servers", methods=["POST"])
    @rate_limit("save_server", max_requests=60, window_seconds=60)
    def save_server_endpoint():
        """서버 등록. body: {name, ip, os_type(auto|windows|UNIX 계열), location?, is_vm?}

        os_type은 linux/windows 외 AIX·Solaris·HP-UX·ESXi·BSD·macOS·unix도 허용한다
        (Windows가 아닌 OS는 모두 범용 UNIX 명령 폴백으로 수집).
        """
        try:
            data = request.get_json(silent=True) or {}
            if not isinstance(data, dict):
                return jsonify({"error": "본문은 JSON 객체여야 합니다"}), 400
            name = _sv_text(data.get("name"))
            ip = _sv_text(data.get("ip"), 64)
            if not name or not ip:
                return jsonify({"error": "name과 ip는 필수입니다"}), 400
            try:
                validated_ip = validate_ipv4(ip, config.collector.get("allowed_ip_ranges"))
            except ValueError as e:
                return jsonify({"error": "IP 거부: %s" % e}), 400
            os_type = data.get("os_type") if data.get("os_type") in _SERVER_OS_TYPES else "auto"
            sid = db.save_server(db_path, name, validated_ip, os_type=os_type,
                                 location=_sv_text(data.get("location"), 60) or None,
                                 is_vm=1 if data.get("is_vm") else 0)
            log_event("info", "server_saved", server_id=sid, ip=validated_ip)
            return jsonify({"ok": True, "id": sid}), 201
        except Exception as e:
            log_event("error", "save_server_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/servers/<int:server_id>", methods=["PUT"])
    @rate_limit("update_server", max_requests=60, window_seconds=60)
    def update_server_endpoint(server_id):
        try:
            if _bad_id(server_id):
                return jsonify({"error": "not found"}), 404
            data = request.get_json(silent=True) or {}
            if not isinstance(data, dict):
                return jsonify({"error": "본문은 JSON 객체여야 합니다"}), 400
            fields = {}
            for k in ("name", "os_type", "hostname"):
                if k in data:
                    fields[k] = _sv_text(data[k]) or None
            # 위치는 '비우기'가 되어야 한다 — 랙에서 뺀 서버가 서버실 현황에
            # 영원히 남지 않도록, 빈 문자열은 None(무시)이 아니라 ""(지우기)로 넘긴다.
            if "location" in data:
                fields["location"] = _sv_text(data["location"], 60)
            # os_type은 허용 계열만(임의 값 저장 방지 — 수집 분기가 문자열에 의존)
            if fields.get("os_type") and fields["os_type"] not in _SERVER_OS_TYPES:
                fields["os_type"] = "auto"
            if "is_vm" in data:
                fields["is_vm"] = 1 if data.get("is_vm") else 0
            if not db.get_server(db_path, server_id):
                return jsonify({"error": "not found"}), 404
            # IP 변경 — 화면 수정 모달이 IP 칸을 보여주고 보내는데 지금까지 조용히 무시됐다.
            if "ip" in data:
                new_ip = _sv_text(data["ip"], 64)
                if new_ip:
                    try:
                        new_ip = validate_ipv4(new_ip, config.collector.get("allowed_ip_ranges"))
                    except ValueError as e:
                        return jsonify({"error": "IP 거부: %s" % e}), 400
                    dup = db.get_server_by_ip(db_path, new_ip)
                    if dup and dup.get("id") != server_id:
                        return jsonify({"error": "이미 등록된 IP입니다: %s" % new_ip}), 409
                    fields["ip"] = new_ip
            db.update_server(db_path, server_id, **fields)
            return jsonify({"ok": True})
        except Exception as e:
            log_event("error", "update_server_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/servers/<int:server_id>", methods=["DELETE"])
    @rate_limit("delete_server", max_requests=30, window_seconds=60)
    def delete_server_endpoint(server_id):
        try:
            if _bad_id(server_id):
                return jsonify({"error": "not found"}), 404
            n = db.delete_server(db_path, server_id)
            return (jsonify({"ok": True}) if n else (jsonify({"error": "not found"}), 404))
        except Exception as e:
            log_event("error", "delete_server_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/servers/<int:server_id>/collect", methods=["POST"])
    @rate_limit("collect_server_host", max_requests=30, window_seconds=60)
    def collect_server_endpoint(server_id):
        """서버 수집(백그라운드). 계정 없이도 무자격 수집(스캔·역DNS·ARP 대조) 수행.
        body: {username?, password?, persist?}"""
        try:
            from core import server_collector
            if _bad_id(server_id):
                return jsonify({"error": "not found"}), 404
            sv = db.get_server(db_path, server_id)
            if not sv:
                return jsonify({"error": "not found"}), 404
            # SSRF 방어: DB에 저장된 IP도 수집 직전 재검증(스위치 일괄수집과 동일 정책).
            # 등록 이후 allowed_ip_ranges가 좁혀졌을 수 있다.
            try:
                validate_ipv4(sv.get("ip"), config.collector.get("allowed_ip_ranges"))
            except ValueError as e:
                return jsonify({"error": "IP 거부: %s" % e}), 400
            data = request.get_json(silent=True) or {}
            if not isinstance(data, dict):
                return jsonify({"error": "본문은 JSON 객체여야 합니다"}), 400
            username = _sv_text(data.get("username"), 128) or None
            password = data.get("password") if isinstance(data.get("password"), str) else None
            # 계정 미입력 시 저장 계정 사용(SSH 상세) — 없으면 무자격 수집만
            if not (username and password):
                blob = db.get_server_credential(db_path, server_id)
                if blob:
                    dec = credentials.decrypt_credential(blob)
                    if dec and "|" in dec:
                        username, password = dec.split("|", 1)
            elif data.get("persist"):
                cred_blob = credentials.encrypt_credential(username, password)
                if cred_blob:
                    db.update_server_cred(db_path, server_id, cred_blob)
                    log_event("info", "server_credential_persisted", server_id=server_id)
            threading.Thread(
                target=server_collector.collect_server,
                args=(db_path, server_id, username, password), daemon=True).start()
            return jsonify({"ok": True, "status": "collecting"}), 202
        except Exception as e:
            log_event("error", "collect_server_host_error",
                      error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/servers/<int:server_id>/diagnose", methods=["POST"])
    @rate_limit("diagnose_server", max_requests=20, window_seconds=60)
    def diagnose_server_endpoint(server_id):
        """서버 1대 무자격 진단(동기) — 도달성·열린 포트·hostname·연결 스위치.

        계정 없이 확인 가능한 것만 본다. 수집과 같은 경로를 쓰되 자격증명을 주지 않는다.
        """
        try:
            from core import server_collector
            if _bad_id(server_id):
                return jsonify({"error": "not found"}), 404
            sv = db.get_server(db_path, server_id)
            if not sv:
                return jsonify({"error": "not found"}), 404
            try:
                validate_ipv4(sv.get("ip"), config.collector.get("allowed_ip_ranges"))
            except ValueError as e:
                return jsonify({"error": "IP 거부: %s" % e}), 400
            res = server_collector.collect_server(db_path, server_id, None, None)
            if res.get("status") == "skipped":
                return jsonify({"error": "이미 수집이 진행 중입니다"}), 409
            ports = res.get("open_ports") or ""
            after = db.get_server(db_path, server_id) or {}
            log_event("info", "server_diagnosed", server_id=server_id,
                      status=res.get("status"))
            return jsonify({"ok": True, "diag": {
                "name": sv.get("name"), "ip": sv.get("ip"),
                "reachable": bool(ports),
                "open_ports": ports,
                "ssh_port": server_collector.pick_ssh_port(
                    [int(p) for p in ports.split(",") if p.strip().isdigit()]),
                "hostname": after.get("hostname") or "",
                "mac": after.get("mac") or "",
                "switch_name": after.get("switch_name") or "",
                "switch_port": after.get("switch_port") or "",
                "os_type": after.get("os_type") or "",
                "has_cred": bool(after.get("has_cred") or
                                 db.get_server_credential(db_path, server_id)),
                "status": res.get("status"),
                "error": res.get("last_error") or "",
            }}), 200
        except Exception as e:
            log_event("error", "diagnose_server_error",
                      error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/servers/diagnose-all", methods=["POST"])
    @rate_limit("diagnose_all_servers", max_requests=6, window_seconds=60)
    def diagnose_all_servers_endpoint():
        """전 서버 무자격 진단 — 계정을 전혀 쓰지 않는다.

        화면은 '계정 없이 도달성·열린 포트·hostname·연결 스위치를 확인'이라고
        안내하는데, 예전엔 collect-all을 호출해 세션 계정·서버별 저장 계정으로
        실제 SSH 접속을 했다(안내와 다름). 진단 전용 경로를 따로 둔다.
        """
        try:
            from core import server_collector
            if server_collector.get_progress().get("running"):
                return jsonify({"error": "이미 서버 수집/진단이 진행 중입니다"}), 409
            # ids(선택): 서버실 화면처럼 일부만 진단할 때 대상 한정.
            data = request.get_json(silent=True) or {}
            raw_ids = data.get("ids")
            if raw_ids is not None and not isinstance(raw_ids, list):
                return jsonify({"error": "ids는 정수 배열이어야 합니다"}), 400
            ids = None
            if raw_ids:
                # 전부 무효면 조용히 "전체 진단"으로 확대되던 fail-open을 막는다.
                ids = [int(i) for i in raw_ids if str(i).isdigit()]
                if not ids:
                    return jsonify({"error": "ids에 유효한 정수가 없습니다"}), 400
                if len(ids) > 1000:
                    return jsonify({"error": "한 번에 최대 1000대까지 지정할 수 있습니다"}), 400
            threading.Thread(
                target=server_collector.collect_all_servers,
                kwargs={"db_path": db_path, "no_cred": True, "ids": ids},
                daemon=True).start()
            log_event("info", "servers_diagnose_all_started", count=len(ids or []))
            return jsonify({"ok": True, "status": "diagnosing"}), 202
        except Exception as e:
            log_event("error", "diagnose_all_servers_error",
                      error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/servers/collect-all", methods=["POST"])
    @rate_limit("collect_all_servers", max_requests=6, window_seconds=60)
    def collect_all_servers_endpoint():
        """등록된 전 서버 일괄 (재)수집(백그라운드).

        body: {username?, password?, persist?} — 공통 계정(선택). 주면 전 서버에
        그 계정으로 접속해 OS 자동 인식 + 상세 수집. 없으면 서버별 저장 계정 사용.
        """
        try:
            from core import server_collector
            data = request.get_json(silent=True) or {}
            if not isinstance(data, dict):
                return jsonify({"error": "본문은 JSON 객체여야 합니다"}), 400
            if server_collector.get_progress().get("running"):
                return jsonify({"error": "이미 서버 수집이 진행 중입니다"}), 409
            cu = _sv_text(data.get("username"), 128) or None
            cp = data.get("password") if isinstance(data.get("password"), str) else None
            # 요청에 계정이 없으면 이 세션에 보관된 수집 계정 사용(메모리·TTL)
            from_session = False
            if not (cu and cp):
                _sc = _session_cred("server")
                if _sc:
                    cu, cp = _sc
                    from_session = True
            # 세션 계정은 '디스크에 안 남긴다'가 설계(core/session_creds.py)다.
            # persist와 함께 오면 세션 계정이 전 서버 cred_blob으로 영구화되므로 막는다.
            persist = bool(data.get("persist"))
            if persist and from_session:
                persist = False
                log_event("info", "server_persist_skipped_session_cred")
            # 선택 수집(없으면 전체) — 검증 없이 넘기면 워커 스레드가 int() 예외로
            # 조용히 죽는데 API는 202를 준다(사용자는 수집이 도는 줄 안다).
            ids = data.get("ids")
            if isinstance(ids, list) and not ids:
                ids = None                      # 빈 배열 = 전체(화면 규약 유지)
            if ids is not None:
                if not isinstance(ids, list):
                    return jsonify({"error": "ids는 배열이어야 합니다"}), 400
                if len(ids) > 1000:
                    return jsonify({"error": "ids가 너무 많습니다(최대 1000)"}), 400
                clean = []
                for raw in ids:
                    try:
                        n = int(raw)
                    except (TypeError, ValueError):
                        return jsonify({"error": "ids는 정수여야 합니다"}), 400
                    if not _bad_id(n):
                        clean.append(n)
                if not clean:
                    return jsonify({"error": "유효한 서버 id가 없습니다"}), 400
                ids = clean
            threading.Thread(
                target=server_collector.collect_all_servers,
                kwargs={"db_path": db_path, "common_user": cu,
                        "common_pass": cp, "persist": persist, "ids": ids},
                daemon=True).start()
            return jsonify({"ok": True, "status": "collecting"}), 202
        except Exception as e:
            log_event("error", "collect_all_servers_error",
                      error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/wall", methods=["GET"])
    def wallboard():
        """관제(월보드) 모드 — 대형 모니터용 읽기 전용 전체화면."""
        return _render_page("wall.html")

    @app.route("/api/wall", methods=["GET"])
    def wall_data():
        """월보드 데이터: 요약 카운터 + 문제 장비 + 최근 알람."""
        try:
            switches = db.get_switches(db_path)
            try:
                from core import reachability
                reach = reachability.get_state()
                fw_reach = reachability.get_fw_state()
            except Exception:
                reach = {}
                fw_reach = {}
            # 방화벽 장애 — 이벤트(firewall_unreachable)는 남는데 관제 화면에는
            # 카테고리도 타일도 없어서 아무 데도 안 보이던 것을 노출한다.
            firewalls = db.list_firewalls(db_path)
            fw_unreach = [f for f in firewalls if fw_reach.get(f["id"]) is False]
            fw_failed = [f for f in firewalls if f.get("status") == "failed"]
            failed = [s for s in switches if s.get("status") == "failed"]
            alerts_sw = [s for s in switches if (s.get("alert") or "none") != "none"]
            unreach = [s for s in switches if reach.get(s["id"]) is False]
            fac = db.get_facility_hosts(db_path)
            fac_off = [h for h in fac if not h.get("online")]
            # 오프라인 설비 MAC의 '마지막 관측 위치'를 1회 배치 조회(호스트별 N쿼리 방지 — 성능).
            _mac_last = (db.get_mac_last_seen(db_path, [h.get("mac") for h in fac_off])
                         if fac_off else {})

            def _hist_by_mac(mac):
                import re as _re
                h = _re.sub(r"[^0-9a-f]", "", (mac or "").lower())
                return _mac_last.get(h) if len(h) == 12 else None

            def _fac_switch_of(h):
                """설비의 연결 스위치명(현재 없으면 과거 이력). 없으면 '미확인'."""
                sn = h.get("switch_name")
                if not sn:
                    sn = (_hist_by_mac(h.get("mac")) or {}).get("switch_name")
                return sn or "미확인"

            # TPS 구역 전원다운 의심: 한 구역의 스위치가 2대 이상이고 전부 도달불가면 정전 의심
            zone_out = []
            try:
                from core import tps_location as _tpsloc
                _zg = {}
                for s in switches:
                    info = _tpsloc.parse(s.get("hostname"))
                    if not info:
                        continue
                    g = "%d공장 %s(%s) %d층" % (info["phase"], info["building_name"],
                                                info["building_code"], info["floor"])
                    z = _zg.setdefault(g, {"total": 0, "down": 0})
                    z["total"] += 1
                    if reach.get(s["id"]) is False:
                        z["down"] += 1
                zone_out = [{"group": g, "total": z["total"], "down": z["down"]}
                            for g, z in _zg.items()
                            if z["total"] >= 2 and z["down"] == z["total"]]
            except Exception:
                zone_out = []

            # 대역별 게이트웨이(수집에 쓴 TPS 스위치) 이름 — '위치 미확인'을 조치 가능하게 안내
            _gw_by_subnet = {}
            try:
                from core import facility as _facmod
                _swname = {s["id"]: s.get("name") for s in switches}
                for _sn, _sid in _facmod.get_band_map(db_path).items():
                    _gw_by_subnet[_sn] = _swname.get(_sid)
            except Exception:
                _gw_by_subnet = {}

            def _facility_detail(h):
                """오프라인 설비의 위치 표기 — 오프라인이면 MAC이 테이블에서
                빠져 포트가 비므로, 마지막 확인 위치/경유/대역으로 상황을 명확히 한다.

                '위치 미확인'은 이 설비 MAC이 등록 스위치 MAC 테이블 어디에도 매칭되지
                않았다는 뜻(=연결 액세스 스위치 미수집). 대역 게이트웨이와 조치를 함께 안내.
                """
                sw = h.get("switch_name")
                port = h.get("port")
                if sw and port:
                    tag = "%s · %s" % (sw, port)
                    return tag if h.get("direct") else (tag + " (경유)")
                if h.get("via"):
                    return "경유 %s" % h["via"]
                if sw:
                    return "%s (포트 미확인)" % sw
                # 과거 연결 이력: 현재 위치를 몰라도, 이 MAC이 과거 스냅샷에서 학습된
                # '마지막 위치'가 있으면 그걸 보여준다(배치 맵 사용 — 성능).
                hist = _hist_by_mac(h.get("mac"))
                if hist and hist.get("switch_name"):
                    when = (hist.get("ts") or "")[:16]
                    return "과거 확인: %s · %s%s (현재 끊김)" % (
                        hist["switch_name"], hist.get("port") or "포트?",
                        (" · " + when) if when else "")
                subnet = h.get("subnet") or "대역 미상"
                gw = _gw_by_subnet.get(h.get("subnet"))
                if gw:
                    return ("위치 미확인 · %s (게이트웨이 %s에서만 관측 — 연결 액세스 스위치를 "
                            "수집 후 설비 '새로고침')" % (subnet, gw))
                return "위치 미확인 · %s (연결 스위치 미수집 — 해당 스위치 수집 후 설비 '새로고침')" % subnet

            # 오프라인 설비는 '위치 확인된 것'(직접 포트) 먼저 노출 — 미확인 노이즈는 뒤로
            fac_off = sorted(
                fac_off,
                key=lambda h: (0 if (h.get("switch_name") and h.get("port")) else
                               1 if h.get("via") or h.get("switch_name") else 2))

            def _alert_detail(sw):
                """경보 스위치의 문제 포트를 switch_logs 이벤트에서 추출해 상세 표기."""
                import json as _json
                crit = sw.get("alert") == "critical"
                kind = "looping" if crit else "flapping"
                ports = []
                try:
                    logs = db.get_switch_logs(db_path, sw["id"])
                    if logs and logs.get("events_json"):
                        ports = collector._extract_event_ports(
                            _json.loads(logs["events_json"]), kind)
                except Exception:
                    pass
                base = "LOOP 경보" if crit else "FLAP 경보"
                return base + ((" [" + ", ".join(ports) + "]") if ports else "")

            # 설비 실패 통계: 어느 '연결 스위치'에서 많이 끊겼고 그 스위치는 어디에 있나.
            # 현재 switch_name이 비면(위치 미확인) 과거 MAC 이력에서 마지막 스위치를 찾는다.
            from core import tps_location as _tl2, serverroom as _sr2

            def _sw_loc(sw):
                """스위치의 위치 문자열: TPS 라벨 / 서버실 랙 / location / hostname 구역."""
                if not sw:
                    return ""
                info = _tl2.parse(sw.get("hostname"))
                if info:
                    return info["label"]
                room = _sr2.parse_rack(sw.get("location"))
                if room:
                    return room["label"]
                if sw.get("location"):
                    return sw["location"]
                m = re.match(r"^[A-Za-z0-9]+_(.+?)_SW(?:ITCH)?[\d_-]*$",
                             (sw.get("hostname") or sw.get("name") or ""), re.I)
                return m.group(1) if m else ""

            _sw_by_name = {s.get("name"): s for s in switches if s.get("name")}
            _fac_sw = {}   # switch_name -> {count, location}
            for h in fac_off:
                sname = h.get("switch_name") or (_hist_by_mac(h.get("mac")) or {}).get("switch_name")
                key = sname or "미확인"
                ent = _fac_sw.setdefault(key, {"count": 0, "location": ""})
                ent["count"] += 1
                if not ent["location"] and sname:
                    ent["location"] = _sw_loc(_sw_by_name.get(sname))
            _fac_subnet_summary = sorted(
                [{"switch": k, "location": v["location"], "count": v["count"]}
                 for k, v in _fac_sw.items()],
                key=lambda x: -x["count"])

            # 카테고리별 정돈된 문제 목록(관제 화면 섹션 렌더용)
            categories = [
                {"key": "zone", "title": "⚡ 구역 전원 다운(의심)", "severity": "bad",
                 "items": [{"name": z["group"], "ip": "",
                            "detail": "구역 스위치 %d대 전부 도달불가 — 전원/정전 의심"
                                      % z["total"]} for z in zone_out]},
                {"key": "unreach", "title": "도달 불가", "severity": "bad",
                 "items": [{"name": s.get("name"), "ip": s.get("ip"),
                            "detail": "TCP-22 응답 없음"} for s in unreach[:30]]},
                {"key": "failed", "title": "수집 실패", "severity": "bad",
                 "items": [{"name": s.get("name"), "ip": s.get("ip"),
                            "detail": (s.get("last_error") or "")[:90]}
                           for s in failed[:30]]},
                {"key": "firewall", "title": "🛡 방화벽 장애", "severity": "bad",
                 "items": ([{"name": f.get("name"), "ip": f.get("host"),
                             "detail": "관리 포트 TCP-%s 응답 없음" % (f.get("port") or 443)}
                            for f in fw_unreach[:30]] +
                           [{"name": f.get("name"), "ip": f.get("host"),
                             "detail": "수집 실패 — " + ((f.get("last_error") or "")[:80] or "원인 미상")}
                            for f in fw_failed[:30]
                            if fw_reach.get(f["id"]) is not False])},
                {"key": "alert", "title": "경보(FLAP/LOOP)", "severity": "warn",
                 "items": [{"name": s.get("name"), "ip": s.get("ip"),
                            "detail": _alert_detail(s)} for s in alerts_sw[:30]]},
                {"key": "facility", "title": "설비 연결 실패", "severity": "warn",
                 "total": len(fac_off),
                 "summary": _fac_subnet_summary,   # 대역별 실패 수(많을 때 한눈에)
                 "items": [{"name": h.get("ip"), "ip": h.get("mac") or "",
                            "fip": h.get("ip"), "subnet": h.get("subnet") or "",
                            "switch": _fac_switch_of(h),   # 칩 클릭 필터용
                            "recollect": True,
                            "detail": _facility_detail(h)}
                           # 칩 필터가 클라이언트에서 동작하므로 넉넉히 내려보낸다
                           # (30건 슬라이스면 '미확인' 칩이 항상 빈 목록이 되던 문제)
                           for h in fac_off[:300]]},
            ]

            # 구버전 호환(problems 평면 목록) — 카테고리에서 파생
            problems = []
            seen = set()
            for s, why in ([(x, "도달 불가") for x in unreach] +
                           [(x, "수집 실패") for x in failed] +
                           [(x, "LOOP 경보" if x.get("alert") == "critical" else "FLAP 경보")
                            for x in alerts_sw]):
                if s["id"] in seen:
                    continue
                seen.add(s["id"])
                problems.append({"name": s.get("name"), "ip": s.get("ip"), "why": why})
            return jsonify({
                "total_switches": len(switches),
                "unreachable": len(unreach),
                "failed": len(failed),
                "alert_switches": len(alerts_sw),
                "facility_total": len(fac),
                "facility_offline": len(fac_off),
                "firewalls_total": len(firewalls),
                "firewalls_down": len(set([f["id"] for f in fw_unreach]) |
                                      set([f["id"] for f in fw_failed])),
                "unacked_alerts": db.count_unacked_events(db_path),
                "categories": categories,
                "problems": problems[:30],
                "recent_events": db.list_device_events(db_path, limit=12),
            })
        except Exception as e:
            log_event("error", "wall_data_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/topology", methods=["GET"])
    def get_topology():
        """서버실 트리 구성도(방화벽+L3/백본, L2 숨김, 대역 박스) + 도달성."""
        try:
            from core import topology, reachability
            topo = topology.build_serverroom_tree(db_path)
            reach = reachability.get_state()
            fw_reach = reachability.get_fw_state()
            for n in topo.get("nodes", []):
                if n.get("kind") == "fw":
                    fid = int(str(n["id"])[1:])
                    if fid in fw_reach:
                        n["reachable"] = fw_reach[fid]
                elif n["id"] in reach:
                    n["reachable"] = reach[n["id"]]
            return jsonify(topo)
        except Exception as e:
            log_event("error", "topology_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    # ── 하이브리드 토폴로지 편집기 (v4.4) ───────────────────────────
    @app.route("/api/topology/diagram", methods=["GET"])
    def get_topology_diagram():
        """저장된 구성도(배치+연결) 로드. 없으면 빈 구성."""
        try:
            import json as _json
            raw = db.get_setting(db_path, "topology_diagram", "") or ""
            data = _json.loads(raw) if raw else {"nodes": [], "edges": []}
            return jsonify(data)
        except Exception as e:
            log_event("error", "topo_diagram_get_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"nodes": [], "edges": []})

    @app.route("/api/topology/diagram", methods=["POST"])
    @rate_limit("save_topo_diagram", max_requests=60, window_seconds=60)
    def save_topology_diagram():
        """구성도 저장(배치 좌표·연결·대역칩). body: {nodes:[...], edges:[...]}"""
        try:
            import json as _json
            data = request.get_json(silent=True) or {}
            nodes = data.get("nodes", [])
            edges = data.get("edges", [])
            if not isinstance(nodes, list) or not isinstance(edges, list):
                return jsonify({"error": "nodes/edges must be lists"}), 400
            if len(nodes) > 2000 or len(edges) > 4000:
                return jsonify({"error": "too large"}), 400
            db.set_setting(db_path, "topology_diagram",
                           _json.dumps({"nodes": nodes, "edges": edges}))
            log_event("info", "topo_diagram_saved", nodes=len(nodes), edges=len(edges))
            return jsonify({"ok": True})
        except Exception as e:
            log_event("error", "topo_diagram_save_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/topology/lookup", methods=["GET"])
    def topology_lookup():
        """IP로 등록 장비 조회(hostname·모델·상태 자동 채움)."""
        try:
            from core import topology, reachability
            ip = (request.args.get("ip") or "").strip()
            dev = topology.lookup_device(db_path, ip)
            if not dev:
                return jsonify({"found": False})
            # 도달성 주입
            if dev["kind"] == "sw":
                r = reachability.get_state().get(dev["id"])
                if r is not None:
                    dev["reachable"] = r
            elif dev["kind"] == "fw":
                r = reachability.get_fw_state().get(dev["id"])
                if r is not None:
                    dev["reachable"] = r
            dev["found"] = True
            return jsonify(dev)
        except Exception as e:
            log_event("error", "topo_lookup_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"found": False, "error": "Internal server error"}), 500

    @app.route("/api/topology/subnet-suggest", methods=["GET"])
    def topology_subnet_suggest():
        """스위치가 나르는 VLAN → 대역 자동 제안(L2 대역 박스 자동 채움)."""
        try:
            from core import topology
            ip = (request.args.get("ip") or "").strip()
            return jsonify({"subnets": topology.subnet_suggest(db_path, ip)})
        except Exception as e:
            log_event("error", "topo_subnet_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"subnets": []})

    @app.route("/api/topology/link-ports", methods=["POST"])
    def topology_link_ports():
        """두 장비 사이 연결 포트 자동 인식. body: {a_ip, b_ip}"""
        try:
            from core import topology
            data = request.get_json(silent=True) or {}
            res = topology.resolve_link_ports(db_path, data.get("a_ip"), data.get("b_ip"))
            return jsonify(res)
        except Exception as e:
            log_event("error", "topo_linkports_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"a_port": None, "b_port": None, "method": "none"})

    @app.route("/api/topology/serverroom", methods=["GET"])
    def topology_serverroom():
        """서버실 등록 장비를 정보 채운 노드로 반환(편집기 '서버실 현황 불러오기')."""
        try:
            from core import topology
            return jsonify(topology.serverroom_devices(db_path))
        except Exception as e:
            log_event("error", "topo_serverroom_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"nodes": []})

    @app.route("/api/topology/subnets", methods=["GET"])
    def topology_subnets():
        """구성도 대역 드롭다운 목록(스위치 SVI + 설비 대역)."""
        try:
            from core import topology
            return jsonify({"subnets": topology.list_subnets(db_path)})
        except Exception as e:
            log_event("error", "topo_subnets_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"subnets": []})

    @app.route("/api/topology/switches", methods=["GET"])
    def topology_switches():
        """선택한 대역의 스위치를 편집기 노드로 반환('스위치 현황 불러오기')."""
        try:
            from core import topology
            subnet = (request.args.get("subnet") or "").strip()
            return jsonify(topology.switches_in_subnet(db_path, subnet))
        except Exception as e:
            log_event("error", "topo_switches_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"nodes": []})

    @app.route("/api/configs/diff", methods=["GET"])
    def config_diff():
        """설정 백업 두 개(a=이전, b=이후)의 줄 단위 diff. b 생략 시 a의 직전 백업과 비교."""
        try:
            import difflib
            b_id = request.args.get("b", "")
            a_id = request.args.get("a", "")
            if not (a_id.isdigit()):
                return jsonify({"error": "a required"}), 400
            newer = db.get_config_backup_content(db_path, int(a_id))
            if not newer:
                return jsonify({"error": "not found"}), 404
            if b_id.isdigit():
                older = db.get_config_backup_content(db_path, int(b_id))
            else:
                # 같은 스위치의 직전 백업
                backups = db.get_config_backups(db_path, newer["switch_id"], limit=20)
                prev = [x for x in backups if x["id"] < newer["id"]]
                older = db.get_config_backup_content(db_path, prev[0]["id"]) if prev else None
            old_lines = (older.get("content") or "").splitlines() if older else []
            new_lines = (newer.get("content") or "").splitlines()
            diff = list(difflib.unified_diff(
                old_lines, new_lines,
                fromfile="이전(%s)" % ((older or {}).get("ts", "-")[:16]),
                tofile="현재(%s)" % (newer.get("ts", "-")[:16]),
                lineterm="", n=3))
            return jsonify({"ok": True, "diff": diff,
                            "same": not diff,
                            "older_ts": (older or {}).get("ts"),
                            "newer_ts": newer.get("ts")})
        except Exception as e:
            log_event("error", "config_diff_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/settings/email", methods=["GET"])
    def get_email_settings():
        """알람 이메일 설정 조회(SMTP 비밀번호는 저장 여부만)."""
        try:
            return jsonify({
                "enabled": db.get_setting(db_path, "email_enabled", "0") == "1",
                "smtp_host": db.get_setting(db_path, "smtp_host", "") or "",
                "smtp_port": db.get_setting(db_path, "smtp_port", "25") or "25",
                "smtp_from": db.get_setting(db_path, "smtp_from", "netdash@localhost") or "",
                "email_to": db.get_setting(db_path, "email_to", "") or "",
                "min_sev": db.get_setting(db_path, "email_min_sev", "warning") or "warning",
                "has_auth": bool(db.get_setting(db_path, "smtp_auth_blob", "")),
            })
        except Exception as e:
            log_event("error", "get_email_settings_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/settings/email", methods=["POST"])
    @rate_limit("set_email", max_requests=20, window_seconds=60)
    def set_email_settings():
        """알람 이메일 설정 저장. SMTP 인증정보는 DPAPI 암호화 저장(입력 시에만 갱신)."""
        try:
            data = request.get_json(silent=True) or {}
            db.set_setting(db_path, "email_enabled", "1" if data.get("enabled") else "0")
            db.set_setting(db_path, "smtp_host", (data.get("smtp_host") or "").strip()[:200])
            port_raw = str(data.get("smtp_port") or "25").strip()
            db.set_setting(db_path, "smtp_port", port_raw if port_raw.isdigit() else "25")
            db.set_setting(db_path, "smtp_from", (data.get("smtp_from") or "netdash@localhost").strip()[:200])
            db.set_setting(db_path, "email_to", (data.get("email_to") or "").strip()[:500])
            sev = data.get("min_sev") or "warning"
            db.set_setting(db_path, "email_min_sev", sev if sev in ("warning", "info") else "warning")
            # 인증(선택): 입력됐을 때만 갱신
            user = (data.get("smtp_user") or "").strip()
            pw = data.get("smtp_pass") or ""
            if user and pw:
                blob = credentials.encrypt_text("%s|%s" % (user, pw))
                if blob:
                    db.set_setting(db_path, "smtp_auth_blob", blob)
            elif data.get("clear_auth"):
                db.set_setting(db_path, "smtp_auth_blob", "")
            log_event("info", "email_settings_saved")
            return jsonify({"ok": True})
        except Exception as e:
            log_event("error", "set_email_settings_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/settings/email/test", methods=["POST"])
    @rate_limit("email_test", max_requests=5, window_seconds=60)
    def test_email():
        """테스트 메일 즉시 발송."""
        try:
            from core import notifier
            ok = notifier.send_email(db_path, "[NetDash] 테스트 메일",
                                     "NetDash 알람 이메일 설정이 정상입니다.")
            return jsonify({"ok": ok,
                            "detail": "발송 성공" if ok else
                            "발송 실패 — 수신 주소, 또는 (직접 전달 모드면) 수신 도메인 메일 서버 도달 여부를 확인하세요"})
        except Exception as e:
            log_event("error", "email_test_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/switches/<int:switch_id>/configs", methods=["GET"])
    def list_switch_configs(switch_id):
        """스위치의 설정(running-config) 백업 목록(내용 제외)."""
        try:
            return jsonify({"backups": db.get_config_backups(db_path, switch_id)})
        except Exception as e:
            log_event("error", "configs_list_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/report/pptx", methods=["GET"])
    @rate_limit("report_pptx", max_requests=10, window_seconds=60)
    def report_pptx():
        """네트워크 구성도 PPTX 자동 생성. ?customer=고객사명&date=YYYY-MM-DD"""
        try:
            from core import pptx_report
            customer = (request.args.get("customer") or "").strip()[:60]
            gdate = (request.args.get("date") or "").strip()[:10]
            if gdate and not re.match(r"^\d{4}-\d{2}-\d{2}$", gdate):
                gdate = None
            data, fname = pptx_report.build_pptx(db_path, customer=customer, generated=gdate)
            import urllib.parse
            quoted = urllib.parse.quote(fname)
            return Response(
                data,
                mimetype="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                headers={"Content-Disposition": "attachment; filename*=UTF-8''%s" % quoted})
        except Exception as e:
            log_event("error", "report_pptx_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "PPTX 생성 실패: " + collector._sanitize_error_msg(str(e))}), 500

    @app.route("/api/configs/export-all", methods=["GET"])
    def export_all_configs():
        """config 백업 ZIP 다운로드. ?ids=1,2,3이면 선택 장비만, 없으면 전체."""
        try:
            import io as _io
            import zipfile as _zip
            from datetime import datetime as _dt
            # 선택된 스위치 id 필터(체크박스 선택 다운로드)
            ids_raw = (request.args.get("ids") or "").strip()
            sel_ids = None
            if ids_raw:
                sel_ids = set()
                for tok in ids_raw.split(","):
                    tok = tok.strip()
                    if tok.isdigit():
                        sel_ids.add(int(tok))
            buf = _io.BytesIO()
            count = 0
            with _zip.ZipFile(buf, "w", _zip.ZIP_DEFLATED) as zf:
                for sw in db.get_switches(db_path):
                    if sel_ids is not None and sw["id"] not in sel_ids:
                        continue
                    backups = db.get_config_backups(db_path, sw["id"], limit=1)
                    if not backups:
                        continue
                    row = db.get_config_backup_content(db_path, backups[0]["id"])
                    if not row or not row.get("content"):
                        continue
                    # 파일명: 이름_IP_백업시각.txt (금지문자 제거)
                    safe = re.sub(r"[^A-Za-z0-9._-]", "_", "%s_%s" % (sw.get("name") or sw["id"], sw.get("ip") or ""))
                    fname = "%s_%s.txt" % (safe, (row.get("ts") or "")[:10])
                    zf.writestr(fname, row["content"])
                    count += 1
            if count == 0:
                msg = ("선택한 장비에 저장된 config 백업이 없습니다." if sel_ids
                       else "저장된 config 백업이 없습니다. 스위치를 수집하면 자동 백업됩니다.")
                return jsonify({"error": msg}), 404
            buf.seek(0)
            stamp = _dt.now().strftime("%Y%m%d")
            return Response(buf.read(), mimetype="application/zip",
                            headers={"Content-Disposition":
                                     "attachment; filename=netdash_configs_%s.zip" % stamp})
        except Exception as e:
            log_event("error", "configs_export_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/serverroom/export", methods=["GET"])
    def export_serverroom():
        """서버실 랙 배치를 엑셀(랙 그리드 형식)로 다운로드."""
        try:
            from core import serverroom
            devices = []
            for sw in db.get_switches(db_path):
                room = serverroom.parse_rack(sw.get("location"))
                if room:
                    devices.append({"name": sw.get("name") or sw.get("hostname"),
                                    "ip": sw.get("ip"), "rack": room["rack"], "unit": room["unit"], "height": room.get("height", 1),
                                    "device_type": sw.get("device_type")})
            for f in db.list_firewalls(db_path):
                room = serverroom.parse_rack(f.get("location"))
                if room:
                    devices.append({"name": f.get("name"), "ip": f.get("host"),
                                    "rack": room["rack"], "unit": room["unit"], "height": room.get("height", 1),
                                    "device_type": "Firewall"})
            # 물리 서버도 랙에 꽂혀 있다. 화면(카드뷰·랙뷰)과 CSV에는 나오는데
            # 엑셀만 빠져 있어서 랙 배치도를 인쇄하면 서버가 통째로 없었다.
            # (VM은 물리 위치가 없으므로 화면과 동일하게 제외)
            for s in db.list_servers(db_path):
                if s.get("is_vm"):
                    continue
                room = serverroom.parse_rack(s.get("location"))
                if room:
                    devices.append({"name": s.get("name"), "ip": s.get("ip"),
                                    "rack": room["rack"], "unit": room["unit"], "height": room.get("height", 1),
                                    "device_type": "Server"})
            if not devices:
                return jsonify({"error": "서버실 소속 장비가 없습니다. location에 'A03U36' 형식으로 랙/유닛을 기재하세요."}), 404
            data = serverroom.build_rack_xlsx(devices)
            return Response(data,
                            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": "attachment; filename=serverroom_rack.xlsx"})
        except Exception as e:
            log_event("error", "serverroom_export_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/configs/<int:backup_id>", methods=["GET"])
    def get_config_content(backup_id):
        """설정 백업 원문 다운로드(txt)."""
        try:
            row = db.get_config_backup_content(db_path, backup_id)
            if not row:
                return jsonify({"error": "not found"}), 404
            fname = "config_sw%s_%s.txt" % (row.get("switch_id"), (row.get("ts") or "")[:10])
            return Response(row.get("content") or "", mimetype="text/plain; charset=utf-8",
                            headers={"Content-Disposition": "attachment; filename=%s" % fname})
        except Exception as e:
            log_event("error", "config_content_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/alerts/ack", methods=["POST"])
    @rate_limit("alerts_ack", max_requests=60, window_seconds=60)
    def ack_alerts():
        """알람 확인 처리. body {ids:[...]} 없으면 전체 확인."""
        try:
            data = request.get_json(silent=True) or {}
            ids = data.get("ids")
            n = db.ack_device_events(db_path, ids if isinstance(ids, list) and ids else None)
            return jsonify({"ok": True, "acked": n})
        except Exception as e:
            log_event("error", "alerts_ack_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/facility/export", methods=["GET"])
    def facility_export():
        """설비 현황 전체를 엑셀(xlsx) 또는 TXT로 내려받기. ?format=xlsx|txt"""
        fmt = (request.args.get("format") or "xlsx").lower()
        try:
            if fmt == "txt":
                data = facility_mod.export_txt(db_path)
                return Response(data, mimetype="text/plain; charset=utf-8",
                                headers={"Content-Disposition": "attachment; filename=facility.txt"})
            data = facility_mod.export_xlsx(db_path)
            return Response(
                data,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=facility.xlsx"})
        except Exception as e:
            log_event("error", "facility_export_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    def _cred_owner():
        """세션 자격증명 소유자 키 — 접속한 원격 주소(브라우저 PC) 기준."""
        return request.remote_addr or "local"

    def _session_cred(kind):
        """(username, password) 또는 None — 그 장비 종류의 세션 계정.

        kind는 'switch'|'server'|'firewall'. 종류를 섞으면 스위치 계정이 전 서버에
        SSH로 시도돼 수집이 실패하고 반복 인증 실패로 계정이 잠길 수 있다.
        """
        from core import session_creds
        return session_creds.get_credential(_cred_owner(), kind)

    @app.route("/api/session/credential", methods=["GET"])
    def session_cred_status():
        """수집 계정 활성 상태(비밀번호는 반환하지 않음). ?kind=로 종류별 조회."""
        from core import session_creds
        return jsonify(session_creds.status(_cred_owner(),
                                            request.args.get("kind") or None))

    @app.route("/api/session/credential", methods=["POST"])
    @rate_limit("session_cred", max_requests=20, window_seconds=60)
    def session_cred_set():
        """수집 계정을 이 세션(메모리)에만 TTL 동안 보관. 디스크 저장 없음."""
        try:
            data = request.get_json(silent=True) or {}
            u = (data.get("username") or "").strip()
            p = data.get("password") or ""
            if not u or not p:
                return jsonify({"error": "username/password required"}), 400
            try:
                u = validate_credential(u)
                p = validate_credential(p)
            except ValueError as e:
                return jsonify({"error": str(e)}), 400
            from core import session_creds
            kind = session_creds._norm_kind(data.get("kind"))
            ttl = session_creds.set_credential(_cred_owner(), u, p, data.get("ttl"),
                                               kind=kind)
            log_event("info", "session_cred_set", owner=_cred_owner(), kind=kind,
                      ttl=ttl, minutes=ttl // 60)
            return jsonify({"ok": True, "ttl": ttl, "kind": kind})
        except Exception as e:
            log_event("error", "session_cred_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/session/credential/lock", methods=["POST"])
    def session_cred_lock():
        """즉시 잠금 — 보관 중인 계정을 폐기. body {kind}로 종류별 잠금도 가능."""
        from core import session_creds
        data = request.get_json(silent=True) or {}
        kind = data.get("kind") if isinstance(data, dict) else None
        kind = session_creds._norm_kind(kind) if kind else None
        session_creds.clear(_cred_owner(), kind)
        log_event("info", "session_cred_locked", owner=_cred_owner(), kind=kind or "all")
        return jsonify({"ok": True})

    @app.route("/api/export/<kind>", methods=["GET"])
    def export_dataset(kind):
        """현황 페이지 공통 내보내기 — CSV/TXT.

        kind: switches | servers | firewalls | serverroom | facility
        ?format=csv|txt (기본 csv, UTF-8 BOM으로 Excel 한글 정상)
        """
        try:
            from core import exporter
            fmt = (request.args.get("format") or "csv").lower()
            data, mime, fname = exporter.export(db_path, kind, fmt)
            return Response(data, mimetype=mime,
                            headers={"Content-Disposition": "attachment; filename=%s" % fname})
        except ValueError:
            return jsonify({"error": "unknown dataset"}), 404
        except Exception as e:
            log_event("error", "export_error", kind=kind,
                      error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/facility/rematch", methods=["POST"])
    @rate_limit("facility_rematch", max_requests=30, window_seconds=60)
    def facility_rematch():
        """설비 현황 새로고침: ping 없이 최신 MAC 스냅샷 기준으로 재대조."""
        try:
            n = facility_mod.rematch(db_path)
            return jsonify({"ok": True, "updated": n})
        except Exception as e:
            log_event("error", "facility_rematch_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/switches/<int:switch_id>/diagnose", methods=["POST"])
    @rate_limit("diagnose_switch", max_requests=20, window_seconds=60)
    def diagnose_switch_endpoint(switch_id):
        """장비 진단: 실제 배너/프롬프트/show version 원문을 반환(벤더 미인식 원인 특정)."""
        try:
            sw = db.get_switch(db_path, switch_id)
            if not sw:
                return jsonify({"error": "not found"}), 404
            data = request.get_json(silent=True) or {}
            username = data.get("username", "")
            password = data.get("password", "")
            if not (username and password):
                blob = db.get_switch_credential(db_path, switch_id)
                if blob:
                    dec = credentials.decrypt_credential(blob)
                    if dec and "|" in dec:
                        username, password = dec.split("|", 1)
            if not (username and password):
                return jsonify({"error": "계정이 필요합니다(입력 또는 저장)"}), 400
            src = pcprofile.get_source_ip(db_path)
            res = collector.diagnose_switch(sw, username, password, source_ip=src)
            # 진단이 벤더를 알아냈고 현재 미지정/오지정이면 자동 교정
            # → 다음 수집부터 올바른 경로(예: Alteon 전용 수집)로 동작
            guess = res.get("guess")
            if guess and guess != (sw.get("vendor") or "").lower():
                try:
                    db.update_switch(db_path, switch_id, vendor=guess)
                    res["vendor_corrected"] = guess
                    log_event("info", "vendor_corrected_by_diagnose",
                              switch_id=switch_id, vendor=guess)
                except Exception:
                    pass
            # 진단이 파싱한 OS/모델/시리얼을 저장(없으면 원문에서 파싱 폴백) — 진단만
            # 눌러도 표의 벤더·모델·버전·시리얼이 채워지도록(누락만 채움; None은 건너뜀).
            if guess:
                try:
                    diag_text = "\n".join(res.get(k) or "" for k in (
                        "version_head", "banner_head", "sysinfo_head", "inventory_head"))
                    osv = res.get("os_version") or collector._parse_os_version(guess, diag_text)
                    model = res.get("model") or collector._parse_model(guess, diag_text)
                    serial = res.get("serial") or collector._parse_serial(guess, diag_text)
                    if osv or model or serial:
                        db.update_switch(db_path, switch_id,
                                         os_version=osv, model=model, serial=serial)
                        res["model_version_filled"] = "%s / %s / SN:%s" % (
                            model or "-", osv or "-", serial or "-")
                except Exception:
                    pass
            log_event("info", "switch_diagnosed", switch_id=switch_id,
                      guess=guess or "unknown", error=res.get("error") or "")
            return jsonify({"ok": True, "diag": res})
        except Exception as e:
            log_event("error", "diagnose_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/switches/bulk-collect/status", methods=["GET"])
    def bulk_collect_status():
        """스위치 일괄 수집 진행 상태 — 다른 화면들과 같은 {running,done,total,message}."""
        with _sw_bulk_lock:
            ids = list(_sw_bulk.get("ids") or [])
            total = _sw_bulk.get("total") or 0
            started = _sw_bulk.get("started")
        if not started or not total:
            return jsonify({"running": False, "done": 0, "total": 0, "message": ""})
        try:
            pending = collector.collecting_ids() & set(ids)
        except Exception:
            pending = set()
        done = total - len(pending)
        running = bool(pending)
        if not running:
            with _sw_bulk_lock:
                _sw_bulk.update(started=False)
        return jsonify({"running": running, "done": done, "total": total,
                        "message": ("스위치 수집 중 (%d/%d)" % (done, total) if running
                                    else "완료(%d대)" % total)})

    @app.route("/api/switches/bulk-collect/stop", methods=["POST"])
    def bulk_collect_stop():
        """대기 중인 스위치 수집을 취소. 이미 접속 중인 장비는 끝까지 마친다."""
        with _sw_bulk_lock:
            if not _sw_bulk.get("started"):
                return jsonify({"ok": False, "error": "진행 중인 수집이 없습니다"}), 400
        n = collector.cancel_pending()
        log_event("info", "bulk_collect_stopped", cancelled=n)
        return jsonify({"ok": True, "cancelled": n})

    @app.route("/api/switches/diagnose-all", methods=["POST"])
    @rate_limit("diagnose_all", max_requests=5, window_seconds=60)
    def diagnose_all_endpoint():
        """등록된 전 스위치를 백그라운드로 일괄 진단(벤더 미지정/오지정 자동 교정)."""
        try:
            with _diag_all_lock:
                if _diag_all["running"]:
                    return jsonify({"error": "이미 일괄 진단이 진행 중입니다",
                                    "running": True}), 409
                switches = db.get_switches(db_path)
                _diag_all.update(running=True, total=len(switches), done=0,
                                 corrected=0, results=[], error=None)
            src = pcprofile.get_source_ip(db_path)
            threading.Thread(target=_run_diagnose_all, args=(db_path, src),
                             daemon=True).start()
            log_event("info", "diagnose_all_started", total=_diag_all["total"])
            return jsonify({"ok": True, "total": _diag_all["total"]}), 202
        except Exception as e:
            with _diag_all_lock:
                _diag_all["running"] = False
            log_event("error", "diagnose_all_error",
                      error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/switches/diagnose-all/status", methods=["GET"])
    def diagnose_all_status_endpoint():
        """일괄 진단 진행 상태 폴링용."""
        with _diag_all_lock:
            return jsonify(dict(_diag_all))

    @app.route("/api/facility/delete-subnet", methods=["POST"])
    @rate_limit("facility_delete_subnet", max_requests=30, window_seconds=60)
    def facility_delete_subnet():
        """설비 현황에서 특정 대역의 수집 결과 전체 삭제."""
        try:
            data = request.get_json(silent=True) or {}
            subnet = (data.get("subnet") or "").strip()
            if not subnet:
                return jsonify({"error": "subnet required"}), 400
            db.clear_facility_subnet(db_path, subnet)
            log_event("info", "facility_subnet_deleted", subnet=subnet)
            return jsonify({"ok": True})
        except Exception as e:
            log_event("error", "facility_delete_subnet_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/facility/detect-subnets", methods=["POST"])
    @rate_limit("facility_detect", max_requests=20, window_seconds=60)
    def facility_detect_subnets():
        """11번 스위치의 directly-connected 대역 자동 도출."""
        try:
            data = request.get_json(silent=True) or {}
            switch_id = data.get("switch_id")
            if not switch_id:
                return jsonify({"error": "switch_id required"}), 400
            username = data.get("username", "")
            password = data.get("password", "")
            if not (username and password):
                blob = db.get_switch_credential(db_path, switch_id)
                if blob:
                    dec = credentials.decrypt_credential(blob)
                    if dec and "|" in dec:
                        username, password = dec.split("|", 1)
            if not (username and password):
                return jsonify({"error": "스위치 계정이 필요합니다(입력 또는 저장)"}), 400
            src = pcprofile.get_source_ip(db_path)
            subnets = facility_mod.detect_subnets(db_path, switch_id, username, password, src)
            return jsonify({"ok": True, "subnets": subnets})
        except Exception as e:
            log_event("error", "facility_detect_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error", "detail": collector._sanitize_error_msg(str(e))}), 500

    @app.route("/api/facility/collect", methods=["POST"])
    @rate_limit("facility_collect", max_requests=10, window_seconds=60)
    def facility_collect():
        """대역 ping sweep + ARP + MAC 대조 (11번 스위치가 직접 ping). 백그라운드."""
        try:
            data = request.get_json(silent=True) or {}
            switch_id = data.get("switch_id")
            subnet = (data.get("subnet") or "").strip()
            username = data.get("username", "")
            password = data.get("password", "")
            if not switch_id or not subnet:
                return jsonify({"error": "switch_id and subnet required"}), 400
            # 대역 검증: 유효 CIDR + 크기 제한(/22 이하, ping 폭증 방지)
            try:
                net = ipaddress.IPv4Network(subnet, strict=False)
            except (ipaddress.AddressValueError, ValueError):
                return jsonify({"error": "invalid subnet (CIDR)"}), 400
            if net.num_addresses > 1024:
                return jsonify({"error": "대역이 너무 큽니다(/22 이하 권장)"}), 400
            # SSRF 유사 차단: allowed_ip_ranges가 설정돼 있으면 스캔 대역이 그 안에
            # 포함돼야 함(게이트웨이 스위치 경유로 임의 대역 ping sweep 지시 방지).
            _allowed = config.collector.get("allowed_ip_ranges")
            if _allowed:
                _ok = False
                for _cidr in _allowed:
                    try:
                        if net.subnet_of(ipaddress.IPv4Network(_cidr, strict=False)):
                            _ok = True
                            break
                    except (ipaddress.AddressValueError, ValueError, TypeError):
                        continue
                if not _ok:
                    return jsonify({"error": "허용되지 않은 대역입니다(allowed_ip_ranges)"}), 400
            sw = db.get_switch(db_path, switch_id)
            if not sw:
                return jsonify({"error": "switch not found"}), 404
            # 게이트웨이 스위치 IP는 등록 시 검증됨.
            # 계정 우선순위: 요청 입력 > 세션 보관(메모리·TTL) > 저장된 자격증명.
            # 설비 대역 수집은 게이트웨이 '스위치'에 SSH로 붙어 ping/ARP를 하므로 스위치 계정이다.
            if not (username and password):
                _sc = _session_cred("switch")
                if _sc:
                    username, password = _sc
            if not (username and password):
                blob = db.get_switch_credential(db_path, switch_id)
                if blob:
                    dec = credentials.decrypt_credential(blob)
                    if dec and "|" in dec:
                        username, password = dec.split("|", 1)
            if not (username and password):
                return jsonify({"error": "스위치 계정이 필요합니다(입력 또는 저장)"}), 400
            src = pcprofile.get_source_ip(db_path)
            started = facility_mod.start_collect_band(db_path, switch_id, subnet, username, password, src)
            if not started:
                return jsonify({"error": "이미 수집 중입니다"}), 409
            # 자동 스캔 대상으로 대역→스위치 매핑 기억
            try:
                facility_mod.remember_band(db_path, subnet, switch_id)
            except Exception:
                pass
            return jsonify({"ok": True, "subnet": subnet})
        except Exception as e:
            log_event("error", "facility_collect_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    def _facility_start_subnet(subnet):
        """대역을 그 대역의 기억된 게이트웨이 스위치·저장 계정으로 재수집 시작.
        반환: (ok, http_status, payload). 관제 재수집·전체 스캔 공용."""
        band_map = facility_mod.get_band_map(db_path)
        sid = band_map.get(subnet)
        if not sid:
            return False, 400, {"error": "이 대역의 게이트웨이 스위치가 기억되지 않았습니다. "
                                          "설비 현황에서 한 번 '대역 수집'을 실행하세요."}
        blob = db.get_switch_credential(db_path, sid)
        username = password = ""
        if blob:
            dec = credentials.decrypt_credential(blob)
            if dec and "|" in dec:
                username, password = dec.split("|", 1)
        if not (username and password):
            cred = pcprofile.get_credential(db_path)
            if cred:
                username, password = cred
        if not (username and password):
            return False, 400, {"error": "게이트웨이 스위치의 저장된 계정이 없습니다."}
        src = pcprofile.get_source_ip(db_path)
        started = facility_mod.start_collect_band(db_path, sid, subnet, username, password, src)
        if not started:
            return False, 409, {"error": "이미 수집 중입니다"}
        return True, 202, {"ok": True, "subnet": subnet}

    @app.route("/api/facility/recollect", methods=["POST"])
    @rate_limit("facility_recollect", max_requests=20, window_seconds=60)
    def facility_recollect():
        """관제/설비에서 특정 설비(IP)의 대역을 연결 게이트웨이에서 재수집."""
        try:
            data = request.get_json(silent=True) or {}
            ip = (data.get("ip") or "").strip()
            subnet = (data.get("subnet") or "").strip()
            if not subnet and ip:
                for h in db.get_facility_hosts(db_path):
                    if h.get("ip") == ip:
                        subnet = h.get("subnet") or ""
                        break
            if not subnet:
                return jsonify({"error": "대역을 찾을 수 없습니다(IP 미등록)"}), 400
            ok, status, payload = _facility_start_subnet(subnet)
            return jsonify(payload), status
        except Exception as e:
            log_event("error", "facility_recollect_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/facility/stop", methods=["POST"])
    def facility_stop():
        """진행 중인 대역 스캔/전체 스캔 중지(부분 결과는 저장)."""
        return jsonify({"ok": facility_mod.request_stop()})

    @app.route("/api/facility/scan-all", methods=["POST"])
    @rate_limit("facility_scan_all", max_requests=4, window_seconds=60)
    def facility_scan_all():
        """기억된 전 대역을 순차 스캔(백그라운드) — 동시 1개 대역만."""
        try:
            if facility_mod.get_status().get("running"):
                return jsonify({"error": "이미 수집 중입니다"}), 409
            bands = facility_mod.get_band_map(db_path)
            if not bands:
                return jsonify({"error": "기억된 대역이 없습니다. 설비 현황에서 대역 수집을 먼저 실행하세요."}), 400
            threading.Thread(target=facility_mod.run_auto_scan,
                             args=(db_path,), daemon=True).start()
            return jsonify({"ok": True, "bands": len(bands)}), 202
        except Exception as e:
            log_event("error", "facility_scan_all_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/vlans", methods=["GET"])
    def get_vlans():
        """전체 VLAN 현황 조회."""
        try:
            vlans = db.get_vlan_summary(db_path)
            return jsonify({"vlans": vlans})
        except Exception as e:
            sanitized = collector._sanitize_error_msg(str(e))
            log_event("error", "get_vlans_error", error=sanitized)
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/reconcile", methods=["GET"])
    def get_reconcile():
        """M7: 장부(엑셀) vs 실측(수집) 대조 결과 조회 (6판정 + summary)."""
        try:
            result = correlator.reconcile(db_path)
            return jsonify(result)
        except Exception as e:
            sanitized = collector._sanitize_error_msg(str(e))
            log_event("error", "reconcile_error", error=sanitized)
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/report", methods=["GET"])
    def get_report():
        """M9: 현재 DB 상태를 4시트 엑셀 보고서로 내려받기."""
        try:
            data = report_builder.build_report(db_path)
            return Response(
                data,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=netdash_report.xlsx"},
            )
        except Exception as e:
            sanitized = collector._sanitize_error_msg(str(e))
            log_event("error", "report_error", error=sanitized)
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/switches/test", methods=["POST"])
    @rate_limit("test_switch", max_requests=20, window_seconds=60)
    def test_switch_connection():
        """M11: 스위치 연결 테스트 (IP+계정 입력값 기반, 저장 전 선검증)."""
        try:
            data = request.get_json(silent=True) or {}
            ip = (data.get("ip") or "").strip()
            if not ip:
                return jsonify({"error": "ip required"}), 400
            try:
                ip = validate_ipv4(ip, config.collector.get("allowed_ip_ranges"))
            except ValueError as e:
                return jsonify({"ok": False, "stage": "reachable", "detail": f"IP rejected: {e}"}), 400
            src = pcprofile.get_source_ip(db_path)
            result = connectivity.test_switch(
                ip, data.get("vendor", ""), data.get("username", ""),
                data.get("password", ""), int(data.get("port", 22)),
                source_ip=src)
            if isinstance(result, dict):
                result["source_ip"] = src or ""  # 화면에 출발지 표시(자동이면 빈값)
            return jsonify(result)
        except Exception as e:
            log_event("error", "test_switch_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/firewalls/test", methods=["POST"])
    @rate_limit("test_firewall", max_requests=20, window_seconds=60)
    def test_firewall_connection():
        """M11: 방화벽 연결 테스트 (입력값 기반, 저장 전 선검증)."""
        try:
            data = request.get_json(silent=True) or {}
            host = (data.get("host") or "").strip()
            vendor = (data.get("vendor") or "").lower()
            if not host:
                return jsonify({"error": "host required"}), 400
            if vendor not in firewall_mod.SUPPORTED_VENDORS:
                return jsonify({"error": "vendor must be fortigate or paloalto"}), 400
            try:
                host = validate_ipv4(host, config.collector.get("allowed_ip_ranges"))
            except ValueError as e:
                return jsonify({"ok": False, "stage": "reachable", "detail": f"host rejected: {e}"}), 400
            port = data.get("port")
            if port not in (None, "") and not (str(port).isdigit() and 1 <= int(port) <= 65535):
                return jsonify({"error": "port must be 1-65535"}), 400
            src = pcprofile.get_source_ip(db_path)
            result = connectivity.test_firewall(
                vendor, host, int(port) if port else None,
                token=data.get("token", ""), username=data.get("username", ""),
                password=data.get("password", ""), verify_ssl=bool(data.get("verify_ssl", False)),
                source_ip=src)
            if isinstance(result, dict):
                result["source_ip"] = src or ""  # 화면에 출발지 표시(자동이면 빈값)
            return jsonify(result)
        except Exception as e:
            log_event("error", "test_firewall_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/firewalls/<int:fid>/diagnose", methods=["POST"])
    @rate_limit("diagnose_firewall", max_requests=20, window_seconds=60)
    def diagnose_firewall_endpoint(fid):
        """방화벽 1대 진단 — 관리 포트/SSH 도달성 + 저장 계정 인증 확인.

        저장된 자격증명만 사용한다(요청으로 계정을 받지 않음).
        FortiGate에 API 토큰만 저장돼 있으면 REST 인증까지 확인되지만
        SSH 터미널에는 쓸 수 없으므로 그 사실을 함께 알린다.
        """
        try:
            if _bad_id(fid):
                return jsonify({"error": "not found"}), 404
            fw = db.get_firewall(db_path, fid)
            if not fw:
                return jsonify({"error": "not found"}), 404
            host = fw.get("host")
            try:
                host = validate_ipv4(host, config.collector.get("allowed_ip_ranges"))
            except ValueError as e:
                return jsonify({"error": "IP 거부: %s" % e}), 400
            vendor = (fw.get("vendor") or "").lower()
            mgmt_port = fw.get("port") or (443 if vendor == "fortigate" else 22)
            token = username = password = ""
            try:
                blob = db.get_firewall_credential(db_path, fid)
                if blob:
                    dec = credentials.decrypt_text(blob)
                    if dec:
                        import json as _json
                        saved = _json.loads(dec)
                        token = saved.get("token", "")
                        username = saved.get("username", "")
                        password = saved.get("password", "")
            except Exception:
                pass
            src = pcprofile.get_source_ip(db_path)
            res = connectivity.test_firewall(vendor, host, mgmt_port, token=token,
                                             username=username, password=password,
                                             verify_ssl=False, source_ip=src)
            log_event("info", "firewall_diagnosed", firewall_id=fid,
                      ok=bool(res.get("ok")), stage=res.get("stage"))
            return jsonify({"ok": True, "diag": {
                "name": fw.get("name"), "host": host, "vendor": vendor,
                "mgmt_port": mgmt_port,
                "tcp_mgmt": connectivity.test_tcp(host, mgmt_port, 3, src),
                "tcp_ssh": connectivity.test_tcp(host, 22, 3, src),
                "has_token": bool(token),
                "has_login": bool(username and password),
                "auth_ok": bool(res.get("ok")) and res.get("stage") == "auth",
                "stage": res.get("stage"), "detail": res.get("detail") or "",
                "source_ip": src or "",
            }}), 200
        except Exception as e:
            log_event("error", "diagnose_firewall_error",
                      error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/switches/<int:switch_id>", methods=["PUT"])
    @rate_limit("update_switch", max_requests=240, window_seconds=60)  # 구분 인라인 변경: 대당 1회 PUT — 연속 편집 허용
    def update_switch_endpoint(switch_id):
        """스위치 등록 정보 수정."""
        try:
            data = request.get_json(silent=True) or {}
            ip = (data.get("ip") or "").strip()
            if ip:  # IP 변경 시 SSRF 검증
                try:
                    ip = validate_ipv4(ip, config.collector.get("allowed_ip_ranges"))
                except ValueError as e:
                    return jsonify({"error": str(e)}), 400
            # device_type 화이트리스트 검증(bulk-set-type과 동일 정책) — 임의 값이
            # 저장돼 UI 드롭다운/토폴로지 분류와 어긋나는 것 방지. 빈 값='미지정' 허용.
            _dt = data.get("device_type") if "device_type" in data else None
            if _dt not in (None, "") and _dt not in DEVICE_TYPES:
                return jsonify({"error": "invalid device_type"}), 400
            # 빈 문자열은 '지우기'다. 예전엔 `or None`으로 접혀 '변경 없음'이 되어,
            # 잘못 들어간 호스트네임·위치를 지우고 저장해도 200 OK만 뜨고 값이 남았다.
            # (키가 아예 없으면 None → 종전대로 '변경 없음')
            def _clr(key, limit=100):
                if key not in data:
                    return None
                return _sv_text(data[key], limit)

            try:
                ok = db.update_switch(
                    db_path, switch_id,
                    # 이름은 표의 주 식별자라 비우기를 허용하지 않는다(빈 값 = 변경 없음)
                    name=(data.get("name") or "").strip() or None,
                    ip=ip or None,
                    hostname=_clr("hostname"),
                    vendor=(collector.canonical_vendor(data.get("vendor"))
                            if (data.get("vendor") or "").strip() else None),
                    location=_clr("location", 60),
                    note=(data.get("note") if "note" in data else None),
                    device_type=_dt,
                )
            except sqlite3.IntegrityError:
                return jsonify({"error": "이미 사용 중인 이름 또는 IP입니다"}), 409
            if not ok:
                return jsonify({"error": "not found"}), 404
            log_event("info", "switch_updated", switch_id=switch_id)
            return jsonify({"ok": True})
        except Exception as e:
            log_event("error", "update_switch_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/switches/<int:switch_id>", methods=["DELETE"])
    @rate_limit("delete_switch", max_requests=60, window_seconds=60)
    def delete_switch_endpoint(switch_id):
        """스위치 삭제 (잘못 등록 시 제거)."""
        try:
            ok = db.delete_switch(db_path, switch_id)
            if not ok:
                return jsonify({"error": "not found"}), 404
            log_event("info", "switch_deleted", switch_id=switch_id)
            return jsonify({"ok": True})
        except Exception as e:
            log_event("error", "delete_switch_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    # 장비 구분(유형) 화이트리스트 — 스위치 현황 '구분' 드롭다운과 동일
    DEVICE_TYPES = {"BackBone", "L3 Switch", "L2 Switch", "L4 Switch",
                    "Server", "Firewall", "AP", "Tablet", "PC", "기타"}

    @app.route("/api/switches/bulk-set-type", methods=["POST"])
    @rate_limit("bulk_set_type", max_requests=60, window_seconds=60)
    def bulk_set_device_type():
        """선택된 스위치들의 구분(장비 유형)을 일괄 변경. body {ids:[...], device_type}"""
        try:
            data = request.get_json(silent=True) or {}
            ids = data.get("ids", [])
            dtype = (data.get("device_type") or "").strip()
            if not isinstance(ids, list) or not ids:
                return jsonify({"error": "ids required"}), 400
            if dtype and dtype not in DEVICE_TYPES:
                return jsonify({"error": "허용되지 않는 구분 값"}), 400
            n = 0
            for raw in ids[:1000]:
                try:
                    if db.update_switch(db_path, int(raw), device_type=dtype):
                        n += 1
                except (TypeError, ValueError):
                    continue
            log_event("info", "bulk_set_type", count=n, device_type=dtype or "(비움)")
            return jsonify({"ok": True, "updated": n})
        except Exception as e:
            log_event("error", "bulk_set_type_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/switches/bulk-zone", methods=["POST"])
    @rate_limit("bulk_zone", max_requests=60, window_seconds=60)
    def bulk_set_zone():
        """선택된 스위치들의 토폴로지 존(구성도 그룹)을 일괄 지정. body {ids:[...], zone}

        zone은 자유 텍스트(예: 'SERVERFARM', 'DMZ', 'ECO-HUB N전산실'). 빈 값이면 지정 해제.
        """
        try:
            data = request.get_json(silent=True) or {}
            ids = data.get("ids", [])
            zone = (data.get("zone") or "").strip()[:60]
            if not isinstance(ids, list) or not ids:
                return jsonify({"error": "ids required"}), 400
            n = 0
            for raw in ids[:1000]:
                try:
                    if db.update_switch(db_path, int(raw), zone=zone):
                        n += 1
                except (TypeError, ValueError):
                    continue
            log_event("info", "bulk_set_zone", count=n, zone=zone or "(해제)")
            return jsonify({"ok": True, "updated": n})
        except Exception as e:
            log_event("error", "bulk_set_zone_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    # ── 저장 계정 관리(관리자) — 목록 조회 + 삭제 ─────────────────
    @app.route("/api/credentials", methods=["GET"])
    def list_credentials_endpoint():
        """저장 계정 현황: 스위치·방화벽(계정 보유만) + PC 프로필. blob은 비노출."""
        try:
            return jsonify({
                "switches": db.list_switch_credentials(db_path),
                "firewalls": db.list_firewall_credentials(db_path),
                "pc_profiles": db.list_pc_profiles(db_path),
            })
        except Exception as e:
            log_event("error", "list_credentials_error",
                      error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/credentials/delete", methods=["POST"])
    @rate_limit("delete_credential", max_requests=30, window_seconds=60)
    def delete_credential_endpoint():
        """저장 계정 삭제. body: {kind: switch|firewall|profile|all, id?, mac?}

        - switch: 해당 스위치의 저장 계정 + enable secret 삭제
        - firewall: 해당 방화벽의 저장 계정 삭제
        - profile: 해당 PC 프로필 삭제(계정·출발지 IP 포함)
        - all: 전체 삭제(스위치·방화벽·프로필 계정 — 프로필 IP는 유지)
        """
        try:
            data = request.get_json(silent=True) or {}
            kind = data.get("kind")
            if kind == "switch":
                sid = int(data.get("id"))
                db.update_cred_blob(db_path, sid, None)
                db.set_setting(db_path, "enable_secret_%d" % sid, "")
                log_event("info", "credential_deleted", kind="switch", switch_id=sid)
                return jsonify({"ok": True})
            if kind == "firewall":
                fid = int(data.get("id"))
                db.clear_firewall_credential(db_path, fid)
                log_event("info", "credential_deleted", kind="firewall", firewall_id=fid)
                return jsonify({"ok": True})
            if kind == "profile":
                mac = (data.get("mac") or "").strip()[:64]
                if not mac:
                    return jsonify({"error": "mac required"}), 400
                n = db.delete_pc_profile(db_path, mac)
                log_event("info", "credential_deleted", kind="profile", mac=mac, deleted=n)
                return jsonify({"ok": True, "deleted": n})
            if kind == "all":
                res = db.clear_all_credentials(db_path)
                log_event("warning", "credential_deleted", kind="all", **res)
                return jsonify({"ok": True, **res})
            return jsonify({"error": "kind must be switch|firewall|profile|all"}), 400
        except (TypeError, ValueError):
            return jsonify({"error": "invalid id"}), 400
        except Exception as e:
            log_event("error", "delete_credential_error",
                      error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/switches/bulk-delete", methods=["POST"])
    @rate_limit("bulk_delete_switch", max_requests=20, window_seconds=60)
    def bulk_delete_switches_endpoint():
        """스위치 여러 대 일괄/선택 삭제. body: {ids:[...]}"""
        try:
            data = request.get_json(silent=True) or {}
            ids = data.get("ids", [])
            if not isinstance(ids, list) or not ids:
                return jsonify({"error": "ids required"}), 400
            if len(ids) > 1000:
                return jsonify({"error": "too many ids"}), 400
            deleted = db.delete_switches_bulk(db_path, ids)
            log_event("info", "switches_bulk_deleted", count=deleted)
            return jsonify({"ok": True, "deleted": deleted})
        except Exception as e:
            log_event("error", "bulk_delete_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/firewalls/<int:fid>", methods=["PUT"])
    @rate_limit("update_firewall", max_requests=60, window_seconds=60)
    def update_firewall_endpoint(fid):
        """방화벽 등록 정보 수정."""
        try:
            data = request.get_json(silent=True) or {}
            host = (data.get("host") or "").strip()
            if host:
                try:
                    host = validate_ipv4(host, config.collector.get("allowed_ip_ranges"))
                except ValueError as e:
                    return jsonify({"error": f"host rejected: {e}"}), 400
            vendor = (data.get("vendor") or "").strip().lower()
            if vendor and vendor not in firewall_mod.SUPPORTED_VENDORS:
                return jsonify({"error": "vendor must be fortigate or paloalto"}), 400
            port = data.get("port")
            if port not in (None, "") and not (str(port).isdigit() and 1 <= int(port) <= 65535):
                return jsonify({"error": "port must be 1-65535"}), 400
            try:
                # location은 빈 문자열("")도 유효(위치 지우기) → 키 존재 시 그대로 전달
                loc = data.get("location")
                ok = db.update_firewall(
                    db_path, fid,
                    name=(data.get("name") or "").strip() or None,
                    vendor=vendor or None,
                    host=host or None,
                    port=int(port) if port not in (None, "") else None,
                    location=(loc.strip() if isinstance(loc, str) else None),
                )
            except sqlite3.IntegrityError:
                return jsonify({"error": "이미 사용 중인 호스트입니다"}), 409
            if not ok:
                return jsonify({"error": "not found"}), 404
            log_event("info", "firewall_updated", firewall_id=fid)
            return jsonify({"ok": True})
        except Exception as e:
            log_event("error", "update_firewall_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/firewalls/<int:fid>", methods=["DELETE"])
    @rate_limit("delete_firewall", max_requests=20, window_seconds=60)
    def delete_firewall_endpoint(fid):
        """방화벽 삭제 (잘못 등록 시 제거)."""
        try:
            ok = db.delete_firewall(db_path, fid)
            if not ok:
                return jsonify({"error": "not found"}), 404
            log_event("info", "firewall_deleted", firewall_id=fid)
            return jsonify({"ok": True})
        except Exception as e:
            log_event("error", "delete_firewall_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/netinfo", methods=["GET"])
    def get_netinfo():
        """M11: PC 로컬 네트워크 정보(이더넷 IP) 조회. 장비 접근에 쓰는 IP 안내용."""
        try:
            info = netinfo.get_network_info()
            info["source_ip"] = db.get_setting(db_path, "source_ip", "") or ""
            # 이 PC 프로필의 실제 적용값(MAC 키 — 다중 PC 운영 시 PC마다 다름)
            try:
                _prof = pcprofile.get_profile(db_path)
                info["source_ip_effective"] = pcprofile.get_source_ip(db_path) or ""
                info["pc_profile"] = ({"mac": _prof["mac"], "hostname": _prof["hostname"],
                                       "has_cred": bool(_prof.get("cred_blob"))}
                                      if _prof else None)
            except Exception:
                info["source_ip_effective"] = info["source_ip"]
                info["pc_profile"] = None
            return jsonify(info)
        except Exception as e:
            log_event("error", "netinfo_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/settings/auto_collect", methods=["GET"])
    def get_auto_collect():
        """자동화 설정 조회(자동 수집·설비 자동 스캔·알람 보존·도달성 감시)."""
        try:
            from core import facility as _fac
            return jsonify({
                "enabled": db.get_setting(db_path, "auto_collect_enabled", "0") == "1",
                "times": db.get_setting(db_path, "auto_collect_times", "06:00,18:00") or "06:00,18:00",
                "facility_enabled": db.get_setting(db_path, "facility_auto_enabled", "0") == "1",
                "facility_time": db.get_setting(db_path, "facility_auto_time", "07:00") or "07:00",
                "facility_bands": len(_fac.get_band_map(db_path)),
                "retention_days": db.get_setting(db_path, "alert_retention_days", "90") or "90",
                "reach_enabled": db.get_setting(db_path, "reach_check_enabled", "1") != "0",
            })
        except Exception as e:
            log_event("error", "get_auto_collect_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/settings/auto_collect", methods=["POST"])
    @rate_limit("set_auto_collect", max_requests=20, window_seconds=60)
    def set_auto_collect():
        """자동화 설정 저장(자동 수집·설비 자동 스캔·알람 보존·도달성 감시)."""
        try:
            data = request.get_json(silent=True) or {}
            enabled = "1" if data.get("enabled") else "0"
            raw = (data.get("times") or "").strip()
            # HH:MM 형식만 허용(쉼표 구분, 최대 6개)
            valid = []
            for t in raw.split(","):
                t = t.strip()
                if re.match(r"^([01]\d|2[0-3]):[0-5]\d$", t):
                    valid.append(t)
            if not valid:
                valid = ["06:00", "18:00"]
            db.set_setting(db_path, "auto_collect_enabled", enabled)
            db.set_setting(db_path, "auto_collect_times", ",".join(valid[:6]))

            # 설비 대역 자동 스캔(1일 1회 HH:MM)
            fac_enabled = "1" if data.get("facility_enabled") else "0"
            fac_time = (data.get("facility_time") or "07:00").strip()
            if not re.match(r"^([01]\d|2[0-3]):[0-5]\d$", fac_time):
                fac_time = "07:00"
            db.set_setting(db_path, "facility_auto_enabled", fac_enabled)
            db.set_setting(db_path, "facility_auto_time", fac_time)

            # 알람 보존 일수(7~365)
            try:
                days = min(365, max(7, int(data.get("retention_days", 90))))
            except (TypeError, ValueError):
                days = 90
            db.set_setting(db_path, "alert_retention_days", str(days))

            # 도달성 감시 on/off
            db.set_setting(db_path, "reach_check_enabled",
                           "1" if data.get("reach_enabled", True) else "0")

            log_event("info", "auto_collect_set", enabled=enabled, times=",".join(valid[:6]),
                      facility=fac_enabled, retention=days)
            return jsonify({"ok": True, "enabled": enabled == "1", "times": ",".join(valid[:6]),
                            "facility_enabled": fac_enabled == "1", "facility_time": fac_time,
                            "retention_days": str(days)})
        except Exception as e:
            log_event("error", "set_auto_collect_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/settings/source_ip", methods=["POST"])
    @rate_limit("set_source_ip", max_requests=20, window_seconds=60)
    def set_source_ip():
        """M12: 장비 접근에 사용할 출발지 IP 설정(빈값=자동/OS 기본). PC 이더넷 IP만 허용."""
        try:
            data = request.get_json(silent=True) or {}
            ip = (data.get("ip") or "").strip()
            if ip and ip not in netinfo.get_local_ipv4_addresses():
                return jsonify({"error": "선택한 IP가 이 PC의 이더넷 IP 목록에 없습니다"}), 400
            db.set_setting(db_path, "source_ip", ip)
            # 이 PC의 프로필(MAC 키)에도 등록 — 다중 PC 운영 시 각 PC가
            # 자기 출발지 IP로 수집하도록(다른 PC 설정에 영향 없음)
            try:
                pcprofile.save_profile(db_path, source_ip=ip or None)
            except Exception as e:
                log_event("warning", "pc_profile_save_failed",
                          error=collector._sanitize_error_msg(str(e)))
            log_event("info", "source_ip_set", source_ip=ip or "(auto)")
            return jsonify({"ok": True, "source_ip": ip})
        except Exception as e:
            log_event("error", "set_source_ip_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    # ── M10: 방화벽 (Palo Alto / Fortinet) ─────────────────────────
    @app.route("/api/firewalls", methods=["GET"])
    def list_firewalls_endpoint():
        """방화벽 장비 목록 조회. location "A09U27"이면 서버실 랙 정보 주입."""
        try:
            from core import serverroom
            fws = db.list_firewalls(db_path)
            for f in fws:
                room = serverroom.parse_rack(f.get("location"))
                if room:
                    f["room_rack"] = room["rack"]
                    f["room_unit"] = room["unit"]
                    f["room_label"] = room["label"]
                    f["room_height"] = room.get("height", 1)
            # 도달성 감시 결과 주입(관리 포트 TCP — True 도달/False 불가/없으면 미확인)
            try:
                from core import reachability
                fr = reachability.get_fw_state()
                for f in fws:
                    if f["id"] in fr:
                        f["reachable"] = fr[f["id"]]
            except Exception:
                pass
            return jsonify({"firewalls": fws})
        except Exception as e:
            log_event("error", "firewalls_list_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/firewalls", methods=["POST"])
    @rate_limit("add_firewall", max_requests=10, window_seconds=60)
    def add_firewall_endpoint():
        """방화벽 장비 등록 (벤더: fortigate | paloalto)."""
        try:
            data = request.get_json(silent=True) or {}
            vendor = (data.get("vendor") or "").lower()
            host = (data.get("host") or "").strip()
            name = (data.get("name") or host).strip()
            if vendor not in firewall_mod.SUPPORTED_VENDORS:
                return jsonify({"error": "vendor must be fortigate or paloalto"}), 400
            if not host:
                return jsonify({"error": "host required"}), 400
            # SSRF: 방화벽 host도 허용 대역 검증
            try:
                host = validate_ipv4(host, config.collector.get("allowed_ip_ranges"))
            except ValueError as e:
                log_event("warning", "firewall_blocked_invalid_ip", host=host, reason=str(e))
                return jsonify({"error": f"host rejected: {e}"}), 400
            # SSRF(CWE-918): port를 정수 1-65535로 강제 검증. SQLite는 타입 강제를
            # 하지 않으므로 '443@evil' 같은 문자열이 저장되어 요청 URL에 주입되는 것을 차단.
            port_raw = data.get("port")
            port = None
            if port_raw is not None and port_raw != "":
                try:
                    port = int(port_raw)
                except (ValueError, TypeError):
                    return jsonify({"error": "port must be an integer 1-65535"}), 400
                if not (1 <= port <= 65535):
                    return jsonify({"error": "port must be an integer 1-65535"}), 400
            location = (data.get("location") or "").strip()
            fid = db.save_firewall(db_path, name, vendor, host,
                                   port, data.get("auth_type", "token"), location=location)
            # M11: 자격증명(토큰/계정)을 DPAPI 암호화하여 저장(입력된 경우만).
            # 저장되면 이후 수집 시 재입력 불필요. 암호화 불가(비Windows) 시 저장 생략.
            cred = {"token": data.get("token", ""), "username": data.get("username", ""),
                    "password": data.get("password", "")}
            if any(cred.values()):
                import json as _json
                blob = credentials.encrypt_text(_json.dumps(cred))
                if blob:
                    db.save_firewall_credential(db_path, fid, blob)
                    log_event("info", "firewall_cred_saved", firewall_id=fid)
            log_event("info", "firewall_added", firewall_id=fid, vendor=vendor)
            return jsonify({"ok": True, "firewall_id": fid, "cred_saved": bool(cred and any(cred.values()))}), 201
        except Exception as e:
            log_event("error", "firewall_add_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/firewalls/<int:fid>", methods=["GET"])
    def get_firewall_detail(fid):
        """방화벽 상세 (인터페이스 + ARP)."""
        try:
            fw = db.get_firewall(db_path, fid)
            if not fw:
                return jsonify({"error": "not found"}), 404
            return jsonify({
                "firewall": fw,
                "interfaces": db.get_firewall_interfaces(db_path, fid),
                "arp": db.get_firewall_arp(db_path, fid),
            })
        except Exception as e:
            log_event("error", "firewall_detail_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/firewalls/<int:fid>/collect", methods=["POST"])
    @rate_limit("collect_firewall", max_requests=10, window_seconds=60)
    def collect_firewall_endpoint(fid):
        """방화벽에서 인터페이스 + ARP 수집 (자격증명은 요청 시점에만 사용)."""
        fw = db.get_firewall(db_path, fid)
        if not fw:
            return jsonify({"error": "not found"}), 404
        # SSRF(CWE-918) 재검증 + 입력 파싱은 동시수집 가드 획득 '전'에 수행 —
        # 이 구간에서 예외(malformed JSON 등)가 나도 잠금이 누수되지 않도록.
        # (이전엔 add(fid) 후 get_json 예외 시 fid가 set에 영구 잔류 → 영구 409 버그)
        try:
            validate_ipv4(fw.get("host"), config.collector.get("allowed_ip_ranges"))
        except ValueError as e:
            db.set_firewall_status(db_path, fid, "failed")
            log_event("warning", "firewall_collect_blocked_invalid_ip", firewall_id=fid, reason=str(e))
            return jsonify({"error": f"firewall host rejected: {e}"}), 400
        fw_port = fw.get("port")
        if fw_port is not None and not (isinstance(fw_port, int) and 1 <= fw_port <= 65535):
            db.set_firewall_status(db_path, fid, "failed")
            return jsonify({"error": "stored firewall port is invalid"}), 400
        data = request.get_json(silent=True) or {}
        token = data.get("token", "")
        username = data.get("username", "")
        password = data.get("password", "")
        provided = bool(token or username or password)  # 요청에 cred 직접 입력 여부
        # M11: 요청에 자격증명이 없으면 저장된(암호화) 자격증명을 복호화해 사용.
        if not (token or username or password):
            blob = db.get_firewall_credential(db_path, fid)
            if blob:
                dec = credentials.decrypt_text(blob)
                if dec:
                    import json as _json
                    try:
                        saved = _json.loads(dec)
                        token = saved.get("token", "")
                        username = saved.get("username", "")
                        password = saved.get("password", "")
                    except (ValueError, TypeError):
                        pass
        # 동시 수집 가드 — 이후 전 구간을 try/finally로 감싸 잠금 해제를 보장.
        with _collecting_fw_lock:
            if fid in _collecting_firewalls:
                return jsonify({"error": "이미 수집 중입니다"}), 409
            _collecting_firewalls.add(fid)
        try:
            # set_status를 try 안으로 이동 — DB 지연/잠금으로 실패해도 아래 finally가
            # fid를 반드시 해제(이전엔 try 밖이라 예외 시 500 + fid 잔류로 영구 409).
            db.set_firewall_status(db_path, fid, "collecting")
            result = firewall_mod.collect_firewall(
                fw["vendor"], fw["host"], fw.get("port"),
                token=token, username=username, password=password,
                verify_ssl=bool(data.get("verify_ssl", False)),
                source_ip=pcprofile.get_source_ip(db_path),
            )
            db.save_firewall_interfaces(db_path, fid, result["interfaces"])
            db.save_firewall_arp(db_path, fid, result["arp"])
            # HA 구성(FortiGate cmdb/system/ha) — 이중화 연결선에 HA 포트 표기용
            if result.get("ha"):
                try:
                    import json as _json
                    db.set_firewall_ha_info(db_path, fid, _json.dumps(result["ha"]))
                except Exception:
                    pass
            # 수집 모달에서 처음 입력한 자격증명은 저장해 다음 수집부터 재입력 불필요.
            if provided:
                try:
                    import json as _json
                    blob = credentials.encrypt_text(_json.dumps(
                        {"token": token, "username": username, "password": password}))
                    if blob:
                        db.save_firewall_credential(db_path, fid, blob)
                except Exception:
                    pass  # 저장 실패는 수집 성공에 영향 없음
            db.set_firewall_status(db_path, fid, "done")
            log_event("info", "firewall_collected", firewall_id=fid,
                      interfaces=len(result["interfaces"]), arp=len(result["arp"]))
            return jsonify({"ok": True,
                            "interfaces": len(result["interfaces"]),
                            "arp": len(result["arp"])})
        except Exception as e:
            try:
                db.set_firewall_status(db_path, fid, "failed")
            except Exception:
                pass  # DB 자체 문제면 상태 기록도 실패 — 무시하고 오류 응답 반환
            sanitized = collector._sanitize_error_msg(str(e))
            log_event("error", "firewall_collect_error", firewall_id=fid, error=sanitized)
            # DB 오류면 원인 힌트를 함께(항상 JSON 응답 → 프론트 '서버 오류' 대신 상세 표시)
            di = db.get_last_db_error()
            if di:
                return jsonify({"error": "DB 오류: " + di.get("reason", ""),
                                "detail": di.get("hint", "") + " (" + (di.get("detail") or "") + ")"}), 503
            return jsonify({"error": "수집 실패", "detail": sanitized}), 502
        finally:
            token = username = password = None
            _collecting_firewalls.discard(fid)

    @app.route("/api/firewalls/collect-all", methods=["POST"])
    @rate_limit("collect_all_firewalls", max_requests=6, window_seconds=60)
    def collect_all_firewalls_endpoint():
        """방화벽 일괄 수집(백그라운드).

        body {ids:[...]} 주면 그 방화벽만, 없으면 전체.
        body에 token 또는 username/password를 주면 그 계정을 공통으로 쓴다
        (방화벽 계정은 스위치·서버와 다르므로 이 화면에서 직접 입력받는다).
        없으면 방화벽 세션 계정 → 각 방화벽 저장 계정 순.
        """
        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return jsonify({"error": "본문은 JSON 객체여야 합니다"}), 400
        ids = data.get("ids") or None
        _tok = data.get("token") if isinstance(data.get("token"), str) else ""
        _u = _sv_text(data.get("username"), 128)
        _p = data.get("password") if isinstance(data.get("password"), str) else ""
        common = (_u, _p) if (_u and _p) else _session_cred("firewall")
        if _tok:
            common = common or (None, None)
        with _fw_all_lock:
            if _fw_all["running"]:
                return jsonify({"error": "이미 수집 중입니다", "status": dict(_fw_all)}), 409
            try:
                _all = db.list_firewalls(db_path)
                if ids:
                    _idset = set(int(x) for x in ids)
                    total = len([f for f in _all if f.get("id") in _idset])
                else:
                    total = len(_all)
            except Exception:
                total = 0
            _fw_all.update(running=True, total=total, done=0, ok=0, message="시작 중", stop=False)
        src = pcprofile.get_source_ip(db_path)
        threading.Thread(target=_run_collect_all_firewalls,
                         args=(db_path, src, ids, common, _tok),
                         daemon=True).start()
        return jsonify({"ok": True, "total": total}), 202

    @app.route("/api/firewalls/diagnose-all", methods=["POST"])
    @rate_limit("diagnose_all_firewalls", max_requests=6, window_seconds=60)
    def diagnose_all_firewalls_endpoint():
        """전 방화벽 도달성·인증 확인. 수집(인터페이스·ARP 저장)은 하지 않는다."""
        with _fw_diag_lock:
            if _fw_diag["running"]:
                return jsonify({"error": "이미 진단 중입니다"}), 409
            _fw_diag.update(running=True, total=0, done=0, ok=0,
                            message="시작 중", results=[])
        src = pcprofile.get_source_ip(db_path)
        data = request.get_json(silent=True) or {}
        raw_ids = data.get("ids")
        if raw_ids is not None and not isinstance(raw_ids, list):
            return jsonify({"error": "ids는 정수 배열이어야 합니다"}), 400
        ids = None
        if raw_ids:
            # 전부 무효면 조용히 "전체 진단"으로 확대되던 fail-open을 막는다.
            ids = [int(i) for i in raw_ids if str(i).isdigit()]
            if not ids:
                return jsonify({"error": "ids에 유효한 정수가 없습니다"}), 400
            if len(ids) > 1000:
                return jsonify({"error": "한 번에 최대 1000대까지 지정할 수 있습니다"}), 400
        threading.Thread(target=_run_diagnose_all_firewalls,
                         args=(db_path, src, ids), daemon=True).start()
        log_event("info", "firewalls_diagnose_all_started", count=len(ids or []))
        return jsonify({"ok": True}), 202

    @app.route("/api/firewalls/diagnose-all/status", methods=["GET"])
    def diagnose_all_firewalls_status():
        with _fw_diag_lock:
            return jsonify(dict(_fw_diag))

    @app.route("/api/firewalls/collect-all/status", methods=["GET"])
    def collect_all_firewalls_status():
        with _fw_all_lock:
            return jsonify(dict(_fw_all))

    @app.route("/api/firewalls/collect-all/stop", methods=["POST"])
    def collect_all_firewalls_stop():
        with _fw_all_lock:
            if not _fw_all.get("running"):
                return jsonify({"ok": False, "error": "진행 중인 수집이 없습니다"}), 400
            _fw_all["stop"] = True
            _fw_all["message"] = "중지 요청됨 — 마무리 중…"
        return jsonify({"ok": True})

    @app.route("/api/servers/collect-all/status", methods=["GET"])
    def collect_all_servers_status():
        from core import server_collector
        return jsonify(server_collector.get_progress())

    @app.route("/api/servers/collect-all/stop", methods=["POST"])
    def collect_all_servers_stop():
        from core import server_collector
        return jsonify({"ok": server_collector.request_stop()})

    @app.route("/api/switches/<int:switch_id>/events", methods=["GET"])
    def get_switch_events(switch_id):
        """스위치의 포트 이벤트(flapping/looping) 조회."""
        try:
            events = db.get_port_events(db_path, switch_id)
            return jsonify({"switch_id": switch_id, "events": events})
        except Exception as e:
            sanitized = collector._sanitize_error_msg(str(e))
            log_event("error", "get_events_error", switch_id=switch_id, error=sanitized)
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/switches/bulk-collect", methods=["POST"])
    @rate_limit("bulk_collect", max_requests=10, window_seconds=60)
    def bulk_collect_endpoint():
        """공통 계정으로 선택된 스위치들을 일괄(비동기 동시) 수집.

        body: {ids:[...], username, password, persist?, enable_secret?}
        각 스위치를 워커 큐에 넣어 동시 수집한다. 계정은 세션 저장소 경유(평문 큐 비노출).
        """
        try:
            data = request.get_json(silent=True) or {}
            ids = data.get("ids", [])
            if not isinstance(ids, list) or not ids:
                return jsonify({"error": "ids required"}), 400
            if len(ids) > 500:
                return jsonify({"error": "too many ids"}), 400
            try:
                username = validate_credential(data.get("username"))
                password = validate_credential(data.get("password"))
                enable_secret = validate_credential(data.get("enable_secret")) \
                    if data.get("enable_secret") else None
            except ValueError as ve:
                return jsonify({"error": str(ve)}), 400
            # 요청에 계정이 없으면 이 세션에 보관된 수집 계정 사용(메모리·TTL)
            if not (username and password):
                _sc = _session_cred("switch")
                if _sc:
                    username, password = _sc
            if not config.app.get("demo_mode") and (not username or not password):
                return jsonify({"error": "username and password required"}), 400

            persist = data.get("persist", False)
            if persist:
                # 일괄 수집 계정도 이 PC 프로필에 등록(1회)
                try:
                    pcprofile.save_profile(db_path, username, password,
                                           source_ip=pcprofile.get_source_ip(db_path))
                except Exception as e:
                    log_event("warning", "pc_profile_save_failed",
                              error=collector._sanitize_error_msg(str(e)))
            queued, skipped = [], []
            allowed = config.collector.get("allowed_ip_ranges")
            for raw in ids:
                try:
                    sid = int(raw)
                except (TypeError, ValueError):
                    continue
                sw = db.get_switch(db_path, sid)
                if not sw:
                    skipped.append({"id": sid, "reason": "not found"})
                    continue
                # SSRF 방어: DB에 저장된 IP도 수집 직전 재검증
                ip = sw.get("ip") if isinstance(sw, dict) else getattr(sw, "ip", None)
                if ip:
                    try:
                        validate_ipv4(ip, allowed)
                    except ValueError as e:
                        skipped.append({"id": sid, "reason": "ip rejected: %s" % e})
                        continue
                result = collector.collect_switch(db_path, sid, username, password,
                                                  enable_secret=enable_secret)
                if result.get("status") == "queued":
                    queued.append(sid)
                    if persist:
                        cred_blob = credentials.encrypt_credential(username, password)
                        if cred_blob:
                            try:
                                db.update_cred_blob(db_path, sid, cred_blob)
                            except Exception:
                                pass
                        if enable_secret:
                            es_blob = credentials.encrypt_text(enable_secret)
                            if es_blob:
                                try:
                                    db.set_setting(db_path, "enable_secret_%d" % sid, es_blob)
                                except Exception:
                                    pass
                else:
                    skipped.append({"id": sid, "reason": result.get("message", "enqueue failed")})
            log_event("info", "bulk_collect", queued=len(queued), skipped=len(skipped))
            # 진행바·중지에 쓸 배치 정보 기록(스위치만 진행 표시가 없어 200대를 걸면
            # alert 한 번이 전부였고, 진척 확인도 중단도 불가능했다)
            with _sw_bulk_lock:
                _sw_bulk.update(ids=list(queued), total=len(queued), started=True)
            return jsonify({"ok": True, "queued": queued, "skipped": skipped,
                            "queued_count": len(queued), "skipped_count": len(skipped)}), 202
        except Exception as e:
            log_event("error", "bulk_collect_error", error=collector._sanitize_error_msg(str(e)))
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/switches/<int:switch_id>/collect", methods=["POST"])
    @rate_limit("collect_switch", max_requests=30, window_seconds=60)
    def collect_switch_endpoint(switch_id):
        log_event("info", "collect_requested", switch_id=switch_id)

        try:
            data = request.get_json(silent=True) or {}

            # HIGH FIX (CWE-20): Validate credential string length, type, character set
            try:
                username = validate_credential(data.get("username"))
                password = validate_credential(data.get("password"))
                # enable secret(선택): enable 비밀번호가 로그인과 다른 장비만 입력
                enable_secret = validate_credential(data.get("enable_secret")) \
                    if data.get("enable_secret") else None
            except ValueError as validation_error:
                log_event("warning", "collect_invalid_credentials", switch_id=switch_id, reason=str(validation_error))
                return jsonify({"error": str(validation_error)}), 400

            # 요청에 계정이 없으면 이 세션에 보관된 수집 계정 사용(메모리·TTL)
            if not (username and password):
                _sc = _session_cred("switch")
                if _sc:
                    username, password = _sc

            # CWE-522 fix: Require credentials in production mode; sanitize log output
            if not config.app.get("demo_mode") and (not username or not password):
                return jsonify({"error": "username and password required"}), 400

            # HARDENING (CWE-918 SSRF): Validate DB switch IP before collection.
            # Prevents legacy/seed data with public IPs from bypassing input validation.
            switch_row = db.get_switch(db_path, switch_id)
            if not switch_row:
                # 존재하지 않는 스위치는 즉시 404(이전엔 워커에 큐잉돼 202 반환 후 실패)
                return jsonify({"error": "not found"}), 404
            switch_ip = switch_row.get("ip") if isinstance(switch_row, dict) else getattr(switch_row, "ip", None)
            if switch_ip:
                try:
                    validate_ipv4(switch_ip, config.collector.get("allowed_ip_ranges"))
                except ValueError as e:
                    log_event("warning", "collect_blocked_invalid_ip", switch_id=switch_id, ip=switch_ip, reason=str(e))
                    return jsonify({"error": f"Switch IP rejected: {e}"}), 400

            persist = data.get("persist", False)

            # M5 (CWE-362 race fix): collect_switch owns session storage. It stores
            # the credential ONLY after passing the in-progress check, so a duplicate
            # request can never overwrite or clear an active job's credential. It
            # also disposes the credential itself on any enqueue failure. The async
            # worker loads it at moment-of-use and clears it when collection finishes.
            result = collector.collect_switch(db_path, switch_id, username, password,
                                              enable_secret=enable_secret)

            # M5 (W3): Map the async submission outcome to an accurate HTTP status.
            status = result.get("status")
            if status == "queued":
                # M5 R2: Persist the DPAPI blob ONLY after a successful enqueue, so a
                # duplicate request (rejected as in-progress below) can never overwrite
                # an active switch's persisted credential blob.
                if persist:
                    cred_blob = credentials.encrypt_credential(username, password)
                    if cred_blob:
                        try:
                            db.update_cred_blob(db_path, switch_id, cred_blob)
                            log_event("info", "credential_persisted", switch_id=switch_id)
                        except Exception as e:
                            sanitized = _sanitize_error_msg(str(e))
                            log_event("warning", "credential_persist_failed", switch_id=switch_id, error=sanitized)
                    # 이 PC 프로필(MAC·IP·계정)에도 등록 — 이 PC가 수집할 때
                    # 자기 계정/IP로 시도(스위치별 blob이 다른 PC 것이어도 폴백 가능)
                    try:
                        pcprofile.save_profile(db_path, username, password,
                                               source_ip=pcprofile.get_source_ip(db_path))
                    except Exception as e:
                        log_event("warning", "pc_profile_save_failed",
                                  error=_sanitize_error_msg(str(e)))
                    # enable secret도 함께 영속화(별도 blob — 자격증명 형식과 분리).
                    # 저장돼 있으면 다음 수집(자동수집 포함)부터 자동 사용.
                    if enable_secret:
                        es_blob = credentials.encrypt_text(enable_secret)
                        if es_blob:
                            try:
                                db.set_setting(db_path, "enable_secret_%d" % switch_id, es_blob)
                                log_event("info", "enable_secret_persisted", switch_id=switch_id)
                            except Exception as e:
                                log_event("warning", "enable_secret_persist_failed",
                                          switch_id=switch_id, error=_sanitize_error_msg(str(e)))
                return jsonify(result), 202
            if "already being collected" in result.get("message", ""):
                return jsonify(result), 409  # Conflict: collection in progress
            return jsonify(result), 503  # Service Unavailable: queue full / enqueue failed
        except Exception as e:
            # CWE-532 fix: Sanitize error messages to prevent credential/path exposure in logs
            sanitized_error = _sanitize_error_msg(str(e))
            log_event("error", "collect_error", switch_id=switch_id, error_type=type(e).__name__, error=sanitized_error)
            return jsonify({"error": "Internal server error"}), 500

    @app.route("/api/switches/<int:switch_id>/detail", methods=["GET"])
    def get_switch_detail(switch_id):
        log_event("info", "detail_requested", switch_id=switch_id)

        try:
            switch = db.get_switch(db_path, switch_id)
            if not switch:
                return jsonify({"error": "Switch not found"}), 404

            from core import tps_location
            _info = tps_location.parse(switch.get("hostname"))
            if _info:
                switch["tps_location"] = _info["label"]

            ports = db.get_ports_by_switch(db_path, switch_id)
            macs = db.get_mac_entries_by_switch(db_path, switch_id)
            arps = db.get_arp_entries_by_switch(db_path, switch_id)
            hosts = db.get_hosts_by_switch(db_path, switch_id)

            # show logging 분석 결과(최근 로그 + 탐지 이벤트)
            logs = None
            raw_logs = db.get_switch_logs(db_path, switch_id)
            if raw_logs:
                import json as _json
                try:
                    events = _json.loads(raw_logs.get("events_json") or "[]")
                except (ValueError, TypeError):
                    events = []
                logs = {
                    "recent": (raw_logs.get("recent_lines") or "").split("\n"),
                    "events": events,
                    "alert": raw_logs.get("log_alert") or "none",
                    "updated": raw_logs.get("updated"),
                }

            return jsonify({
                "switch": switch,
                "ports": ports,
                "macs": macs,
                "arps": arps,
                "hosts": hosts,
                "logs": logs
            })
        except Exception as e:
            # CWE-532 fix: Sanitize error messages to prevent credential/path exposure in logs
            sanitized_error = _sanitize_error_msg(str(e))
            log_event("error", "detail_error", switch_id=switch_id, error_type=type(e).__name__, error=sanitized_error)
            return jsonify({"error": "Internal server error"}), 500

    @app.errorhandler(404)
    def not_found(e):
        return jsonify({"error": "Not found"}), 404

    @app.errorhandler(500)
    def internal_error(e):
        # CWE-532 fix: Sanitize error messages to prevent credential/path exposure in logs
        sanitized_error = _sanitize_error_msg(str(e))
        log_event("error", "internal_error", error_type=type(e).__name__, error=sanitized_error)
        return jsonify({"error": "Internal server error"}), 500

    # ── 웹 SSH 터미널(WebSocket) — 장비 클릭 → 브라우저에서 PuTTY처럼 CLI ──
    try:
        from flask_sock import Sock
        from core import webshell
        sock = Sock(app)

        def _ws_same_origin():
            """CSWSH(교차 사이트 WebSocket 하이재킹) 방어.

            WebSocket에는 동일 출처 정책이 적용되지 않는다 — 브라우저는 크로스
            오리진 연결도 프리플라이트 없이 보낸다. Origin 검사가 없으면 운영자가
            연 아무 웹페이지나 ws://127.0.0.1:8082/ws/shell/<id> 를 열어 **저장된
            장비 계정으로 실제 SSH 세션**을 잡을 수 있다(로컬 면제 탓에 토큰도 불필요).

            Origin이 없는 요청(브라우저가 아닌 클라이언트)은 아래 토큰 검사에 맡긴다.
            """
            origin = request.headers.get("Origin")
            if not origin:
                return True
            host = request.host or ""
            return origin in ("http://" + host, "https://" + host)

        def _ws_authorized(token):
            """WS 인증: 로컬 루프백은 면제, 원격은 토큰 검증(before_request와 동일 정책)."""
            if config.app.get("demo_mode") and not config.api_token:
                return True
            bind_host = config.app.get("host", "127.0.0.1")
            if bind_host in ("127.0.0.1", "localhost", "::1") and request.remote_addr in ("127.0.0.1", "::1"):
                return True
            expected = config.api_token
            # bytes 비교: 비ASCII 토큰이 오면 compare_digest가 TypeError를 던져
            # try 밖에서 터진다(HTTP 경로는 이미 같은 이유로 encode 처리됨).
            return bool(expected and token and hmac.compare_digest(
                token.encode("utf-8", "replace"), expected.encode("utf-8", "replace")))

        def _run_ws_shell(ws, kind, target_id):
            token = request.args.get("token", "")
            if not _ws_same_origin():
                log_event("warning", "webshell_bad_origin",
                          origin=str(request.headers.get("Origin"))[:80])
                try:
                    ws.send("\r\n[NetDash] 허용되지 않은 출처입니다.\r\n")
                except Exception:
                    pass
                return
            if not _ws_authorized(token):
                try:
                    ws.send("\r\n[NetDash] 인증 실패(토큰).\r\n")
                except Exception:
                    pass
                return
            # 자격증명은 저장된 것(세션/DPAPI)만 사용 — 쿼리 파라미터로 받지 않는다
            # (URL 쿼리는 access log/프록시 로그에 남을 수 있어 자격증명 노출 위험).
            src = pcprofile.get_source_ip(db_path)
            ranges = config.collector.get("allowed_ip_ranges")

            def _vip(ip):
                return validate_ipv4(ip, allowed_ip_ranges=ranges)
            try:
                webshell.run_shell(ws, db_path, target_id, "", "", src,
                                   validate_ip=_vip, client_ip=_client_ip(), kind=kind)
            except Exception as e:
                log_event("error", "webshell_error", error=collector._sanitize_error_msg(str(e)))

        @sock.route("/ws/shell/<int:switch_id>")
        def ws_shell(ws, switch_id):
            _run_ws_shell(ws, "switch", switch_id)

        @sock.route("/ws/shell/<kind>/<int:target_id>")
        def ws_shell_kind(ws, kind, target_id):
            """스위치·방화벽·서버 공통 터미널. 종류가 이상하면 스위치로 떨어뜨리지 않고 거부."""
            if kind not in webshell.KINDS:
                try:
                    ws.send("\r\n[NetDash] 알 수 없는 장비 종류입니다.\r\n")
                except Exception:
                    pass
                return
            _run_ws_shell(ws, kind, target_id)
        log_event("info", "webshell_enabled")
    except Exception as e:
        # flask-sock 미설치 등 → 터미널 비활성(나머지 기능은 정상)
        log_event("warning", "webshell_unavailable", error=str(e))

    return app


def _open_browser_when_ready(url, port):
    """서버 기동(포트 LISTEN)을 확인한 뒤 기본 브라우저를 연다.

    headless 웹앱(console=False)이라 더블클릭 시 화면이 안 뜨는 문제를 해결한다.
    별도 스레드에서 포트를 폴링하므로 app.run()을 막지 않는다.
    """
    import socket
    import time
    import webbrowser

    for _ in range(60):  # 최대 ~30초 대기
        try:
            with socket.create_connection(("127.0.0.1", port), timeout=0.5):
                break
        except OSError:
            time.sleep(0.5)
    try:
        webbrowser.open(url)
    except Exception:
        pass


def _warm_openpyxl():
    """onefile exe 첫 /api/report ~15초 지연 해소(성능 이슈 후속).

    openpyxl 모듈은 기동 시 import되지만, 첫 Workbook 저장 시점에 lazy 로드되는
    writer/serializer 서브모듈이 onefile 압축 해제·로드로 오래 걸린다.
    기동 직후 백그라운드에서 미니 워크북을 한 번 저장해 미리 데운다.
    실패해도 무해 — 첫 보고서 요청이 기존처럼 로드할 뿐이다.
    """
    try:
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        wb.active["A1"] = "warmup"
        buf = io.BytesIO()
        wb.save(buf)
        log_event("info", "openpyxl_warmed")
    except Exception:
        pass


if __name__ == "__main__":
    import traceback
    _safe_stdout()          # 콘솔 인코딩 오류로 배너가 프로세스를 죽이지 않게
    try:
        parser = argparse.ArgumentParser(description="NetDash - Network switch current status dashboard")
        parser.add_argument("--demo", action="store_true", help="Run in demo mode with sample data")
        args = parser.parse_args()

        # CLI --demo flag takes precedence over DEMO_MODE environment variable
        demo_mode = args.demo if args.demo else (os.getenv("DEMO_MODE", "").lower() == "true")

        # Create config with determined demo_mode
        reset_config()
        try:
            config = get_config(demo_mode=demo_mode)
        except (ValueError, TypeError, AttributeError, RuntimeError) as _cfg_err:
            # 설정 오류는 사용자가 고칠 수 있는 문제다. 파이썬 트레이스백을 던지면
            # 콘솔이 닫히며 아무것도 못 보고 끝난다 — 원인과 조치를 한글로 안내한다.
            # (특히 원격 접속용 host: 0.0.0.0 만 바꾸고 api_token을 빠뜨리는 경우)
            # ValueError 외에도 섹션을 빈 값으로 둔 편집(`collector:` 만 남김 →
            # TypeError)·YAML 문법 오류(RuntimeError)가 흔해 함께 잡는다.
            _m = str(_cfg_err)
            print("=" * 60)
            print("  NetDash 를 시작할 수 없습니다 - 설정(config.yaml) 문제")
            print("-" * 60)
            if "api_token" in _m:
                print("  원인: 외부 접속(host: 0.0.0.0)으로 설정했는데")
                print("        접속 토큰(api_token)이 비어 있습니다.")
                print()
                print("  조치: config.yaml 을 열어 아래처럼 32자 이상 토큰을 넣으세요.")
                print('        api_token: "여기에-32자-이상의-임의-문자열"')
                print()
                print("  참고: host 를 127.0.0.1 로 두면 토큰 없이 이 PC에서만 씁니다.")
            else:
                print("  원인: " + _m)
                print()
                print("  조치: config.yaml 의 해당 항목을 확인하세요.")
            print("=" * 60, flush=True)
            log_event("error", "app_start_config_error", reason=_m[:200])
            try:
                input("계속하려면 Enter 를 누르세요...")
            except (EOFError, OSError):
                pass
            sys.exit(1)

        # 단일 인스턴스 보장 — DB를 열기 전에(create_app 이전) 검사해야 한다.
        # 다른 PC가 같은 공유폴더의 exe/DB로 이미 실행 중이면 SQLite(WAL)가
        # 다중 호스트 접근을 지원하지 않아 db_error가 나므로, 명확히 안내 후 종료.
        from core import instance_lock
        from core.config_loader import get_data_dir
        _lk_host = config.app.get("host", "127.0.0.1")
        _lk_port = config.app.get("port", 8082) or 8082
        _lk_open_host = "127.0.0.1" if _lk_host in ("0.0.0.0", "::") else _lk_host
        # 잠금은 **DB가 있는 폴더**에 건다. 예전엔 exe 폴더(get_data_dir())에 걸었는데,
        # 각 PC가 자기 로컬에 exe를 두고 db_path만 공유 절대경로로 가리키면
        # (문서상 자연스러운 구성) 서로의 잠금을 보지 못해 **두 PC가 동시에 주
        # 서버가 되어 같은 SQLite에 쓴다** — 이 모듈이 막으려던 바로 그 상황이다.
        try:
            _lock_dir = config.get_db_path().parent
        except Exception:
            _lock_dir = get_data_dir()
        _acquired, _other = instance_lock.acquire(
            _lock_dir, url=f"http://{_lk_open_host}:{_lk_port}")
        _readonly_info = None
        if not _acquired:
            # 주 서버가 다른 곳에서 실행 중 → 종료하지 않고 '읽기 전용'으로 기동.
            # 조회는 자유롭게, 수집/수정 시도 시 UI에 안내 메시지가 뜬다.
            _readonly_info = _other or {}
            _o_host = _readonly_info.get("hostname", "알 수 없음")
            print("=" * 56)
            print("  NetDash가 이미 다른 곳(" + str(_o_host) + ")에서 실행 중입니다.")
            print("  이 프로그램은 [읽기 전용 모드]로 시작합니다.")
            print("-" * 56)
            print("  - 현황 조회: 가능 (수집된 정보 실시간 확인)")
            print("  - 정보 수집·수정: 불가 - 주 서버(" + str(_o_host) + ")에서 하세요.")
            print("  - 주 서버가 종료되면 이 프로그램이 자동으로 주 서버가 됩니다.")
            print("=" * 56, flush=True)
            log_event("warning", "app_start_readonly",
                      other_host=str(_o_host))

        # 주의: _readonly_info가 빈 dict여도(정보 파일 유실) 읽기 전용 + 감시는 켠다
        app = create_app(demo_mode=demo_mode, readonly_info=_readonly_info,
                         promote_watch=_readonly_info is not None)

        # CWE-306 fix: In production mode, API token MUST be configured.
        # (config_loader auto-generates a token for loopback binds; this is a safety net.)
        if not demo_mode and not config.api_token:
            api_token_env = os.getenv("API_TOKEN", "")
            if not api_token_env:
                log_event("error", "app_start_failed", reason="API_TOKEN required in production mode")
                raise RuntimeError("API_TOKEN environment variable required in production mode")
            config.api_token = api_token_env

        host = config.app.get("host", "127.0.0.1")
        port = config.app.get("port", 8082)
        # CRITICAL FIX (CWE-489): In production mode, force debug=False to prevent credential/stack-trace exposure.
        # Do NOT allow debug override via environment variables in production (app.run() receives final value here).
        debug = config.app.get("debug", False) and demo_mode

        log_event("info", "app_start", host=host, port=port, debug=debug, demo_mode=demo_mode)

        open_host = "127.0.0.1" if host in ("0.0.0.0", "::") else host
        url = f"http://{open_host}:{port}"

        # 가시적 콘솔 배너 (console=True): 사용자가 실행 상태/접속 주소/종료법을 명확히 인지.
        mode_label = "데모" if demo_mode else "운영"
        print("=" * 56)
        print("  NetDash 가 시작되었습니다.  (모드: " + mode_label + ")")
        print("  접속 주소:  " + url)
        print("  종료하려면 이 창을 닫거나 Ctrl+C 를 누르세요.")
        print("=" * 56, flush=True)

        # 편의: 서버가 뜨면 브라우저를 자동으로 연다. 콘솔이 떠 있으므로
        # 백그라운드 은닉이 아니라 보조 기능이다(자동으로 안 열려도 위 주소로 접속).
        browser_thread = threading.Thread(
            target=_open_browser_when_ready,
            args=(url, port),
            daemon=True,
        )
        browser_thread.start()

        # 첫 /api/report 지연(openpyxl lazy 서브모듈) 예열 — 백그라운드, 실패 무해
        threading.Thread(target=_warm_openpyxl, daemon=True).start()

        # threaded=True: 웹 SSH 터미널(WebSocket)이 요청당 스레드를 점유하므로 필수
        app.run(host=host, port=port, debug=debug, threaded=True)
    except Exception:
        # console=False(windowed) exe에서는 콘솔에 트레이스백이 보이지 않으므로
        # 데이터 디렉터리(exe 폴더 또는 LOCALAPPDATA 폴백)에 에러 로그를 남긴다.
        # cwd는 관리자 실행 시 System32라 쓰기 실패할 수 있어 사용하지 않는다.
        try:
            from core.config_loader import get_data_dir
            _err_path = get_data_dir() / "netdash_error.log"
        except Exception:
            _err_path = Path("netdash_error.log")
        try:
            with open(_err_path, "w", encoding="utf-8") as _f:
                _f.write(traceback.format_exc())
        except OSError:
            pass
        raise
