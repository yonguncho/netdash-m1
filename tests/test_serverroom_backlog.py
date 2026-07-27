# -*- coding: utf-8 -*-
"""백로그 B-4·B-13·B-14 — 서버실 배치도/카드 결함."""
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

from core import db, serverroom

ROOT = Path(__file__).parent.parent
APPJS = (ROOT / "web" / "static" / "app.js").read_text(encoding="utf-8")
CSS = (ROOT / "web" / "static" / "style.css").read_text(encoding="utf-8")


@pytest.fixture
def cli(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    from config import reset_config
    reset_config()
    from app import create_app
    app = create_app(demo_mode=True)
    app.config["TESTING"] = True
    with app.test_client() as c:
        yield c


# ── B-4: 랙배치 엑셀에 서버가 빠져 있었다 ─────────────────────────
def test_rack_xlsx_includes_physical_servers(cli):
    p = Path.cwd() / "netdash.db"
    sw = db.save_switch(p, "RACK-SW", "10.80.0.1", "cisco_ios")
    db.update_switch(p, sw, location="A09U10")
    sid = db.save_server(p, "RACK-SRV", "10.80.0.2", location="A09U27")
    db.save_server(p, "RACK-VM", "10.80.0.3", location="A09U28", is_vm=1)

    r = cli.get("/api/serverroom/export")
    assert r.status_code == 200, r.get_data(as_text=True)[:200]
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.get_data()))
    text = "\n".join(
        str(c.value) for row in wb.active.iter_rows() for c in row if c.value)
    assert "RACK-SW" in text
    assert "RACK-SRV" in text, "물리 서버가 랙배치 엑셀에서 빠졌다"
    assert "RACK-VM" not in text, "VM은 물리 위치가 없으므로 제외돼야 한다"
    assert sid


def test_rack_xlsx_still_404_when_empty(cli):
    assert cli.get("/api/serverroom/export").status_code == 404


# ── B-14: 같은 랙·유닛 중복이 조용히 사라졌다 ─────────────────────
def test_duplicate_unit_is_not_silently_dropped():
    devices = [
        {"name": "SW-A", "ip": "10.1.0.1", "rack": "A09", "unit": 27,
         "device_type": "L2 Switch"},
        {"name": "SRV-B", "ip": "10.1.0.2", "rack": "A09", "unit": 27,
         "device_type": "Server"},
    ]
    data = serverroom.build_rack_xlsx(devices)
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data))
    text = "\n".join(
        str(c.value) for row in wb.active.iter_rows() for c in row if c.value)
    assert "SW-A" in text and "SRV-B" in text, \
        "같은 유닛의 장비 하나가 조용히 사라졌다: %s" % text[:200]


def test_normal_units_unaffected():
    devices = [
        {"name": "SW-A", "ip": "10.1.0.1", "rack": "A09", "unit": 27},
        {"name": "SW-B", "ip": "10.1.0.2", "rack": "A09", "unit": 28},
    ]
    data = serverroom.build_rack_xlsx(devices)
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data))
    text = "\n".join(
        str(c.value) for row in wb.active.iter_rows() for c in row if c.value)
    assert "SW-A" in text and "SW-B" in text
    assert "SW-A + SW-B" not in text, "중복이 아닌데 합쳐졌다"


# ── B-13: 서버 카드 상태 색이 안 나왔다 ───────────────────────────
def test_server_card_uses_defined_status_classes():
    """CSS에 없는 클래스를 쓰면 컬러바·상태 dot이 아예 안 보인다."""
    i = APPJS.index("function _srvCardHTML(")
    block = APPJS[i:i + 900]
    assert '"done"' not in block, "CSS에 없는 sw-card--done 을 쓰고 있다"
    for cls in ("critical", "collecting", "new", "ok"):
        assert '"%s"' % cls in block, cls
        assert ".sw-card--%s" % cls in CSS or cls == "ok", cls
    assert ".sw-card--ok" in CSS and ".dot--ok" in CSS


def test_uncollected_server_card_is_not_shown_as_normal():
    i = APPJS.index("function _srvCardHTML(")
    block = APPJS[i:i + 900]
    assert 's.status === "new"' in block, "미수집 서버가 정상 카드로 보인다"
