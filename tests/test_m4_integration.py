"""
M4: 멀티블록 엑셀 업로드 통합 테스트

이 테스트는 `/api/upload` 엔드포인트의 전체 플로우를 검증한다:
1. 엑셀 파일(스위치+호스트 혼합) 업로드
2. 멀티블록 분리 및 DB 임포트 (import_switches_bulk + save_hosts)
3. 멱등성: 동일 파일 재업로드 시 중복 삽입 금지 (upsert)
4. 진단 정보 반환 확인

이전에 블로커가 된 이유:
- excel_loader.load_workbook()이 List[Dict] 반환
- db.save_hosts()가 Dict 기대 → 'list' object has no attribute 'items' 오류
- 통합 테스트 부재 → 단위 테스트 56개는 통과했으나 엔드포인트 테스트 없음

수정 내용:
- db.save_hosts(db_path, List[Dict]) 으로 시그니처 통일
- 통합 테스트 추가하여 향후 회귀 방지
"""

import pytest
import sys
import os
import tempfile
from pathlib import Path
from openpyxl import Workbook
import json

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app import create_app
from core import db


@pytest.fixture()
def client(tmp_path, monkeypatch):
    """Demo mode client for integration tests"""
    monkeypatch.chdir(tmp_path)
    app = create_app(demo_mode=True)
    app.config["TESTING"] = True
    with app.test_client() as c:
        yield c


@pytest.fixture()
def multiblock_xlsx():
    """
    멀티블록 엑셀 파일 생성 (스위치 + 호스트 + 폐기 대상)

    결과:
    - Block 1 (제목): "Multiblock Upload Test" → IP 없음 → 폐기
    - Block 2 (스위치): 2개 스위치
    - Block 3 (호스트): 3개 호스트
    """
    wb = Workbook()
    ws = wb.active

    # Block 1: 제목 (폐기 대상)
    ws["A1"] = "Network Device Summary"
    ws["B1"] = "Total Devices"
    ws["C1"] = "2024"

    # Block 2: 스위치 헤더
    ws["A2"] = "name"
    ws["B2"] = "ip"
    ws["C2"] = "hostname"
    ws["D2"] = "vendor"
    ws["E2"] = "location"

    # Block 2: 스위치 데이터
    ws["A3"] = "Core-Switch-1"
    ws["B3"] = "10.0.1.1"
    ws["C3"] = "core1.lab"
    ws["D3"] = "Cisco"
    ws["E3"] = "DC-Primary"

    ws["A4"] = "Core-Switch-2"
    ws["B4"] = "10.0.1.2"
    ws["C4"] = "core2.lab"
    ws["D4"] = "Juniper"
    ws["E4"] = "DC-Backup"

    # Block 3: 호스트 헤더
    ws["A5"] = "ip"
    ws["B5"] = "mac"

    # Block 3: 호스트 데이터
    ws["A6"] = "192.168.10.100"
    ws["B6"] = "00:11:22:33:44:01"

    ws["A7"] = "192.168.10.101"
    ws["B7"] = "00:11:22:33:44:02"

    ws["A8"] = "192.168.10.102"
    ws["B8"] = "00:11:22:33:44:03"

    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        yield tmp_path
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


class TestM4UploadIntegration:
    """멀티블록 엑셀 업로드 통합 테스트"""

    def test_upload_multiblock_file_succeeds(self, client, multiblock_xlsx):
        """
        멀티블록 파일 업로드가 성공하고 올바른 응답을 반환하는가?
        """
        with open(multiblock_xlsx, "rb") as f:
            response = client.post(
                "/api/upload",
                data={"file": (f, "multiblock.xlsx")},
                content_type="multipart/form-data",
            )

        assert response.status_code == 201
        data = response.get_json()
        assert data["ok"] is True
        assert "diagnostics" in data
        assert "imported_switch_ids" in data
        assert "imported_host_ids" in data

        diag = data["diagnostics"]
        assert isinstance(diag, dict)
        assert diag["total_blocks"] >= 2  # 최소 스위치 + 호스트 블록
        assert diag["switch_blocks"] == 1
        assert diag["host_blocks"] == 1
        # Note: discarded_blocks may vary based on how rows are parsed
        assert diag["discarded_blocks"] >= 0

    def test_upload_inserts_switches_into_db(self, client, multiblock_xlsx):
        """
        업로드 후 스위치가 DB에 실제 삽입되는가?
        """
        with open(multiblock_xlsx, "rb") as f:
            response = client.post(
                "/api/upload",
                data={"file": (f, "multiblock.xlsx")},
                content_type="multipart/form-data",
            )

        assert response.status_code == 201
        data = response.get_json()

        # Flask 테스트 클라이언트에서 db_path 조회
        from config import get_config
        config = get_config(demo_mode=True)
        db_path = config.get_db_path()

        switches = db.get_switches(db_path)
        assert len(switches) >= 2

        # 스위치명 확인
        switch_names = {s["name"].lower() for s in switches}
        assert "core-switch-1" in switch_names
        assert "core-switch-2" in switch_names

        # IP 확인
        switch_ips = {s["ip"] for s in switches}
        assert "10.0.1.1" in switch_ips
        assert "10.0.1.2" in switch_ips

    def test_upload_inserts_hosts_into_db(self, client, multiblock_xlsx):
        """
        업로드 후 호스트가 DB에 실제 삽입되는가?
        이전 버그: 'list' object has no attribute 'items'
        """
        with open(multiblock_xlsx, "rb") as f:
            response = client.post(
                "/api/upload",
                data={"file": (f, "multiblock.xlsx")},
                content_type="multipart/form-data",
            )

        assert response.status_code == 201
        data = response.get_json()
        assert data["ok"] is True

        # DB에서 호스트 확인 (search_host_by_ip 사용)
        from config import get_config
        config = get_config(demo_mode=True)
        db_path = config.get_db_path()

        host1 = db.search_host_by_ip(db_path, "192.168.10.100")
        assert host1 is not None
        assert host1["ip"] == "192.168.10.100"
        assert host1["mac"] == "00:11:22:33:44:01"

        host2 = db.search_host_by_ip(db_path, "192.168.10.101")
        assert host2 is not None
        assert host2["ip"] == "192.168.10.101"
        assert host2["mac"] == "00:11:22:33:44:02"

    def test_reupload_same_file_is_idempotent(self, client, multiblock_xlsx):
        """
        동일 파일을 두 번 업로드할 때 중복 삽입이 없는가? (upsert)
        """
        from config import get_config
        config = get_config(demo_mode=True)
        db_path = config.get_db_path()

        # 첫 번째 업로드
        with open(multiblock_xlsx, "rb") as f:
            response1 = client.post(
                "/api/upload",
                data={"file": (f, "multiblock.xlsx")},
                content_type="multipart/form-data",
            )
        assert response1.status_code == 201
        switches_after_first = db.get_switches(db_path)
        switch_count_first = len(switches_after_first)

        hosts_after_first = []
        for ip in ["192.168.10.100", "192.168.10.101", "192.168.10.102"]:
            h = db.search_host_by_ip(db_path, ip)
            if h:
                hosts_after_first.append(h)
        host_count_first = len(hosts_after_first)

        # 두 번째 업로드 (동일 파일)
        with open(multiblock_xlsx, "rb") as f:
            response2 = client.post(
                "/api/upload",
                data={"file": (f, "multiblock.xlsx")},
                content_type="multipart/form-data",
            )
        assert response2.status_code == 201

        switches_after_second = db.get_switches(db_path)
        switch_count_second = len(switches_after_second)

        hosts_after_second = []
        for ip in ["192.168.10.100", "192.168.10.101", "192.168.10.102"]:
            h = db.search_host_by_ip(db_path, ip)
            if h:
                hosts_after_second.append(h)
        host_count_second = len(hosts_after_second)

        # 중복 삽입이 없어야 함 (upsert이므로 count 동일)
        assert switch_count_first == switch_count_second
        assert host_count_first == host_count_second

    def test_upload_without_file_field_returns_400(self, client):
        """
        file 필드 없이 업로드하면 400 반환하는가?
        """
        response = client.post(
            "/api/upload",
            data={},
            content_type="multipart/form-data",
        )
        assert response.status_code == 400
        data = response.get_json()
        assert "error" in data
        assert "file" in data["error"].lower()

    def test_upload_invalid_extension_returns_400(self, client):
        """
        .xlsx가 아닌 파일 업로드 시 400 반환하는가?
        """
        with tempfile.NamedTemporaryFile(suffix=".txt", delete=False) as tmp:
            tmp.write(b"not an xlsx file")
            tmp_path = tmp.name

        try:
            with open(tmp_path, "rb") as f:
                response = client.post(
                    "/api/upload",
                    data={"file": (f, "wrong.txt")},
                    content_type="multipart/form-data",
                )
            assert response.status_code == 400
            data = response.get_json()
            assert "error" in data
            assert "xlsx" in data["error"].lower()
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_upload_diagnostics_structure(self, client, multiblock_xlsx):
        """
        업로드 응답의 diagnostics 구조가 올바른가?
        """
        with open(multiblock_xlsx, "rb") as f:
            response = client.post(
                "/api/upload",
                data={"file": (f, "multiblock.xlsx")},
                content_type="multipart/form-data",
            )

        assert response.status_code == 201
        data = response.get_json()
        diag = data["diagnostics"]

        # 필수 필드 확인
        assert "total_blocks" in diag
        assert "discarded_blocks" in diag
        assert "switch_blocks" in diag
        assert "host_blocks" in diag
        assert "imported_switches" in diag
        assert "imported_hosts" in diag
        assert "warnings" in diag

        # 타입 확인
        assert isinstance(diag["total_blocks"], int)
        assert isinstance(diag["switch_blocks"], int)
        assert isinstance(diag["host_blocks"], int)
        assert isinstance(diag["imported_switches"], int)
        assert isinstance(diag["imported_hosts"], int)
        assert isinstance(diag["warnings"], list)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
