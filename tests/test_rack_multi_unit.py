# -*- coding: utf-8 -*-
"""랙 다중 유닛(U 높이) — 'A09U13-U15'처럼 여러 U를 차지하는 장비.

한 장비가 U13~U15를 차지하면 서버실 현황·엑셀 배치도에서 그만큼 크게 보이고,
랙뷰에서 아래 손잡이를 끌어 높이를 바꾸면 location에 저장된다.
"""
import io
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

from core import db, serverroom

ROOT = Path(__file__).parent.parent
APPJS = (ROOT / "web" / "static" / "app.js").read_text(encoding="utf-8")
CSS = (ROOT / "web" / "static" / "style.css").read_text(encoding="utf-8")


@pytest.fixture
def cli(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    from config import reset_config
    reset_config()
    from app import create_app
    app = create_app(demo_mode=True)
    app.config["TESTING"] = True
    with app.test_client() as c:
        yield c


# ── 파싱 ─────────────────────────────────────────────────────────
def test_single_unit_unchanged():
    r = serverroom.parse_rack("A09U27")
    assert r["rack"] == "A09" and r["unit"] == 27
    assert r["unit_end"] == 27 and r["height"] == 1
    assert r["label"] == "A09랙 U27"


@pytest.mark.parametrize("loc", ["A09U13-U15", "A09U13-15", "a09u13 - u15",
                                 "A09U13~U15", "A09U15-U13"])
def test_multi_unit_forms(loc):
    r = serverroom.parse_rack(loc)
    assert r["rack"] == "A09", loc
    assert r["unit"] == 13 and r["unit_end"] == 15 and r["height"] == 3, loc


def test_multi_unit_label_shows_range():
    assert serverroom.parse_rack("A09U13-U15")["label"] == "A09랙 U13-U15 (3U)"


def test_height_is_capped():
    r = serverroom.parse_rack("A01U1-U99")
    assert r["height"] == serverroom.MAX_HEIGHT
    assert r["unit_end"] == 1 + serverroom.MAX_HEIGHT - 1


def test_invalid_still_none():
    for bad in ("", None, "A09", "U27", "그냥위치", "A09U-U15"):
        assert serverroom.parse_rack(bad) is None, bad


def test_occupied_units():
    assert serverroom.occupied_units(serverroom.parse_rack("A09U13-U15")) == [13, 14, 15]
    assert serverroom.occupied_units(serverroom.parse_rack("A09U27")) == [27]


def test_format_rack_roundtrip():
    assert serverroom.format_rack("A09", 13, 3) == "A09U13-U15"
    assert serverroom.format_rack("a09", 27, 1) == "A09U27"
    assert serverroom.parse_rack(serverroom.format_rack("B02", 5, 4))["height"] == 4


# ── API 응답 ─────────────────────────────────────────────────────
def test_api_exposes_room_height(cli):
    p = Path.cwd() / "netdash.db"
    sid = db.save_server(p, "SRV-3U", "10.95.0.1", location="A09U13-U15")
    row = [s for s in cli.get("/api/servers").get_json()["servers"] if s["id"] == sid][0]
    assert row["room_rack"] == "A09" and row["room_unit"] == 13
    assert row["room_height"] == 3
    assert "U13-U15" in row["room_label"]


def test_single_unit_height_is_one(cli):
    p = Path.cwd() / "netdash.db"
    sid = db.save_server(p, "SRV-1U", "10.95.0.2", location="A09U20")
    row = [s for s in cli.get("/api/servers").get_json()["servers"] if s["id"] == sid][0]
    assert row["room_height"] == 1


# ── 엑셀 배치도 ──────────────────────────────────────────────────
def _xlsx_cells(data):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    return ws, [str(c.value) for row in ws.iter_rows() for c in row if c.value]


def test_xlsx_merges_multi_unit_device():
    data = serverroom.build_rack_xlsx([
        {"name": "SRV-3U", "ip": "10.1.0.1", "rack": "A09", "unit": 13,
         "height": 3, "device_type": "Server"}])
    ws, values = _xlsx_cells(data)
    assert any("SRV-3U" in v for v in values)
    assert any("3U" in v for v in values), "높이 표기가 없다"
    # 세로 병합이 실제로 생겼는지
    merges = [str(m) for m in ws.merged_cells.ranges]
    assert any(m.split(":")[0][0] == m.split(":")[1][0] for m in merges), merges


def test_xlsx_multi_unit_blocks_overlap():
    """3U 장비가 차지하는 유닛에 다른 장비를 넣으면 조용히 덮어쓰지 않는다."""
    data = serverroom.build_rack_xlsx([
        {"name": "SRV-3U", "ip": "10.1.0.1", "rack": "A09", "unit": 13, "height": 3},
        {"name": "SW-CONFLICT", "ip": "10.1.0.2", "rack": "A09", "unit": 14},
    ])
    _, values = _xlsx_cells(data)
    joined = " ".join(values)
    assert "SRV-3U" in joined and "SW-CONFLICT" in joined, \
        "겹친 장비가 조용히 사라졌다: %s" % joined[:200]


def test_export_endpoint_passes_height(cli):
    p = Path.cwd() / "netdash.db"
    db.save_server(p, "SRV-3U", "10.95.1.1", location="A09U13-U15")
    r = cli.get("/api/serverroom/export")
    assert r.status_code == 200
    _, values = _xlsx_cells(r.get_data())
    assert any("3U" in v for v in values), "엔드포인트가 height를 넘기지 않는다"


# ── 화면 ─────────────────────────────────────────────────────────
def test_rackview_renders_span_not_duplicate_rows():
    i = APPJS.index("function _rackHtml(")
    block = APPJS[i:i + 2600]
    assert "span[u] != null) continue" in block, \
        "차지된 유닛에 빈 칸을 또 그리면 장비가 여러 개처럼 보인다"
    assert "--ru-span:" in block
    assert "ru--multi" in block


def test_rackview_put_does_not_overwrite_occupied():
    i = APPJS.index("function _put(d)")
    block = APPJS[i:i + 700]
    assert "return;   // 이미 다른 장비가 차지" in block or "이미 다른 장비가 차지" in block


def test_resize_grip_exists_and_saves():
    assert "ru__grip" in APPJS and "ru__grip" in CSS
    assert "function _ruLocation(" in APPJS
    assert '"-U"' in APPJS or "'-U'" in APPJS
    i = APPJS.index("function _ruEndpoint(")
    block = APPJS[i:i + 400]
    for ep in ("/api/firewalls/", "/api/servers/", "/api/switches/"):
        assert ep in block, ep


def test_multi_unit_css_height_scales():
    m = CSS[CSS.index(".ru--multi {"):CSS.index(".ru--multi {") + 220]
    assert "--ru-span" in m and "calc(" in m, m


def test_resize_does_not_open_detail():
    """손잡이 드래그가 장비 상세 팝업을 열면 안 된다."""
    i = APPJS.index('.closest(".ru__grip")')
    block = APPJS[i:i + 500]
    assert "stopPropagation()" in block and "preventDefault()" in block
