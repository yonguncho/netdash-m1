# -*- coding: utf-8 -*-
"""claude-security 스캔(2026-07-29) 지적 반영 (v6.7.8).

Anthropic claude-security 플러그인이 전체 저장소를 medium effort로 훑어
후보 46건 → 중복 제거 44건 → 3인 적대 패널 검증 통과 10건을 보고했다.
그중 **직접 재현해 확인한 것만** 여기에 담는다.

내 자체 감사(v6.7.5)가 놓친 두 건이 특히 중요하다.
① serverroom 랙 엑셀의 수식 주입 — exporter.py만 보고 "clean"으로 판정했는데
   serverroom.py는 별도 경로였다.
② /api/upload 의 XML 엔티티 폭탄 — _read_xlsx_safe를 안 거치는 별도 경로라
   v6.7.6 수정이 닿지 않았다.
"""
import io
import sys
import zipfile
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

import app as appmod  # noqa: E402
from core import serverroom  # noqa: E402

ROOT = Path(__file__).parent.parent


# ── F10: 랙 엑셀 수식 주입 ──────────────────────────────────────
def _rack_cells(name):
    from openpyxl import load_workbook
    data = serverroom.build_rack_xlsx([
        {"name": name, "ip": "10.0.0.1", "rack": "A09", "unit": 13,
         "height": 1, "device_type": "Server"}])
    ws = load_workbook(io.BytesIO(data)).active
    return [c.value for row in ws.iter_rows() for c in row if c.value]


@pytest.mark.parametrize("payload", [
    "=cmd|'/c calc'!A1", "+1+1", "-2+3", "@SUM(A1)",
])
def test_rack_export_neutralizes_formulas(payload):
    """장비 이름이 수식으로 저장되면 받은 사람이 열 때 실행된다."""
    vals = _rack_cells(payload)
    hit = [v for v in vals if payload[:4] in str(v)]
    assert hit, "장비 이름이 시트에 없다 — 테스트 전제가 깨졌다"
    for v in hit:
        assert not str(v).startswith(("=", "+", "-", "@")), v


def test_rack_export_keeps_normal_names_intact():
    """과잉 방어로 정상 이름이 바뀌면 안 된다."""
    assert any(v == "SRV-WEB01" for v in _rack_cells("SRV-WEB01"))


def test_rack_export_strips_control_chars():
    vals = _rack_cells("SRV\tWEB\r\n01")
    assert any("SRV WEB" in str(v) for v in vals), vals


def test_serverroom_has_its_own_sanitizer():
    """exporter.py와 별도 경로다 — 한쪽만 고치면 다른 쪽이 뚫린다."""
    src = (ROOT / "core" / "serverroom.py").read_text(encoding="utf-8")
    assert "_FORMULA_PREFIX" in src and "def _cellv" in src


# ── F2: /api/upload 의 XML 엔티티 폭탄 ──────────────────────────
_BOMB = ('<?xml version="1.0"?><!DOCTYPE lolz [<!ENTITY lol "lol">'
         '<!ENTITY lol2 "&lol;&lol;&lol;&lol;&lol;&lol;&lol;&lol;&lol;&lol;">'
         ']><worksheet>&lol2;</worksheet>')


def _xlsx(sheet_xml):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", "<Types/>")
        z.writestr("xl/worksheets/sheet1.xml", sheet_xml)
    return buf.getvalue()


def test_entity_detector_flags_bomb():
    with zipfile.ZipFile(io.BytesIO(_xlsx(_BOMB))) as z:
        assert appmod._xlsx_has_xml_entities(z) is True


def test_entity_detector_passes_normal():
    ok = '<?xml version="1.0"?><worksheet><sheetData/></worksheet>'
    with zipfile.ZipFile(io.BytesIO(_xlsx(ok))) as z:
        assert appmod._xlsx_has_xml_entities(z) is False


def test_upload_route_checks_entities():
    """/api/upload 는 _read_xlsx_safe 를 안 거치는 별도 경로다."""
    src = (ROOT / "app.py").read_text(encoding="utf-8")
    i = src.index("def upload_excel(")
    block = src[i:i + 4000]
    assert "_xlsx_has_xml_entities(zf)" in block, \
        "업로드 경로에 엔티티 검사가 없다 — 크기 검사만으로는 못 막는다"


def test_entity_check_is_shared_not_duplicated():
    """두 경로가 각자 구현하면 한쪽만 고쳐지는 일이 반복된다."""
    src = (ROOT / "app.py").read_text(encoding="utf-8")
    assert src.count("def _xlsx_has_xml_entities") == 1
    assert src.count("_xlsx_has_xml_entities(zf)") >= 2


# ── F1/F7: 진단·설비 라우트의 자격증명 미검증 ───────────────────
def test_newline_in_password_is_rejected():
    """개행이 통과하면 원격 CLI에서 '한 줄 더'로 실행된다."""
    with pytest.raises(ValueError):
        appmod.validate_credential("pw\nshow running-config")
    with pytest.raises(ValueError):
        appmod.validate_credential("pw\r\nconfigure terminal")


def test_normal_password_still_accepted():
    for pw in ("S3cret!", "P@ssw0rd#2026", "a" * 200):
        assert appmod.validate_credential(pw) == pw


def test_credential_validation_applied_to_all_cred_routes():
    """수집 라우트만 검증하고 진단·설비 라우트는 빠져 있었다."""
    src = (ROOT / "app.py").read_text(encoding="utf-8")
    for fn in ("def diagnose_switch_endpoint(", "def facility_detect_subnets(",
               "def facility_collect(", "def collect_firewall_endpoint("):
        i = src.index(fn)
        block = src[i:i + 1800]
        assert "validate_credential(password)" in block, fn


# ── F6: 세션 계정 소유자 키 ─────────────────────────────────────
def test_cred_owner_uses_client_ip():
    """프록시 뒤에서 remote_addr는 전부 같아져 세션 계정 칸을 공유하게 된다."""
    src = (ROOT / "app.py").read_text(encoding="utf-8")
    i = src.index("def _cred_owner(")
    block = src[i:i + 700]
    assert "_client_ip()" in block, "raw remote_addr를 그대로 쓰고 있다"
    assert "request.remote_addr" not in block.split("return")[-1]
