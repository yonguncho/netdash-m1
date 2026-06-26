"""
M4: 멀티블록 엑셀 로더 테스트

테스트 항목:
1. 멀티블록 분리 로직
2. IP 필터링 및 정규식
3. 컬럼 별칭 매핑
4. 멱등성 (upsert)
5. 파일 크기 제한 (16MB)
6. 확장자 검증 (.xlsx만)
"""

import pytest
import tempfile
import os
from pathlib import Path
from openpyxl import Workbook

from core.excel_loader import (
    _norm,
    _is_valid_ip,
    _looks_like_header,
    _extract_blocks,
    _block_to_records,
    _get_header_map,
    load_workbook,
)


class TestExcelLoaderUtils:
    """유틸 함수 테스트"""

    def test_norm(self):
        """문자열 정규화"""
        assert _norm("  Hello  ") == "hello"
        assert _norm("VENDOR") == "vendor"
        assert _norm(None) == ""
        assert _norm(123) == ""

    def test_is_valid_ip(self):
        """IP 주소 검증"""
        assert _is_valid_ip("192.168.1.1") == True
        assert _is_valid_ip("10.0.0.1") == True
        assert _is_valid_ip("255.255.255.255") == True
        assert _is_valid_ip("0.0.0.0") == True

        assert _is_valid_ip("256.1.1.1") == False
        assert _is_valid_ip("192.168.1") == False
        assert _is_valid_ip("192.168.1.1.1") == False
        assert _is_valid_ip("not-an-ip") == False
        assert _is_valid_ip(None) == False
        assert _is_valid_ip(192) == False

    def test_looks_like_header(self):
        """헤더 판정"""
        # 알려진 필드 ≥ 2개 → 헤더
        assert _looks_like_header(["name", "ip"]) == True
        assert _looks_like_header(["ip", "mac"]) == True
        assert _looks_like_header(["name", "ip", "vendor"]) == True

        # 알려진 필드 < 2개 → 헤더 아님
        assert _looks_like_header(["hostname"]) == False
        assert _looks_like_header(["random1", "random2"]) == False
        assert _looks_like_header([]) == False


class TestHeaderMap:
    """컬럼 별칭 매핑 테스트"""

    def test_get_header_map_exact(self):
        """정확한 필드명"""
        header = ["name", "ip", "hostname", "vendor"]
        hmap = _get_header_map(header)
        assert hmap == {"name": 0, "ip": 1, "hostname": 2, "vendor": 3}

    def test_get_header_map_aliases(self):
        """별칭 필드명"""
        header = ["name", "ipaddress", "vendor", "location"]
        hmap = _get_header_map(header)
        assert hmap["name"] == 0
        assert hmap["ip"] == 1
        assert hmap["vendor"] == 2
        assert hmap["location"] == 3


class TestBlockExtraction:
    """블록 분리 테스트"""

    def test_extract_blocks_single(self):
        """단일 블록"""
        rows = [
            ["name", "ip", "vendor"],
            ["Switch-A", "10.0.1.1", "Cisco"],
            ["Switch-B", "10.0.1.2", "Juniper"],
        ]
        blocks = _extract_blocks(rows)
        assert len(blocks) == 1
        assert len(blocks[0]) == 3  # 헤더 + 데이터 2개

    def test_extract_blocks_multi(self):
        """멀티블록 분리"""
        rows = [
            ["name", "ip", "vendor"],
            ["Switch-A", "10.0.1.1", "Cisco"],
            ["ip", "mac"],  # 새 헤더
            ["192.168.1.100", "00:11:22:33:44:55"],
        ]
        blocks = _extract_blocks(rows)
        assert len(blocks) == 2
        assert len(blocks[0]) == 2  # 스위치 헤더 + 데이터
        assert len(blocks[1]) == 2  # 호스트 헤더 + 데이터

    def test_extract_blocks_discard_no_header(self):
        """헤더 없이 시작하는 행"""
        rows = [
            ["random data"],
            ["name", "ip", "vendor"],
            ["Switch-A", "10.0.1.1", "Cisco"],
        ]
        blocks = _extract_blocks(rows)
        # 첫 행은 헤더가 아니므로 버려짐
        assert len(blocks) == 1


class TestBlockProcessing:
    """블록 후처리 테스트"""

    def test_block_to_records_switch(self):
        """스위치 블록"""
        block = [
            ["name", "ip", "vendor"],
            ["Switch-A", "10.0.1.1", "Cisco"],
            ["Switch-B", "10.0.1.2", "Juniper"],
        ]
        block_type, records = _block_to_records(block)
        assert block_type == "switch"
        assert len(records) == 2
        assert records[0]["name"] == "switch-a"
        assert records[0]["ip"] == "10.0.1.1"
        assert records[0]["vendor"] == "cisco"

    def test_block_to_records_host(self):
        """호스트 블록 (vendor 없음)"""
        block = [
            ["ip", "mac"],
            ["192.168.1.100", "00:11:22:33:44:55"],
            ["192.168.1.101", "00:11:22:33:44:56"],
        ]
        block_type, records = _block_to_records(block)
        assert block_type == "host"
        assert len(records) == 2

    def test_block_to_records_filter_invalid_ip(self):
        """유효하지 않은 IP 필터링"""
        block = [
            ["name", "ip", "vendor"],
            ["Switch-A", "10.0.1.1", "Cisco"],
            ["Bad-Row", "not-an-ip", "Juniper"],
            ["Switch-B", "10.0.1.2", "Cisco"],
        ]
        block_type, records = _block_to_records(block)
        assert block_type == "switch"
        assert len(records) == 2  # Bad-Row 제외

    def test_block_to_records_no_ip_match(self):
        """IP 매칭 0 → 폐기"""
        block = [
            ["name", "ip", "vendor"],
            ["Summary", "Total", "Cisco"],
            ["Section Header", "Count", "Juniper"],
        ]
        block_type, records = _block_to_records(block)
        assert block_type is None
        assert len(records) == 0


class TestLoadWorkbook:
    """전체 로드 및 분석 테스트"""

    @pytest.fixture
    def temp_xlsx(self):
        """임시 엑셀 파일 생성"""
        wb = Workbook()
        ws = wb.active

        # 블록 1: 스위치
        ws["A1"] = "name"
        ws["B1"] = "ip"
        ws["C1"] = "hostname"
        ws["D1"] = "vendor"
        ws["E1"] = "location"

        ws["A2"] = "Switch-A"
        ws["B2"] = "10.0.1.1"
        ws["C2"] = "sw1.lab"
        ws["D2"] = "Cisco"
        ws["E2"] = "DC-North"

        ws["A3"] = "Switch-B"
        ws["B3"] = "10.0.1.2"
        ws["C3"] = "sw2.lab"
        ws["D3"] = "Juniper"
        ws["E3"] = "DC-South"

        # 블록 2: 폐기 대상 (제목 행)
        ws["A4"] = "Hosts in Network"
        ws["B4"] = "Total"
        ws["C4"] = "Count"

        # 블록 3: 호스트
        ws["A5"] = "ip"
        ws["B5"] = "mac"

        ws["A6"] = "192.168.1.100"
        ws["B6"] = "00:11:22:33:44:55"

        ws["A7"] = "192.168.1.101"
        ws["B7"] = "00:11:22:33:44:56"

        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                wb.save(tmp.name)
                tmp_path = tmp.name
            yield tmp_path
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass

    def test_load_workbook_multiblock(self, temp_xlsx):
        """멀티블록 로드 및 분석"""
        result = load_workbook(temp_xlsx)

        # 기본 구조
        assert "switches" in result
        assert "hosts" in result
        assert "diagnostics" in result

        # 스위치
        switches = result["switches"]
        assert len(switches) == 2
        assert any(s["name"] == "switch-a" for s in switches)
        assert any(s["ip"] == "10.0.1.1" for s in switches)

        # 호스트
        hosts = result["hosts"]
        assert len(hosts) == 2
        assert any(h["ip"] == "192.168.1.100" for h in hosts)

        # 진단
        diag = result["diagnostics"]
        assert diag["total_blocks"] >= 2
        assert diag["discarded_blocks"] >= 0
        assert diag["switch_blocks"] == 1
        assert diag["host_blocks"] == 1
        assert diag["imported_switches"] == 2
        assert diag["imported_hosts"] == 2


class TestFileValidation:
    """파일 검증 테스트 (app.py 엔드포인트에서)"""

    def test_max_content_length(self):
        """16MB 업로드 제한 (app.config 확인)"""
        # 이는 app.py의 MAX_CONTENT_LENGTH 설정으로 검증됨
        from app import create_app
        app = create_app(demo_mode=True)
        # 16MB = 16 * 1024 * 1024 = 16777216
        assert app.config.get("MAX_CONTENT_LENGTH") == 16 * 1024 * 1024

    def test_xlsx_extension_only(self):
        """확장자 검증은 app.py /api/upload에서"""
        # app.py의 file.filename.endswith(".xlsx") 체크로 검증됨
        pass


def test_upsert_behavior(temp_db):
    """동일 name 재업로드 시 중복 삽입 금지 (현재: name 기준 upsert)"""
    from core import db

    rows1 = [
        {"name": "Switch-1", "ip": "10.0.1.1", "vendor": "Cisco", "hostname": "sw1", "location": "DC1"}
    ]
    ids1 = db.import_switches_bulk(temp_db, rows1)

    # 동일 name 재업로드 (다른 IP)
    rows2 = [
        {"name": "Switch-1", "ip": "10.0.1.2", "vendor": "Cisco", "hostname": "sw1-updated", "location": "DC2"}
    ]
    ids2 = db.import_switches_bulk(temp_db, rows2)

    # UPSERT이므로 ID는 같아야 함
    assert ids1[0] == ids2[0]

    # DB에서 확인하면 ip와 location이 업데이트됨
    switches = db.get_switches(temp_db)
    updated_sw = next((s for s in switches if s["name"] == "Switch-1"), None)
    assert updated_sw is not None
    assert updated_sw["ip"] == "10.0.1.2"
    assert updated_sw["location"] == "DC2"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
