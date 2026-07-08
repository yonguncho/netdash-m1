"""서버실 랙 뷰 — TPS 그룹핑 키 주입 + UI 테스트."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

ROOT = Path(__file__).parent.parent
APP_JS = ROOT / "web" / "static" / "app.js"
CSS = ROOT / "web" / "static" / "style.css"
HTML = ROOT / "web" / "templates" / "index.html"


def test_state_injects_rack_group(client):
    client.post("/api/switches/manual",
                json={"ip": "10.8.8.1", "name": "FA_SW1", "vendor": "cisco",
                      "hostname": "TPS-F1B02_1F01_FA_SW1"})
    sw = next(x for x in client.get("/api/state").get_json()["switches"] if x["ip"] == "10.8.8.1")
    assert "tps_group" in sw
    assert "1공장" in sw["tps_group"]
    assert "Assembly" in sw["tps_group"]
    assert sw["tps_num"] == "TPS01"


def test_rack_view_ui():
    js = APP_JS.read_text(encoding="utf-8")
    assert "function renderRackView" in js
    assert "_viewMode" in js
    assert "tps_group" in js and "tps_num" in js
    css = CSS.read_text(encoding="utf-8")
    assert ".rack-unit" in css
    html = HTML.read_text(encoding="utf-8")
    assert 'id="btn-view-rack"' in html
    assert 'id="rack-view"' in html


def test_serverroom_rack_xlsx_layout():
    """랙 배치 엑셀: 랙명 헤더 + U42→U1 행 + 장비 종류색 채움."""
    from core import serverroom
    from openpyxl import load_workbook
    import io
    devs = [
        {"name": "FW_M", "ip": "10.1.1.1", "rack": "A03", "unit": 36, "device_type": "Firewall"},
        {"name": "L3_1", "ip": "10.1.1.2", "rack": "A03", "unit": 30, "device_type": "L3 Switch"},
        {"name": "L2_1", "ip": "10.1.1.3", "rack": "B03", "unit": 10, "device_type": "L2 Switch"},
    ]
    data = serverroom.build_rack_xlsx(devs)
    ws = load_workbook(io.BytesIO(data)).active
    assert ws["A1"].value == "A03"                       # 첫 랙명 헤더
    assert ws["A2"].value == "U42"                        # U42가 최상단
    # U36 = 헤더(1) + (42-36) + 1 = 8행, 장비명 컬럼(B)
    dev_cell = ws.cell(row=8, column=2)
    assert dev_cell.value == "FW_M"
    assert dev_cell.fill.fgColor.rgb == "FFEF4444"        # 방화벽=빨강 채움


def test_serverroom_rack_xlsx_letter_grouping():
    """같은 열(A)의 랙 3개 이상이 한 줄에 모두 나란히 표기(2개 제한 없음)."""
    from core import serverroom
    from openpyxl import load_workbook
    import io
    devs = [
        {"name": "A3", "ip": "10.1.1.1", "rack": "A03", "unit": 36, "device_type": "Firewall"},
        {"name": "A4", "ip": "10.1.1.2", "rack": "A04", "unit": 30, "device_type": "L3 Switch"},
        {"name": "A6", "ip": "10.1.1.3", "rack": "A06", "unit": 20, "device_type": "L4 Switch"},
        {"name": "B3", "ip": "10.1.1.4", "rack": "B03", "unit": 10, "device_type": "L2 Switch"},
    ]
    ws = load_workbook(io.BytesIO(serverroom.build_rack_xlsx(devs))).active
    # 첫 줄(A열) 헤더에 A03/A04/A06 모두 존재
    row1 = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert "A03" in row1 and "A04" in row1 and "A06" in row1
    # 전체에서 B03도 존재(다른 letter 블록)
    allv = set()
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            allv.add(ws.cell(row=r, column=c).value)
    assert "B03" in allv


def test_serverroom_export_endpoint_404_when_empty(client):
    """서버실 소속 장비 없으면 404 + 안내."""
    r = client.get("/api/serverroom/export")
    assert r.status_code == 404


def test_serverroom_export_endpoint_xlsx(client):
    """location A03U30 장비 등록 → 엑셀 다운로드 200."""
    client.post("/api/switches/manual",
                json={"ip": "10.9.9.9", "name": "SR_L3", "vendor": "cisco",
                      "location": "A03U30"})
    r = client.get("/api/serverroom/export")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("Content-Type", "")
    assert len(r.data) > 1000
