# -*- coding: utf-8 -*-
"""엑셀 현황 보고서 빌더 테스트 (v3.27 재설계: 요약/스위치/설비/알람/포트 5시트)."""
import io
import sys
from pathlib import Path

import pytest
import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent))

from core import db, report_builder

SHEETS = ["요약", "스위치 현황", "설비 현황", "알람 이력", "포트 현황"]


def _load(data):
    return openpyxl.load_workbook(io.BytesIO(data))


def _seed(db_path):
    """스위치 + 스냅샷(포트/맥) + 설비 + 알람 시드."""
    sid = db.save_switch(db_path, "ACC-SW01", "10.0.0.20", "cisco_ios")
    snap = db.save_snapshot(db_path, sid)
    db.save_ports(db_path, snap, sid, [
        {"name": "Gi1/0/1", "status": "up", "vlan": 10, "speed": "1G", "description": "uplink"},
        {"name": "Gi1/0/2", "status": "down", "vlan": 20, "speed": "1G", "description": ""},
    ])
    db.save_mac_entries(db_path, snap, sid, [
        {"vlan": 10, "mac": "00:11:22:33:44:55", "port": "Gi1/0/1", "type": "dynamic"},
    ])
    db.save_facility_hosts(db_path, [
        {"subnet": "10.0.1.0/24", "ip": "10.0.1.50", "mac": "00:11:22:33:44:66",
         "switch_id": sid, "switch_name": "ACC-SW01", "port": "Gi1/0/2",
         "online": 1, "direct": 1}])
    db.save_device_event(db_path, "new_device", "warning", subnet="10.0.1.0/24",
                         ip="10.0.1.50", message="새 설비 감지: 10.0.1.50")
    return sid


# 1. bytes 반환 + 재로딩 가능
def test_build_report_returns_bytes(temp_db):
    _seed(temp_db)
    data = report_builder.build_report(temp_db)
    assert isinstance(data, bytes) and len(data) > 0
    assert _load(data) is not None


# 2. 5시트 구성(요약이 첫 시트)
def test_report_sheets(temp_db):
    _seed(temp_db)
    wb = _load(report_builder.build_report(temp_db))
    assert wb.sheetnames == SHEETS


# 3. 요약 시트에 통계·생성시각
def test_summary_sheet(temp_db):
    _seed(temp_db)
    ws = _load(report_builder.build_report(temp_db))["요약"]
    text = " ".join(str(v) for row in ws.iter_rows(values_only=True) for v in row if v is not None)
    assert "등록 스위치" in text and "1대" in text
    assert "설비(대역 스캔)" in text
    assert "생성 시각" in text


# 4. 스위치 시트: 한국어 헤더 + 포트 사용/전체 + 상태 한국어
def test_switch_sheet(temp_db):
    _seed(temp_db)
    ws = _load(report_builder.build_report(temp_db))["스위치 현황"]
    headers = [c.value for c in ws[1]]
    assert headers[:6] == ["구분", "IP", "호스트네임", "벤더", "위치", "상태"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    sw_row = [r for r in rows if r[0] == "ACC-SW01"][0]
    assert sw_row[7] == "1 / 2"      # up 1 / 전체 2
    assert sw_row[5] in ("미수집", "정상", "수집 실패", "수집 중")  # 상태 한국어


# 5. 설비 시트에 설비 반영
def test_facility_sheet(temp_db):
    _seed(temp_db)
    ws = _load(report_builder.build_report(temp_db))["설비 현황"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert any(r[1] == "10.0.1.50" for r in rows)


# 6. 알람 시트에 이벤트 반영(한국어 종류)
def test_alerts_sheet(temp_db):
    _seed(temp_db)
    ws = _load(report_builder.build_report(temp_db))["알람 이력"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert any(r[1] == "새 설비" and r[4] == "10.0.1.50" for r in rows)


# 7. 포트 시트
def test_port_sheet(temp_db):
    _seed(temp_db)
    ws = _load(report_builder.build_report(temp_db))["포트 현황"]
    names = [r[1] for r in ws.iter_rows(min_row=2, values_only=True)]
    assert "Gi1/0/1" in names and "Gi1/0/2" in names


# 8. 빈 DB → 예외 없음 + 시트 구성 유지
def test_build_report_empty_db(temp_db):
    wb = _load(report_builder.build_report(temp_db))
    assert wb.sheetnames == SHEETS
    assert wb["스위치 현황"].max_row == 1  # 헤더만


# 9. GET /api/report → 200 + xlsx
def test_api_report_endpoint(client):
    r = client.get("/api/report")
    assert r.status_code == 200
    assert "spreadsheetml.sheet" in r.headers.get("Content-Type", "")
    assert "attachment" in r.headers.get("Content-Disposition", "")
    wb = _load(r.data)
    assert len(wb.sheetnames) == 5
