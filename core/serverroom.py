# -*- coding: utf-8 -*-
"""서버실 랙 위치 해석. location 필드에 "A09U27"처럼 기재하면 랙/유닛으로 해석.

포맷: {랙}U{유닛}
  예) A09U27 → A09 랙, U27 유닛
      B12U04 → B12 랙, U4 유닛
랙 = 영문+숫자, 유닛 = U 뒤 숫자. 대소문자·공백 허용.
이 포맷과 일치하는 장비만 "서버실 소속"으로 본다.
"""
import re

from . import utils

# {랙: 영문1자이상+숫자1자이상}U{유닛}[-[U]{끝유닛}]
#   A09U27        → A09랙 27U 1개
#   A09U13-U15    → A09랙 13~15U (3U 장비)
#   A09U13-15     → 같음(U 생략 허용)
# 접미 공백·대소문자·하이픈 주변 공백 허용.
_PAT = re.compile(
    r"^\s*([A-Za-z]+\d+)\s*[Uu]\s*(\d+)\s*(?:[-~]\s*[Uu]?\s*(\d+)\s*)?$")

MAX_HEIGHT = 42       # 랙 하나 높이 — 그보다 큰 값은 오기로 본다


def parse_rack(location):
    """location에서 "{랙}U{유닛}" 또는 "{랙}U{시작}-U{끝}" 패턴을 해석.

    일치하지 않으면 None(= 서버실 소속 아님).
    Returns: {rack, unit, unit_end, height, label} | None
      unit     — 시작 유닛(가장 작은 번호). 기존 코드가 쓰던 의미 그대로.
      unit_end — 끝 유닛(단일이면 unit과 같음)
      height   — 차지하는 U 수(1 이상)
    """
    if not location:
        return None
    m = _PAT.match(str(location))
    if not m:
        return None
    rack = m.group(1).upper()
    try:
        unit = int(m.group(2))
        end = int(m.group(3)) if m.group(3) else unit
    except ValueError:
        return None
    # 'U15-U13'처럼 거꾸로 적어도 받아들인다(사람이 흔히 하는 실수)
    if end < unit:
        unit, end = end, unit
    height = end - unit + 1
    if height > MAX_HEIGHT:
        end = unit + MAX_HEIGHT - 1
        height = MAX_HEIGHT
    label = ("%s랙 U%d" % (rack, unit) if height == 1
             else "%s랙 U%d-U%d (%dU)" % (rack, unit, end, height))
    return {"rack": rack, "unit": unit, "unit_end": end,
            "height": height, "label": label}


def format_rack(rack, unit, height=1):
    """랙/유닛/높이 → location 문자열. 드래그로 크기를 바꿀 때 쓴다."""
    try:
        unit = int(unit)
        height = max(1, min(int(height or 1), MAX_HEIGHT))
    except (TypeError, ValueError):
        return ""
    if height == 1:
        return "%sU%d" % (str(rack).upper(), unit)
    return "%sU%d-U%d" % (str(rack).upper(), unit, unit + height - 1)


def occupied_units(info):
    """parse_rack 결과 → 그 장비가 차지하는 유닛 번호 목록."""
    if not info:
        return []
    return list(range(info["unit"], info.get("unit_end", info["unit"]) + 1))


# 장비 종류 → 엑셀 셀 채움색(ARGB, 프론트 _RACK_KIND와 동일 팔레트)
_KIND_FILL = {
    "Firewall": "FFEF4444", "BackBone": "FFA855F7", "L3 Switch": "FF8B5CF6",
    "L4 Switch": "FFF59E0B", "L2 Switch": "FF14B8A6", "Server": "FF3B82F6",
    "AP": "FF22C55E",
}
_RACK_U = 42          # 랙 높이(U)
_RACKS_PER_ROW = 2    # 한 줄에 랙 2개(첨부 엑셀 형식)


def _infer_dt(name):
    t = (name or "").upper()
    import re as _re
    if _re.search(r"_FW|-FW|FIREWALL|ASA|PALO|FORTI", t):
        return "Firewall"
    if _re.search(r"L4|SLB|ADC|ALTEON|OASVR", t):
        return "L4 Switch"
    if _re.search(r"BACKBONE|\bBB\b|BB\d|CORE", t):
        return "BackBone"
    if _re.search(r"L3|DSW", t):
        return "L3 Switch"
    if _re.search(r"L2|FASW|ASW|ACC|SW", t):
        return "L2 Switch"
    return ""


# Excel이 수식으로 해석하는 접두 문자 — 셀 값 앞에 오면 무력화한다.
# exporter._s 와 같은 정책이다. 이쪽은 별도 경로라 방어가 빠져 있었고,
# 장비 이름(엑셀 임포트·수동 입력·장비 hostname에서 온다)이 그대로 셀에 들어가
# 받은 사람이 열면 실행됐다(예: =cmd|'/c calc'!A1).
_FORMULA_PREFIX = ("=", "+", "-", "@")


def _cellv(v):
    """셀 값 문자열화 + 수식 접두 무력화(앞에 ' 를 붙이면 Excel이 텍스트로 본다)."""
    if v is None:
        return ""
    s = str(v)
    for _ch in (chr(9), chr(13), chr(10)):      # tab, CR, LF
        s = s.replace(_ch, " ")
    s = s.strip()
    if s[:1] in _FORMULA_PREFIX:
        s = "'" + s
    return s


def build_rack_xlsx(devices):
    """서버실 랙 배치를 첨부 엑셀 형식으로 생성.

    devices: [{name, ip, rack, unit, device_type}] — parse_rack 통과분만.
    Returns: xlsx bytes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    # 랙 → {unit: dev}
    # 같은 랙·유닛에 두 장비가 등록돼 있으면 예전엔 뒤엣것이 앞엣것을 조용히
    # 덮어써서 배치도에서 장비가 사라졌다(화면과 엑셀이 서로 다른 장비를 보여줌).
    # 덮어쓰지 않고 한 칸에 함께 적고, 중복 사실을 로그로 남긴다.
    # 다중 U 장비(예: U13-U15)는 시작 유닛에 두고, 나머지 유닛은 'span'으로 표시해
    # 그 칸을 비워 둔다(엑셀에서는 세로 병합으로 한 덩어리처럼 보인다).
    racks = {}
    spans = {}          # {rack: {unit: 시작유닛}} — 이 유닛은 위 장비가 차지 중
    dup = []
    for d in devices:
        rk, u = d.get("rack"), d.get("unit")
        if not rk or not u:
            continue
        h = max(1, min(int(d.get("height") or 1), MAX_HEIGHT))
        slot = racks.setdefault(rk, {})
        span = spans.setdefault(rk, {})
        taken = [x for x in range(u, u + h) if x in slot or x in span]
        if taken:
            first = taken[0]
            owner = slot.get(first) or slot.get(span.get(first))
            dup.append("%s U%s: %s / %s" % (rk, first,
                                            (owner or {}).get("name"), d.get("name")))
            if owner is not None:
                owner["name"] = "%s + %s" % (owner.get("name") or "?", d.get("name") or "?")
                owner["ip"] = "%s / %s" % (owner.get("ip") or "", d.get("ip") or "")
                owner["dup"] = True
            continue
        dev = dict(d)
        dev["height"] = h
        slot[u] = dev
        for x in range(u + 1, u + h):
            span[x] = u
    if dup:
        utils.log_event("warning", "serverroom_duplicate_unit", count=len(dup),
                        samples=dup[:5])
    rack_names = sorted(racks.keys())

    # 열(letter: A/B...)별 그룹핑 — 랙뷰와 동일하게 같은 열의 랙을 한 줄에 나란히.
    import re as _re
    by_letter = {}
    for rk in rack_names:
        m = _re.match(r"[A-Za-z]+", rk)
        letter = m.group(0).upper() if m else "#"
        by_letter.setdefault(letter, []).append(rk)
    letters = sorted(by_letter)
    max_racks_row = max((len(by_letter[l]) for l in letters), default=1)

    wb = Workbook()
    ws = wb.active
    ws.title = "ServerRoom"

    thin = Side(style="thin", color="FFBFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="FF1F2937")
    hdr_font = Font(bold=True, color="FFFFFFFF", size=12)
    u_font = Font(color="FF64748B", size=9)
    dev_font = Font(bold=True, color="FFFFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # 각 랙 블록: 2컬럼(U라벨, 장비명) + 1컬럼 간격
    BLOCK_COLS = 3
    row0 = 1
    for letter in letters:
        chunk = sorted(by_letter[letter])   # 이 열의 모든 랙(A03,A04,A06...)을 나란히
        # 이 줄에서 사용할 최대 U (기본 42, 초과 시 확장)
        max_u = _RACK_U
        for rk in chunk:
            for u, dv in racks[rk].items():
                top = u + max(1, dv.get("height") or 1) - 1
                if top > max_u:
                    max_u = top
        # 헤더(랙명) — U라벨+장비명 두 칸 병합
        for ci, rk in enumerate(chunk):
            c0 = 1 + ci * BLOCK_COLS  # U라벨 컬럼
            c1 = c0 + 1               # 장비명 컬럼
            ws.merge_cells(start_row=row0, start_column=c0, end_row=row0, end_column=c1)
            cell = ws.cell(row=row0, column=c0, value=_cellv(rk))
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center
            cell.border = border
            ws.cell(row=row0, column=c1).border = border
        # U 행 (U{max_u} → U1)
        for idx, u in enumerate(range(max_u, 0, -1)):
            r = row0 + 1 + idx
            for ci, rk in enumerate(chunk):
                c0 = 1 + ci * BLOCK_COLS
                c1 = c0 + 1
                uc = ws.cell(row=r, column=c0, value="U%d" % u)
                uc.font = u_font
                uc.alignment = center
                uc.border = border
                dc = ws.cell(row=r, column=c1)
                dc.border = border
                dc.alignment = left
                dev = racks[rk].get(u)
                if not dev:
                    # 위 장비가 차지 중인 유닛이면 색만 이어 칠해 한 덩어리로 보이게
                    owner_u = spans.get(rk, {}).get(u)
                    if owner_u is not None:
                        owner = racks[rk].get(owner_u) or {}
                        dt = owner.get("device_type") or _infer_dt(owner.get("name"))
                        fill = _KIND_FILL.get(dt)
                        if fill:
                            dc.fill = PatternFill("solid", fgColor=fill)
                    continue
                h = max(1, dev.get("height") or 1)
                if h > 1:
                    # 위쪽(U가 큰 쪽)이 행 번호가 작다 → 시작 유닛의 행이 아래쪽
                    top_r = row0 + 1 + (max_u - (u + h - 1))
                    ws.merge_cells(start_row=top_r, start_column=c1,
                                   end_row=r, end_column=c1)
                    dc = ws.cell(row=top_r, column=c1)
                    dc.border = border
                    dc.alignment = center
                dc.value = _cellv((dev.get("name") or "")
                                  + (" (%dU)" % h if h > 1 else ""))
                dc.font = dev_font
                dt = dev.get("device_type") or _infer_dt(dev.get("name"))
                fill = _KIND_FILL.get(dt)
                if fill:
                    dc.fill = PatternFill("solid", fgColor=fill)
        row0 = row0 + 1 + max_u + 2  # 다음 랙-줄 블록(간격 2행)

    # 컬럼 너비 — 한 줄 최대 랙 수 기준
    ncols = max_racks_row * BLOCK_COLS
    for c in range(1, ncols + 1):
        col = get_column_letter(c)
        if (c - 1) % BLOCK_COLS == 0:
            ws.column_dimensions[col].width = 6      # U라벨
        elif (c - 1) % BLOCK_COLS == 1:
            ws.column_dimensions[col].width = 26     # 장비명
        else:
            ws.column_dimensions[col].width = 2      # 간격

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
