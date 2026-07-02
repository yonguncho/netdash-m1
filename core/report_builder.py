# -*- coding: utf-8 -*-
"""엑셀 현황 보고서 빌더 (v3.27 재설계 — 필요한 정보만 한국어로 보기 좋게).

build_report(db_path) -> bytes
시트 구성:
  1. 요약        — 생성 시각, 장비 통계, 문제 장비(수집 실패·경보·연결 끊김) 한눈에
  2. 스위치 현황 — 구분/IP/호스트네임/벤더/위치/상태/경보/포트 사용/MAC 수/마지막 수집
  3. 설비 현황   — 설비 탭과 동일(대역/IP/MAC/연결 스위치/포트/직접연결/상태)
  4. 알람 이력   — 최근 변경 이벤트(새 설비/연결 끊김/스위치 실패/flapping 등)
  5. 포트 현황   — 스위치별 포트 상태(up/down/disabled)/VLAN/속도/설명

(구버전의 '이벤트 로그'·'호스트 대장(장부 대조)' 시트는 UI 기능 정리에 맞춰 제거)
"""
import io
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from . import db as _db

logger = logging.getLogger(__name__)


# ── 스타일 ──────────────────────────────────────────────────────
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F497D")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_TITLE_FONT = Font(bold=True, size=14)
_FILL_OK = PatternFill(fill_type="solid", fgColor="DCFCE7")       # 초록(정상/온라인)
_FILL_BAD = PatternFill(fill_type="solid", fgColor="FEE2E2")      # 빨강(실패/끊김)
_FILL_WARN = PatternFill(fill_type="solid", fgColor="FEF3C7")     # 노랑(경보/미확인)

_STATUS_KO = {"done": "정상", "failed": "수집 실패", "collecting": "수집 중", "new": "미수집"}
_ALERT_KO = {"none": "", "warning": "⚠ FLAP", "critical": "⚠ LOOP"}
_EVENT_KO = {
    "new_device": "새 설비", "device_offline": "설비 연결 끊김", "device_online": "설비 복구",
    "device_moved": "설비 이동", "config_changed": "설정 변경",
    "switch_unreachable": "스위치 연결 실패", "switch_recovered": "스위치 복구",
    "flapping": "포트 flapping", "looping": "포트 looping",
}


def _style_header_row(ws, row=1):
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _auto_col_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                # 한글은 폭이 넓어 1.6배 가중
                v = str(cell.value) if cell.value is not None else ""
                w = sum(1.6 if ord(c) > 0x1100 else 1 for c in v)
                if w > max_len:
                    max_len = w
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(int(max_len) + 3, 55)


def _switch_location(sw):
    """스위치 위치 라벨: TPS(hostname 해석) > 서버실 랙 > 자유 텍스트."""
    try:
        from . import tps_location, serverroom
        info = tps_location.parse(sw.get("hostname"))
        if info:
            return info["label"]
        room = serverroom.parse_rack(sw.get("location"))
        if room:
            return "서버실 " + room["label"]
    except Exception:
        pass
    return sw.get("location") or ""


# ── 시트 빌더 ────────────────────────────────────────────────────
def _build_summary_sheet(wb, db_path, switches, fac_hosts):
    ws = wb.create_sheet("요약")
    ws.append(["NetDash 네트워크 현황 보고서"])
    ws["A1"].font = _TITLE_FONT
    ws.append(["생성 시각", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws.append([])

    done = sum(1 for s in switches if s.get("status") == "done")
    failed = sum(1 for s in switches if s.get("status") == "failed")
    alerts = [s for s in switches if (s.get("alert") or "none") != "none"]
    fac_on = sum(1 for h in fac_hosts if h.get("online"))
    fac_off = len(fac_hosts) - fac_on
    try:
        firewalls = _db.list_firewalls(db_path)
    except Exception:
        firewalls = []
    unacked = _db.count_unacked_events(db_path)

    ws.append(["항목", "값"])
    _style_header_row(ws, ws.max_row)
    # 전체 포트 사용률(증설 계획 참고)
    total_up = total_ports = 0
    for s in switches:
        snap_id = _db.latest_snapshot_id(db_path, s["id"])
        ports = _db.get_ports(db_path, snap_id) if snap_id else []
        total_up += sum(1 for p in ports if p.get("status") == "up")
        total_ports += len(ports)
    usage = ("%d%% (%d/%d포트 사용)" % (round(total_up * 100 / total_ports), total_up, total_ports)
             if total_ports else "수집 전")

    rows = [
        ("등록 스위치", "%d대 (정상 %d · 수집실패 %d · 기타 %d)" % (
            len(switches), done, failed, len(switches) - done - failed)),
        ("등록 방화벽", "%d대" % len(firewalls)),
        ("설비(대역 스캔)", "%d대 (온라인 %d · 연결실패 %d)" % (len(fac_hosts), fac_on, fac_off)),
        ("전체 포트 사용률", usage),
        ("경보 스위치(FLAP/LOOP)", "%d대" % len(alerts)),
        ("미확인 알람", "%d건" % unacked),
    ]
    for k, v in rows:
        ws.append([k, v])

    # 문제 장비 목록(있을 때만)
    problems = [(s.get("name"), s.get("ip"), "수집 실패") for s in switches if s.get("status") == "failed"]
    problems += [(s.get("name"), s.get("ip"), _ALERT_KO.get(s.get("alert"), s.get("alert"))) for s in alerts]
    if problems:
        ws.append([])
        ws.append(["⚠ 확인 필요 장비", "IP", "사유"])
        _style_header_row(ws, ws.max_row)
        for name, ip, why in problems:
            ws.append([name or "", ip or "", why or ""])
            for c in ws[ws.max_row]:
                c.fill = _FILL_BAD if "실패" in str(why) else _FILL_WARN

    _auto_col_width(ws)
    return ws


def _build_switch_sheet(wb, db_path, switches):
    ws = wb.create_sheet("스위치 현황")
    ws.append(["구분", "IP", "호스트네임", "벤더", "위치", "상태", "경보",
               "포트(사용/전체)", "사용률", "MAC 수", "마지막 수집"])
    _style_header_row(ws)

    total_up = total_ports = 0
    for sw in switches:
        snap_id = _db.latest_snapshot_id(db_path, sw["id"])
        ports = _db.get_ports(db_path, snap_id) if snap_id else []
        up = sum(1 for p in ports if p.get("status") == "up")
        total_up += up
        total_ports += len(ports)
        mac_count = _db.get_mac_count(db_path, snap_id) if snap_id else 0
        status = sw.get("status") or "new"
        ws.append([
            sw.get("name") or "",
            sw.get("ip") or "",
            sw.get("hostname") or "",
            sw.get("vendor") or "",
            _switch_location(sw),
            _STATUS_KO.get(status, status),
            _ALERT_KO.get(sw.get("alert") or "none", sw.get("alert") or ""),
            ("%d / %d" % (up, len(ports))) if ports else "",
            ("%d%%" % round(up * 100 / len(ports))) if ports else "",
            mac_count or "",
            (sw.get("last_collected") or "").replace("T", " ")[:16],
        ])
        # 상태 색상
        cell = ws.cell(row=ws.max_row, column=6)
        if status == "done":
            cell.fill = _FILL_OK
        elif status == "failed":
            cell.fill = _FILL_BAD
        if (sw.get("alert") or "none") != "none":
            ws.cell(row=ws.max_row, column=7).fill = _FILL_WARN

    ws.freeze_panes = "A2"
    _auto_col_width(ws)
    return ws


def _build_facility_sheet(wb, db_path):
    from . import facility
    ws = wb.create_sheet("설비 현황")
    cols = facility._EXPORT_COLS
    ws.append(cols)
    _style_header_row(ws)
    for r in facility._export_rows(db_path):
        ws.append([r[c] for c in cols])
        st = ws.cell(row=ws.max_row, column=cols.index("상태") + 1)
        st.fill = _FILL_OK if r["상태"] == "온라인" else _FILL_BAD
    ws.freeze_panes = "A2"
    _auto_col_width(ws)
    return ws


def _build_alerts_sheet(wb, db_path):
    ws = wb.create_sheet("알람 이력")
    ws.append(["시각", "종류", "심각도", "장비/스위치", "IP", "대역", "내용", "확인"])
    _style_header_row(ws)
    for ev in _db.list_device_events(db_path, limit=500):
        kind = _EVENT_KO.get(ev.get("kind"), ev.get("kind") or "")
        ws.append([
            (ev.get("ts") or "").replace("T", " ")[:16],
            kind,
            ev.get("severity") or "",
            ev.get("label") or "",
            ev.get("ip") or "",
            ev.get("subnet") or "",
            ev.get("message") or "",
            "확인" if ev.get("ack") else "미확인",
        ])
        if not ev.get("ack"):
            ws.cell(row=ws.max_row, column=8).fill = _FILL_WARN
    ws.freeze_panes = "A2"
    _auto_col_width(ws)
    return ws


def _build_port_sheet(wb, db_path, switches):
    ws = wb.create_sheet("포트 현황")
    ws.append(["스위치", "포트", "상태", "VLAN", "속도", "설명"])
    _style_header_row(ws)

    for sw in switches:
        snap_id = _db.latest_snapshot_id(db_path, sw["id"])
        if not snap_id:
            continue
        for port in _db.get_ports(db_path, snap_id):
            ws.append([
                sw.get("name") or "",
                port.get("name") or "",
                port.get("status") or "",
                port.get("vlan") if port.get("vlan") is not None else "",
                port.get("speed") or "",
                port.get("description") or "",
            ])
    ws.freeze_panes = "A2"
    _auto_col_width(ws)
    return ws


# ── 공개 API ─────────────────────────────────────────────────────
def build_report(db_path: str) -> bytes:
    """현재 DB 상태를 5시트 엑셀로 빌드해 bytes 반환.

    시트: 요약 / 스위치 현황 / 설비 현황 / 알람 이력 / 포트 현황.
    빈 DB여도 헤더만 있는 정상 워크북을 반환한다.
    """
    logger.info("build_report: start db_path=%s", db_path)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    switches = _db.get_switches(db_path)
    fac_hosts = _db.get_facility_hosts(db_path)

    _build_summary_sheet(wb, db_path, switches, fac_hosts)
    _build_switch_sheet(wb, db_path, switches)
    _build_facility_sheet(wb, db_path)
    _build_alerts_sheet(wb, db_path)
    _build_port_sheet(wb, db_path, switches)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    result = buf.read()

    logger.info("build_report: done size=%d", len(result))
    return result
