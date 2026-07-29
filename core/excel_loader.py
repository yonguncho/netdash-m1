"""
M4: 멀티블록 엑셀 로더

단일 엑셀 파일에서 여러 개의 데이터 블록(스위치, 호스트, 제목 등)을 자동으로 분리하고
각 블록을 해당 테이블에 DB 반영하는 모듈.

알려진 필드:
  - 스위치: name, ip, hostname, vendor, location
  - 호스트: ip, mac

멀티블록 분리:
  - 헤더 후보: 알려진 필드 ≥ MIN_MATCHED_COLS(기본값 2) 매칭
  - 새 헤더 발견 시 이전 블록 확정
  - IP 정규식 매칭 0개 블록은 폐기 (제목, 요약, 섹션 헤더)

멱등성:
  - IP 기준 UPSERT (동일 IP 재업로드 시 중복 삽입 금지)
"""

import re
import logging
from typing import Tuple, List, Dict, Any, Optional
from openpyxl import load_workbook as openpyxl_load

logger = logging.getLogger(__name__)

MIN_MATCHED_COLS = 2

# 알려진 필드와 별칭 매핑 (모두 정규화된 형태: 공백 제거 + 소문자)
ALIASES = {
    # 주의: 같은 헤더 문자열을 두 필드에 넣지 말 것(콜맵 충돌로 한쪽이 유실됨).
    # '호스트명'은 hostname 전용 — 이름(name)이 없으면 파서가 hostname으로 대체한다.
    'name': {'name', 'switchname', 'devicename', '이름', '장비명', '장비이름', '서버이름'},
    'ip': {'ip', 'ipv4', 'ipaddress', 'address', '호스트', '호스트ip', '관리ip', 'host',
           '대표ip', '대표아이피', '아이피', 'ip주소', 'ip address'},
    'hostname': {'hostname', 'dnsname', 'fqdn', '사용서버명', '서버명', '호스트명', '호스트네임'},
    'vendor': {'vendor', 'manufacturer', 'os', 'platform', '벤더', '제조사'},
    # 위치는 랙 위치(A09U27)만 — '용도/용도상세'는 랙 파싱을 깨뜨리므로 매핑하지 않는다.
    'location': {'location', 'site', 'datacenter', 'dc', '랙위치', '위치', '랙'},
    'mac': {'mac', 'macaddress', 'hwaddr'},
    # M7: 장부(ledger) 위치 — 호스트 블록에서 기대 연결 위치를 적재
    'ledger_switch': {'연결스위치', 'connectedswitch', 'ledgerswitch', 'uplinkswitch'},
    'ledger_port': {'연결포트', 'connectedport', 'ledgerport', 'uplinkport'},
    # 장비 인벤토리 업로드용 서브넷
    'subnet': {'subnet', 'ipsubnet', 'cidr', 'network', 'netmask', '서브넷', '대역', '네트워크'},
    # 방화벽/서버 인벤토리용
    'port': {'port', '포트', 'mgmtport', '관리포트'},
    'os_type': {'ostype', 'osname', '운영체제', 'osversion', 'os version', 'os버전', '운영체제버전'},
    # 서버 사양 일괄 등록용 — 실측 수집 없이 엑셀로 CPU·메모리·디스크를 채운다.
    # 단위가 섞여 들어오는 칸이 많아(메모리를 GB로 적는 경우가 대부분) MB/GB를
    # 별도 필드로 받아 파서에서 환산한다.
    'cpu_model': {'cpumodel', 'cpu', 'processor', 'cpu모델', '프로세서', 'cpu종류'},
    'cpu_cores': {'cpucores', 'cores', 'corecount', 'cpu코어', '코어수', '코어', 'cpu코어수'},
    'mem_total_mb': {'memmb', 'memorymb', 'ram_mb', '메모리mb', '메모리(mb)'},
    'mem_total_gb': {'memgb', 'memorygb', 'ram', 'ram_gb', '메모리', '메모리gb', '메모리(gb)',
                     'memory'},
    'disk_total_gb': {'diskgb', 'disktotal', 'disk', 'storage', '디스크', '디스크gb',
                      '디스크(gb)', '디스크용량', '전체디스크'},
    'disk_used_gb': {'diskused', 'diskusedgb', '사용디스크', '디스크사용량', '사용중디스크'},
}

IP_REGEX = re.compile(
    r'^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$'
)


def _norm(s: Optional[Any]) -> str:
    """정규화: 앞뒤 공백 제거 + 내부 공백 모두 제거 + 소문자 (매칭·키용)"""
    if not isinstance(s, str):
        return ""
    return s.strip().lower().replace(" ", "")


# 표시용 컬럼 — _norm(소문자/공백제거) 대신 원문 strip만 적용해 표기 보존
_DISPLAY_COLS = {"name", "hostname", "location", "vendor", "note", "model", "description"}


def _is_valid_ip(ip_str: Optional[Any]) -> bool:
    """IP 주소 검증 (IPv4)"""
    if not isinstance(ip_str, str):
        return False
    return bool(IP_REGEX.match(ip_str.strip()))


def parse_switch_inventory(source) -> List[Dict[str, Any]]:
    """IP/SUBNET/HOSTNAME 인벤토리 엑셀 → 스위치 등록 행 목록.

    벤더 무관: IP가 있는 모든 행을 스위치로 본다(벤더는 'unknown', 이후 수정).
    헤더에서 ip/hostname/name/subnet/location 컬럼을 자동 매핑.
    Returns: [{name, ip, hostname, subnet, location, vendor='unknown'}]
    """
    wb = openpyxl_load(source, read_only=True, data_only=True)
    wanted = ('ip', 'hostname', 'name', 'subnet', 'location')
    field_alias = {f: ALIASES[f] for f in wanted if f in ALIASES}
    out = []
    seen = set()
    for ws in wb.worksheets:
        colmap = None
        for row in ws.iter_rows(values_only=True):
            if colmap is None:
                m = {}
                for ci, cell in enumerate(row):
                    n = _norm(cell)
                    for f, al in field_alias.items():
                        if n in al and f not in m:   # first-wins(뒤 컬럼이 앞을 덮지 않음)
                            m[f] = ci
                if 'ip' in m:  # ip 컬럼이 있는 행을 헤더로 인식
                    colmap = m
                continue

            def _get(field):
                ci = colmap.get(field)
                if ci is None or ci >= len(row):
                    return ""
                v = row[ci]
                return str(v).strip() if v is not None else ""

            ip = _get('ip')
            if not _is_valid_ip(ip) or ip in seen:
                continue
            seen.add(ip)
            host = _get('hostname')
            name = _get('name') or host or ip
            out.append({
                "name": name, "ip": ip, "hostname": host,
                "subnet": _get('subnet'), "location": _get('location'),
                "vendor": "unknown",
            })
    return out


def _inventory_rows(source, wanted, key_field, builder):
    """공통 인벤토리 파서: wanted 컬럼을 헤더에서 자동 매핑, key_field 값이 유효한 행만.

    key_field='ip'면 IP 검증, 그 외엔 비어있지 않으면 채택. builder(_get)로 dict 생성.
    """
    field_alias = {f: ALIASES[f] for f in wanted if f in ALIASES}
    out, seen = [], set()
    wb = openpyxl_load(source, read_only=True, data_only=True)
    for ws in wb.worksheets:
        colmap = None
        for row in ws.iter_rows(values_only=True):
            if colmap is None:
                m = {}
                for ci, cell in enumerate(row):
                    n = _norm(cell)
                    for f, al in field_alias.items():
                        if n in al and f not in m:
                            m[f] = ci
                if key_field in m:
                    colmap = m
                continue

            def _get(field):
                ci = colmap.get(field)
                if ci is None or ci >= len(row):
                    return ""
                v = row[ci]
                return str(v).strip() if v is not None else ""

            key = _get(key_field)
            if not key:
                continue
            if key_field == 'ip' and not _is_valid_ip(key):
                continue
            if key in seen:
                continue
            seen.add(key)
            item = builder(_get)
            if item:
                out.append(item)
    return out


def _norm_os_type(raw):
    """엑셀 OS 문자열 → (os_type, os_info).

    os_type은 수집 분기에 쓰이므로 계열로 정규화하고(windows/linux/aix/solaris/
    hpux/esxi/bsd/macos), 원문('Ubuntu 22.04' 등)은 표시용 os_info로 보존한다.
    linux·windows가 아닌 OS도 그대로 인식해 범용 UNIX 경로로 수집된다.
    """
    s = (raw or "").strip()
    if not s:
        return "auto", ""
    low = s.lower()
    if any(k in low for k in ("windows", "win2", "win 2", "msserver", "microsoft")):
        return "windows", s
    if any(k in low for k in ("esxi", "vmware", "vsphere", "vmkernel")):
        return "esxi", s
    if "aix" in low:
        return "aix", s
    if any(k in low for k in ("solaris", "sunos")):
        return "solaris", s
    if any(k in low for k in ("hp-ux", "hpux")):
        return "hpux", s
    if any(k in low for k in ("mac os", "macos", "darwin", "osx")):
        return "macos", s
    if "bsd" in low:
        return "bsd", s
    if any(k in low for k in ("linux", "ubuntu", "centos", "redhat", "red hat", "rhel",
                              "debian", "rocky", "alma", "suse", "oracle linux", "amazon linux")):
        return "linux", s
    if "unix" in low:
        return "unix", s
    return "auto", s


def parse_server_inventory(source) -> List[Dict[str, Any]]:
    """이름/IP(필수) + 위치/OS(선택) 엑셀 → 서버 등록 행 목록.

    Returns: [{name, ip, location, os_type, os_info}] — 계정·상세는 이후 '수집'에서.
    """
    def _b(_get):
        ip = _get('ip')
        os_type, os_info = _norm_os_type(_get('os_type'))
        return {
            "name": _get('name') or _get('hostname') or ip,
            "ip": ip,
            "hostname": _get('hostname'),
            "location": _get('location'),
            "os_type": os_type,
            "os_info": os_info,
        }
    return _inventory_rows(source, ('name', 'ip', 'hostname', 'location', 'os_type'),
                           'ip', _b)


def _to_num(v):
    """엑셀 셀 → float. '32 GB'·'1,024'처럼 단위·구분자가 섞여도 숫자만 뽑는다."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^0-9.\-]", "", str(v))
    if not s or s in ("-", "."):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_server_specs(source) -> List[Dict[str, Any]]:
    """IP + CPU·메모리·디스크 사양 엑셀 → 서버 사양 갱신 행 목록.

    실측 수집(SSH·WMI·SNMP)이 막힌 서버가 있어, 이미 파악해 둔 사양을 엑셀로
    한 번에 채우기 위한 경로다. **IP만 매칭용**이고 새 서버를 만들지 않는다
    (호출부가 기존 서버와만 매칭해야 한다) — 이름·위치는 여기서 건드리지 않는다.

    Returns: [{ip, cpu_model, cpu_cores, mem_total_mb, disk_total_gb, disk_used_gb}]
    값이 비었거나 해석 안 되는 칸은 아예 키에서 빠진다(= 그 필드는 안 건드림).
    메모리는 GB 칸으로 들어와도 MB로 환산해 mem_total_mb 하나로 합친다.
    """
    def _b(_get):
        ip = _get('ip')
        out = {"ip": ip}
        model = _get('cpu_model')
        if model:
            out["cpu_model"] = model[:150]
        cores = _to_num(_get('cpu_cores'))
        if cores and cores > 0:
            out["cpu_cores"] = int(cores)
        mem_mb = _to_num(_get('mem_total_mb'))
        if not mem_mb:
            mem_gb = _to_num(_get('mem_total_gb'))
            if mem_gb and mem_gb > 0:
                mem_mb = mem_gb * 1024
        if mem_mb and mem_mb > 0:
            out["mem_total_mb"] = int(round(mem_mb))
        disk_total = _to_num(_get('disk_total_gb'))
        if disk_total and disk_total > 0:
            out["disk_total_gb"] = round(disk_total, 1)
        disk_used = _to_num(_get('disk_used_gb'))
        if disk_used and disk_used >= 0:
            out["disk_used_gb"] = round(disk_used, 1)
        # 사양 칸이 전부 비어 있으면 반영할 게 없다 — IP만 있는 행은 버린다.
        if len(out) <= 1:
            return None
        return out
    return _inventory_rows(
        source,
        ('ip', 'cpu_model', 'cpu_cores', 'mem_total_mb', 'mem_total_gb',
         'disk_total_gb', 'disk_used_gb'),
        'ip', _b)


def parse_firewall_inventory(source) -> List[Dict[str, Any]]:
    """이름/벤더/호스트IP(필수) + 포트/위치(선택) 엑셀 → 방화벽 등록 행 목록.

    Returns: [{name, vendor, host, port, location}].
    벤더는 fortigate/paloalto로 정규화(미상이면 fortigate 기본).
    """
    def _b(_get):
        host = _get('ip')
        v = _norm(_get('vendor'))
        if 'palo' in v or 'pan' in v:
            vendor = "paloalto"
        elif 'forti' in v or 'fgt' in v:
            vendor = "fortigate"
        else:
            vendor = "fortigate"
        port = _get('port')
        try:
            port = int(port) if port else None
        except ValueError:
            port = None
        return {
            "name": _get('name') or host,
            "vendor": vendor,
            "host": host,
            "port": port,
            "location": _get('location'),
        }
    return _inventory_rows(source, ('name', 'vendor', 'ip', 'port', 'location'),
                           'ip', _b)


def _looks_like_header(row: List[Any]) -> bool:
    """
    헤더 후보인가?
    알려진 필드의 별칭과 MIN_MATCHED_COLS 개 이상 매칭하면 헤더로 간주.
    """
    if not row:
        return False

    normalized_row = [_norm(cell) for cell in row]
    matched_fields = set()

    for norm_cell in normalized_row:
        if not norm_cell:  # 빈 셀 스킵
            continue
        for known_field, aliases in ALIASES.items():
            if norm_cell in aliases:
                matched_fields.add(known_field)
                break

    return len(matched_fields) >= MIN_MATCHED_COLS


def _extract_blocks(rows: List[List[Any]]) -> List[List[List[Any]]]:
    """
    행 리스트를 멀티블록으로 분리.
    새 헤더 발견 시 이전 블록 확정.

    반환:
      blocks: 각 블록은 [헤더_행, 데이터_행1, 데이터_행2, ...]
    """
    blocks = []
    current_block = []
    current_header = None

    for row in rows:
        if _looks_like_header(row):
            if current_header is not None:
                # 새 헤더 발견 → 이전 블록 확정
                if current_block:
                    blocks.append(current_block)
                current_block = [row]
            else:
                # 첫 헤더 발견
                current_block = [row]
            current_header = row
        else:
            # 첫 헤더 이후 데이터만 추가
            if current_header is not None:
                current_block.append(row)

    # 마지막 블록 확정
    if current_block:
        blocks.append(current_block)

    return blocks


def _get_header_map(header_row: List[Any]) -> Dict[str, int]:
    """
    헤더 행을 정규화하고 별칭 매핑해서 {canonical_field: column_index} 딕셔너리 반환.

    예:
      header_row = ['Switch Name', 'IPv4 Address', 'Vendor']
      반환: {'name': 0, 'ip': 1, 'vendor': 2}
    """
    header_map = {}
    for idx, cell in enumerate(header_row):
        norm_cell = _norm(cell)
        for canonical, aliases in ALIASES.items():
            if norm_cell in aliases:
                header_map[canonical] = idx
                break
    return header_map


def _block_to_records(block: List[List[Any]]) -> Tuple[Optional[str], List[Dict[str, Any]]]:
    """
    블록(헤더+데이터)을 레코드로 변환.

    입력:
      block: [헤더_행, 데이터_행1, ...]

    출력:
      (block_type, records)
      block_type: 'switch', 'host', 또는 None (폐기)
      records: [{field: value}, ...]

    로직:
      1. 헤더 파싱
      2. 데이터 행만 필터링
      3. IP 정규식 매칭된 행만 추출
      4. IP 매칭 = 0 → 폐기 (None)
      5. 장비 식별 컬럼(vendor/name/hostname/subnet/location) 있음 → 'switch'
      6. 장부 전용 컬럼(mac/연결스위치/연결포트)만 있음 → 'host'
      7. IP만 있음 → 'switch' (현황판에 등록. 장부 대조 기능 제거됨)
    """
    if not block or len(block) < 1:
        return None, []

    header_row = block[0]
    data_rows = block[1:]

    header_map = _get_header_map(header_row)

    if not header_map:
        return None, []

    # 데이터를 딕셔너리로 변환
    records = []
    for row in data_rows:
        record = {}
        for canonical, idx in header_map.items():
            value = row[idx] if idx < len(row) else None
            if isinstance(value, str):
                # BUGFIX: _norm(소문자+공백제거)은 매칭용 컬럼(ip/subnet)에만.
                # 표시용(name/hostname/location 등)에 적용하면 'Seoul DC A03'→
                # 'seouldca03'처럼 표기가 훼손된다 → 표시용은 strip만.
                record[canonical] = (value.strip() if canonical in _DISPLAY_COLS
                                     else _norm(value))
            else:
                record[canonical] = value

        # IP 검증: 유효한 IP만 포함
        if 'ip' in record and _is_valid_ip(record['ip']):
            records.append(record)

    # IP 매칭 = 0 → 폐기
    if not records:
        return None, []

    # 분류(우선순위):
    #  1) 장부 전용 컬럼(연결스위치/연결포트) → host
    #  2) 장비 식별 컬럼(vendor/name/subnet/location) → switch
    #  3) mac만 있는 엔드포인트 목록 → host
    #  4) 그 외(ip/hostname만) → switch  ← 현황판에 등록(장부 대조 기능 제거됨)
    # (예전엔 'vendor'가 있어야만 스위치라, vendor 열 없는 장비 목록이 전부 host로 분류돼
    #  현황판에 안 나오던 문제 수정. hostname은 장부(사용서버명)와 공용이라 스위치 신호에서 제외.)
    hdr = set(header_map)
    if hdr & {'ledger_switch', 'ledger_port'}:
        block_type = 'host'
    elif hdr & {'vendor', 'name', 'subnet', 'location'}:
        block_type = 'switch'
    elif 'mac' in hdr:
        block_type = 'host'
    else:
        block_type = 'switch'

    return block_type, records


def load_workbook(
    file_path: str,
    read_only: bool = True,
    data_only: bool = True,
) -> Dict[str, Any]:
    """
    엑셀 파일을 로드하고 멀티블록으로 분리해서 DB 임포트 가능한 포맷으로 변환.

    입력:
      file_path: 엑셀 파일 경로 (.xlsx)
      read_only: openpyxl read_only 모드
      data_only: openpyxl data_only 모드

    출력:
      {
        "switches": [{name, ip, hostname, vendor, location}, ...],
        "hosts": [{ip, mac}, ...],
        "diagnostics": {
          "total_blocks": int,
          "discarded_blocks": int,
          "switch_blocks": int,
          "host_blocks": int,
          "imported_switches": int,
          "imported_hosts": int,
          "warnings": [str]
        }
      }

    예외:
      - FileNotFoundError: 파일 없음
      - openpyxl.utils.exceptions: 엑셀 형식 오류
    """
    try:
        wb = openpyxl_load(file_path, read_only=read_only, data_only=data_only)
    except Exception as e:
        logger.error(f"Failed to load workbook {file_path}: {e}")
        raise

    try:
        ws = wb.active
        if not ws:
            logger.warning(f"No active sheet in {file_path}")
            return {
                "switches": [],
                "hosts": [],
                "diagnostics": {
                    "total_blocks": 0,
                    "discarded_blocks": 0,
                    "switch_blocks": 0,
                    "host_blocks": 0,
                    "imported_switches": 0,
                    "imported_hosts": 0,
                    "warnings": ["No active sheet found"],
                }
            }

        # 행 추출 (None 값 포함, 나중에 필터링)
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row) if row else [])
    finally:
        wb.close()

    if not rows:
        return {
            "switches": [],
            "hosts": [],
            "diagnostics": {
                "total_blocks": 0,
                "discarded_blocks": 0,
                "switch_blocks": 0,
                "host_blocks": 0,
                "imported_switches": 0,
                "imported_hosts": 0,
                "warnings": ["No rows found"],
            }
        }

    # 멀티블록 분리
    blocks = _extract_blocks(rows)

    switches = []
    hosts = []
    warnings = []
    switch_block_count = 0
    host_block_count = 0
    discarded_block_count = 0

    for block_idx, block in enumerate(blocks):
        block_type, records = _block_to_records(block)

        if block_type is None:
            discarded_block_count += 1
            warnings.append(f"Block #{block_idx + 1} discarded (no IP matches)")
            continue

        if block_type == 'switch':
            switch_block_count += 1
            switches.extend(records)
        elif block_type == 'host':
            host_block_count += 1
            hosts.extend(records)

    return {
        "switches": switches,
        "hosts": hosts,
        "diagnostics": {
            "total_blocks": len(blocks),
            "discarded_blocks": discarded_block_count,
            "switch_blocks": switch_block_count,
            "host_blocks": host_block_count,
            "imported_switches": len(switches),
            "imported_hosts": len(hosts),
            "warnings": warnings,
        }
    }
